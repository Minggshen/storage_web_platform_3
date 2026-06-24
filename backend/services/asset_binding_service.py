
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from fastapi import UploadFile
from openpyxl import load_workbook

from models.project_model import (
    AssetRef,
    AssetValidationMessage,
    AssetValidationReport,
    DeviceRecord,
    ValidationStatus,
)
from services.project_model_service import ProjectModelService


@dataclass
class ParsedTable:
    headers: List[str]
    rows: List[Dict[str, Any]]


DEVICE_LIBRARY_SCHEMA_VERSION = "device_library_v2"
RUNTIME_FILE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
TARIFF_FILE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
DEVICE_LIBRARY_FILE_SUFFIXES = {".xlsx", ".xlsm"}
V2_DEVICE_LIBRARY_COLUMNS = [
    "enabled",
    "manufacturer",
    "device_model",
    "rated_power_kw",
    "rated_energy_kwh",
    "duration_hour",
    "battery_chemistry",
    "cooling_class",
    "cooling_note",
    "ip_system",
    "ip_pack",
    "ip_pcs",
    "corrosion_grade",
    "corrosion_optional_grade",
    "manual_safety_grade",
    "round_trip_efficiency",
    "c_rate_charge_max",
    "c_rate_discharge_max",
    "energy_unit_price_yuan_per_kwh",
    "power_related_capex_yuan_per_kw",
    "annual_om_ratio",
    "soc_min",
    "soc_max",
    "operating_temp_min_c",
    "operating_temp_max_c",
    "cycle_life",
    "fire_detection_class",
    "fire_suppression_class",
    "explosion_protection_class",
    "propagation_protection_class",
    "ems_model",
    "certification_tokens",
    "weight_kg",
    "dimensions_mm",
    "is_default_candidate",
    "ems_package_name",
]


class AssetBindingService:
    def __init__(self, project_service: ProjectModelService | None = None) -> None:
        self.project_service = project_service or ProjectModelService()

    def upload_runtime_files(
        self,
        project_id: str,
        node_id: str,
        year_map_file: UploadFile,
        model_library_file: UploadFile,
    ) -> tuple[AssetRef, AssetValidationReport, AssetRef, AssetValidationReport, Any, Path]:
        self.project_service.ensure_load_node(project_id, node_id)
        self._ensure_upload_suffix(year_map_file, RUNTIME_FILE_SUFFIXES, "年度模型映射文件")
        self._ensure_upload_suffix(model_library_file, RUNTIME_FILE_SUFFIXES, "典型日模型库文件")

        year_map_asset, year_map_path, _, _ = self.project_service.save_asset_upload(
            project_id=project_id,
            upload_file=year_map_file,
            category="runtime",
            subfolder=node_id,
            metadata={"runtime_kind": "year_map", "node_id": node_id},
        )
        try:
            year_map_report = self.validate_runtime_year_map(year_map_path)
            year_map_asset.metadata["validation"] = year_map_report.model_dump(mode="json")
        except Exception as exc:
            self._remove_staged_file(year_map_path)
            raise ValueError(f"年度模型映射文件无法读取或校验：{exc}") from exc

        model_asset, model_path, _, _ = self.project_service.save_asset_upload(
            project_id=project_id,
            upload_file=model_library_file,
            category="runtime",
            subfolder=node_id,
            metadata={"runtime_kind": "model_library", "node_id": node_id},
        )
        try:
            model_report = self.validate_runtime_model_library(model_path)
            model_asset.metadata["validation"] = model_report.model_dump(mode="json")
        except Exception as exc:
            self._remove_staged_file(year_map_path)
            self._remove_staged_file(model_path)
            raise ValueError(f"典型日模型库文件无法读取或校验：{exc}") from exc
        if not year_map_report.ok or not model_report.ok:
            self._remove_staged_file(year_map_path)
            self._remove_staged_file(model_path)
            raise ValueError(self._validation_failure_message([year_map_report, model_report], "负荷 runtime 文件校验未通过"))

        project, project_file = self.project_service.bind_runtime_assets(
            project_id=project_id,
            node_id=node_id,
            year_map_asset=year_map_asset,
            model_library_asset=model_asset,
        )
        project.assets[year_map_asset.file_id] = year_map_asset
        project.assets[model_asset.file_id] = model_asset
        _, project_file = self.project_service.save_project(project)
        return year_map_asset, year_map_report, model_asset, model_report, project, project_file

    def upload_tariff_file(
        self,
        project_id: str,
        tariff_file: UploadFile,
    ) -> tuple[AssetRef, AssetValidationReport, Any, Path]:
        self._ensure_upload_suffix(tariff_file, TARIFF_FILE_SUFFIXES, "电价表")
        asset, target_path, _, _ = self.project_service.save_asset_upload(
            project_id=project_id,
            upload_file=tariff_file,
            category="tariff",
            metadata={"asset_kind": "tariff_annual"},
        )
        try:
            report = self.validate_tariff_file(target_path)
            asset.metadata["validation"] = report.model_dump(mode="json")
        except Exception as exc:
            self._remove_staged_file(target_path)
            raise ValueError(f"电价表无法读取或校验：{exc}") from exc
        if not report.ok:
            self._remove_staged_file(target_path)
            raise ValueError(self._validation_failure_message([report], "电价表校验未通过"))
        detected_year = report.parsed_preview.get("detected_year")
        project, project_file = self.project_service.bind_tariff_asset(
            project_id=project_id,
            asset=asset,
            tariff_year=int(detected_year) if detected_year is not None else None,
        )
        project.assets[asset.file_id] = asset
        _, project_file = self.project_service.save_project(project)
        return asset, report, project, project_file

    def upload_device_library_file(
        self,
        project_id: str,
        device_file: UploadFile,
    ) -> tuple[AssetRef, AssetValidationReport, List[DeviceRecord], Any, Path]:
        self._ensure_upload_suffix(device_file, DEVICE_LIBRARY_FILE_SUFFIXES, "设备策略库")
        asset, target_path, _, _ = self.project_service.save_asset_upload(
            project_id=project_id,
            upload_file=device_file,
            category="device_library",
            metadata={"asset_kind": "device_library"},
        )
        try:
            report, records = self.validate_device_library_file(target_path)
            asset.metadata["validation"] = report.model_dump(mode="json")
        except Exception as exc:
            self._remove_staged_file(target_path)
            raise ValueError(f"设备策略库无法读取或校验：{exc}") from exc
        if not report.ok:
            self._remove_staged_file(target_path)
            raise ValueError(self._validation_failure_message([report], "设备策略库校验未通过"))
        project, project_file = self.project_service.replace_device_library(
            project_id=project_id,
            asset=asset,
            records=records,
        )
        project.assets[asset.file_id] = asset
        _, project_file = self.project_service.save_project(project)
        return asset, report, records, project, project_file

    @staticmethod
    def _ensure_upload_suffix(file: UploadFile, allowed_suffixes: set[str], label: str) -> None:
        suffix = Path(file.filename or "").suffix.lower()
        if suffix not in allowed_suffixes:
            allowed = "、".join(sorted(allowed_suffixes))
            raise ValueError(f"{label}文件格式不支持：{file.filename or '未命名文件'}。请上传 {allowed} 格式。")

    @staticmethod
    def _validation_failure_message(reports: Sequence[AssetValidationReport], prefix: str) -> str:
        details: list[str] = []
        for report in reports:
            for message in report.messages:
                if message.status == ValidationStatus.ERROR:
                    detail = f"{message.title}：{message.message}"
                    if message.detail:
                        detail = f"{detail}（{message.detail}）"
                    details.append(detail)
        if not details:
            return prefix
        return f"{prefix}；" + "；".join(details[:5])

    @staticmethod
    def _remove_staged_file(path: Path) -> None:
        try:
            if path.exists() and path.is_file():
                path.unlink()
        except OSError:
            pass

    def validate_runtime_year_map(self, file_path: str | Path) -> AssetValidationReport:
        table = self._read_table(file_path)
        report = self._new_report("runtime_year_map")
        header_map = self._normalize_header_map(table.headers)
        required = ["date", "internal_model_id"]
        missing = [field for field in required if field not in header_map]
        if missing:
            for field in missing:
                self._push_message(report, ValidationStatus.ERROR, "字段完整性", f"缺少字段：{field}")
            self._finalize_report(report)
            return report
        self._push_message(report, ValidationStatus.PASS, "字段完整性", "核心字段齐全")

        dates: List[date] = []
        model_ids: List[str] = []
        for idx, row in enumerate(table.rows, start=2):
            dt = self._parse_date(row.get(header_map["date"]))
            if dt is None:
                self._push_message(report, ValidationStatus.ERROR, "日期解析", "存在无法解析的日期", detail=f"第 {idx} 行")
            else:
                dates.append(dt)
            model_id = str(row.get(header_map["internal_model_id"], "")).strip()
            if not model_id:
                self._push_message(report, ValidationStatus.ERROR, "模型ID", "存在空 internal_model_id", detail=f"第 {idx} 行")
            else:
                model_ids.append(model_id)

        if len(table.rows) == 365:
            self._push_message(report, ValidationStatus.PASS, "天数检查", "当前 365 天")
        else:
            level = ValidationStatus.WARNING if len(table.rows) > 0 else ValidationStatus.ERROR
            self._push_message(report, level, "天数检查", f"当前 {len(table.rows)} 天，建议为 365 天")

        if dates:
            if len(set(dates)) == len(dates):
                self._push_message(report, ValidationStatus.PASS, "日期去重", "日期无重复")
            else:
                self._push_message(report, ValidationStatus.ERROR, "日期去重", "存在重复日期")
            if self._is_contiguous(sorted(dates)):
                self._push_message(report, ValidationStatus.PASS, "日期连续", "日期按天连续")
            else:
                self._push_message(report, ValidationStatus.WARNING, "日期连续", "日期未完全连续")

        report.parsed_preview = {
            "row_count": len(table.rows),
            "sample_model_ids": model_ids[:5],
            "date_start": dates[0].isoformat() if dates else None,
            "date_end": dates[-1].isoformat() if dates else None,
        }
        self._finalize_report(report)
        return report

    def validate_runtime_model_library(self, file_path: str | Path) -> AssetValidationReport:
        table = self._read_table(file_path)
        report = self._new_report("runtime_model_library")
        header_map = self._normalize_header_map(table.headers)
        required = ["internal_model_id", *[f"h{i:02d}" for i in range(24)]]
        missing = [field for field in required if field not in header_map]
        if missing:
            for field in missing:
                self._push_message(report, ValidationStatus.ERROR, "字段完整性", f"缺少字段：{field}")
            self._finalize_report(report)
            return report
        self._push_message(report, ValidationStatus.PASS, "字段完整性", "模型库字段齐全")

        model_ids: List[str] = []
        for idx, row in enumerate(table.rows, start=2):
            model_id = str(row.get(header_map["internal_model_id"], "")).strip()
            if not model_id:
                self._push_message(report, ValidationStatus.ERROR, "模型ID", "存在空 internal_model_id", detail=f"第 {idx} 行")
            model_ids.append(model_id)
            for hour_key in [f"h{i:02d}" for i in range(24)]:
                value = row.get(header_map[hour_key])
                if self._coerce_float(value) is None:
                    self._push_message(report, ValidationStatus.ERROR, "24点曲线", f"{hour_key} 无法解析为数值", detail=f"第 {idx} 行")
                    break

        if len(set(filter(None, model_ids))) == len(list(filter(None, model_ids))):
            self._push_message(report, ValidationStatus.PASS, "模型ID唯一性", "模型 ID 无重复")
        else:
            self._push_message(report, ValidationStatus.ERROR, "模型ID唯一性", "存在重复的 internal_model_id")

        report.parsed_preview = {"row_count": len(table.rows), "sample_model_ids": model_ids[:5]}
        self._finalize_report(report)
        return report

    def validate_tariff_file(self, file_path: str | Path) -> AssetValidationReport:
        table = self._read_table(file_path)
        report = self._new_report("tariff_annual")
        header_map = self._normalize_header_map(table.headers)
        required = ["date", *[f"电价_{i:02d}" for i in range(24)]]
        missing = [field for field in required if field not in header_map]
        if missing:
            for field in missing:
                self._push_message(report, ValidationStatus.ERROR, "字段完整性", f"缺少字段：{field}")
            self._finalize_report(report)
            return report
        self._push_message(report, ValidationStatus.PASS, "字段完整性", "电价表字段完整")

        dates: List[date] = []
        detected_year: Optional[int] = None
        for idx, row in enumerate(table.rows, start=2):
            dt = self._parse_date(row.get(header_map["date"]))
            if dt is None:
                self._push_message(report, ValidationStatus.ERROR, "日期解析", "存在无法解析的日期", detail=f"第 {idx} 行")
            else:
                dates.append(dt)
                if detected_year is None:
                    detected_year = dt.year
            for hour_key in [f"电价_{i:02d}" for i in range(24)]:
                value = self._coerce_float(row.get(header_map[hour_key]))
                if value is None:
                    self._push_message(report, ValidationStatus.ERROR, "电价值解析", f"{hour_key} 无法解析为数值", detail=f"第 {idx} 行")
                    break

        if len(table.rows) == 365:
            self._push_message(report, ValidationStatus.PASS, "天数检查", "当前 365 天")
        else:
            level = ValidationStatus.WARNING if len(table.rows) > 0 else ValidationStatus.ERROR
            self._push_message(report, level, "天数检查", f"当前 {len(table.rows)} 天，建议为 365 天")

        if dates:
            if len(set(dates)) == len(dates):
                self._push_message(report, ValidationStatus.PASS, "日期去重", "日期无重复")
            else:
                self._push_message(report, ValidationStatus.ERROR, "日期去重", "存在重复日期")
            if self._is_contiguous(sorted(dates)):
                self._push_message(report, ValidationStatus.PASS, "日期连续", "日期逐天连续")
            else:
                self._push_message(report, ValidationStatus.WARNING, "日期连续", "日期未完全连续")

        report.parsed_preview = {
            "row_count": len(table.rows),
            "detected_year": detected_year,
            "date_start": dates[0].isoformat() if dates else None,
            "date_end": dates[-1].isoformat() if dates else None,
        }
        self._finalize_report(report)
        return report

    def validate_device_library_file(self, file_path: str | Path) -> tuple[AssetValidationReport, List[DeviceRecord]]:
        report = self._new_report("device_library")
        schema_version = self._read_device_library_schema_version(file_path)
        if schema_version != DEVICE_LIBRARY_SCHEMA_VERSION:
            self._push_message(
                report,
                ValidationStatus.ERROR,
                "模板版本",
                f"设备策略库必须使用 v2 模板，元数据 schema_version 应为 {DEVICE_LIBRARY_SCHEMA_VERSION}",
                detail=f"当前 schema_version：{schema_version or '未找到'}",
            )
            self._finalize_report(report)
            return report, []
        self._push_message(report, ValidationStatus.PASS, "模板版本", "已识别 device_library_v2 模板")

        table = self._read_device_library_table(file_path)
        header_map = self._normalize_header_map(table.headers)
        missing = [field for field in V2_DEVICE_LIBRARY_COLUMNS if field not in table.headers]
        if missing:
            for field in missing:
                self._push_message(report, ValidationStatus.ERROR, "字段完整性", f"缺少字段：{field}")
            self._finalize_report(report)
            return report, []
        self._push_message(report, ValidationStatus.PASS, "字段完整性", "设备主字段完整")

        records: List[DeviceRecord] = []
        enabled_count = 0
        for idx, row in enumerate(table.rows, start=2):
            if self._looks_like_device_description_row(row, header_map):
                continue
            vendor = str(row.get(header_map["vendor"], "")).strip()
            model = str(row.get(header_map["model"], "")).strip()
            if not vendor or not model:
                self._push_message(report, ValidationStatus.ERROR, "记录合法性", "存在空厂家或型号", detail=f"第 {idx} 行")
                continue

            duration = self._value_by_alias(row, header_map, ["duration_hour"])
            rated_power = self._value_by_alias(row, header_map, ["rated_power_kw"])
            rated_energy = self._value_by_alias(row, header_map, ["rated_energy_kwh"])
            if duration is None and rated_power and rated_energy and rated_power > 0:
                duration = round(rated_energy / rated_power, 4)

            energy_unit_price = self._value_by_alias(
                row,
                header_map,
                ["energy_unit_price_yuan_per_kwh"],
            )
            price_yuan_per_wh = round(float(energy_unit_price) / 1000.0, 6) if energy_unit_price is not None else None

            safety_level = self._string_by_alias(row, header_map, ["manual_safety_grade"])
            ems_package = self._string_by_alias(row, header_map, ["ems_package_name"])

            core_keys = {
                "enabled",
                "vendor",
                "model",
                "manufacturer",
                "device_model",
                "is_default_candidate",
                "ems_package_name",
                "battery_chemistry",
                "rated_power_kw",
                "rated_energy_kwh",
                "duration_hour",
                "cooling_class",
                "manual_safety_grade",
                "round_trip_efficiency",
                "ip_system",
                "corrosion_grade",
                "weight_kg",
                "energy_unit_price_yuan_per_kwh",
                "power_related_capex_yuan_per_kw",
                "cycle_life",
                "soc_min",
                "soc_max",
            }
            extra: Dict[str, Any] = {}
            for norm_key, original in header_map.items():
                if norm_key in core_keys:
                    continue
                extra[norm_key] = row.get(original)

            record = DeviceRecord(
                enabled=self._bool_by_alias(row, header_map, ["enabled", "启用"], default=True),
                vendor=vendor,
                model=model,
                series_name=self._string_by_alias(row, header_map, ["series_name"]),
                device_family=self._string_by_alias(row, header_map, ["device_family"]),
                system_topology_type=self._string_by_alias(row, header_map, ["system_topology_type"]),
                application_scene=self._string_by_alias(row, header_map, ["application_scene"]),
                cni_fit_level=self._string_by_alias(row, header_map, ["cni_fit_level"]),
                is_default_candidate=self._bool_by_alias(row, header_map, ["is_default_candidate"], default=False) if "is_default_candidate" in header_map else None,
                ems_package=ems_package,
                ems_package_name=self._string_by_alias(row, header_map, ["ems_package_name"]),
                has_builtin_ems=self._bool_by_alias(row, header_map, ["has_builtin_ems"], default=False) if "has_builtin_ems" in header_map else None,
                requires_external_pcs=self._bool_by_alias(row, header_map, ["requires_external_pcs"], default=False) if "requires_external_pcs" in header_map else None,
                supports_black_start=self._bool_by_alias(row, header_map, ["supports_black_start"], default=False) if "supports_black_start" in header_map else None,
                supports_offgrid_microgrid=self._bool_by_alias(row, header_map, ["supports_offgrid_microgrid"], default=False) if "supports_offgrid_microgrid" in header_map else None,
                battery_chemistry=self._string_by_alias(row, header_map, ["battery_chemistry"]),
                rated_power_kw=rated_power,
                rated_energy_kwh=rated_energy,
                usable_energy_kwh_at_fat=self._value_by_alias(row, header_map, ["usable_energy_kwh_at_fat"]),
                duration_hour=duration,
                dc_voltage_range_v=self._string_by_alias(row, header_map, ["dc_voltage_range_v"]),
                ac_grid_voltage_v=self._string_by_alias(row, header_map, ["ac_grid_voltage_v"]),
                battery_config=self._string_by_alias(row, header_map, ["battery_config"]),
                cooling_type=self._string_by_alias(row, header_map, ["cooling_class"]),
                fire_detection=self._string_by_alias(row, header_map, ["fire_detection_class"]),
                fire_suppression=self._string_by_alias(row, header_map, ["fire_suppression_class"]),
                backup_system=self._string_by_alias(row, header_map, ["backup_system"]),
                accident_ventilation=self._bool_by_alias(row, header_map, ["accident_ventilation"], default=False) if "accident_ventilation" in header_map else None,
                pack_level_firefighting_optional=self._bool_by_alias(row, header_map, ["pack_level_firefighting_optional"], default=False) if "pack_level_firefighting_optional" in header_map else None,
                explosion_relief_optional=self._bool_by_alias(row, header_map, ["explosion_relief_optional"], default=False) if "explosion_relief_optional" in header_map else None,
                msd_required=self._bool_by_alias(row, header_map, ["msd_required"], default=False) if "msd_required" in header_map else None,
                communication_protocol=self._string_by_alias(row, header_map, ["communication_protocol"]),
                safety_level=safety_level,
                manual_safety_grade=self._string_by_alias(row, header_map, ["manual_safety_grade"]),
                manual_safety_notes=self._string_by_alias(row, header_map, ["manual_safety_notes"]),
                cycle_life=self._int_by_alias(row, header_map, ["cycle_life"]),
                soc_min=self._value_by_alias(row, header_map, ["soc_min"]),
                soc_max=self._value_by_alias(row, header_map, ["soc_max"]),
                efficiency_pct=self._value_by_alias(row, header_map, ["round_trip_efficiency"]),
                ip_system=self._string_by_alias(row, header_map, ["ip_system"]),
                corrosion_grade=self._string_by_alias(row, header_map, ["corrosion_grade"]),
                install_mode=self._string_by_alias(row, header_map, ["install_mode"]),
                aux_power_interface=self._string_by_alias(row, header_map, ["aux_power_interface"]),
                dimension_w_mm=self._value_by_alias(row, header_map, ["dimension_w_mm"]),
                dimension_d_mm=self._value_by_alias(row, header_map, ["dimension_d_mm"]),
                dimension_h_mm=self._value_by_alias(row, header_map, ["dimension_h_mm"]),
                weight_kg=self._value_by_alias(row, header_map, ["weight_kg"]),
                price_yuan_per_wh=price_yuan_per_wh,
                energy_unit_price_yuan_per_kwh=energy_unit_price,
                power_related_capex_yuan_per_kw=self._value_by_alias(row, header_map, ["power_related_capex_yuan_per_kw"]),
                station_integration_capex_ratio=self._value_by_alias(row, header_map, ["station_integration_capex_ratio"]),
                fire_protection_capex_ratio=self._value_by_alias(row, header_map, ["fire_protection_capex_ratio"]),
                annual_insurance_rate_on_capex=self._value_by_alias(row, header_map, ["annual_insurance_rate_on_capex"]),
                annual_safety_maintenance_rate_on_capex=self._value_by_alias(row, header_map, ["annual_safety_maintenance_rate_on_capex"]),
                annual_fire_system_inspection_rate_on_capex=self._value_by_alias(row, header_map, ["annual_fire_system_inspection_rate_on_capex"]),
                price_status=self._string_by_alias(row, header_map, ["price_status"]),
                quote_source=self._string_by_alias(row, header_map, ["quote_source"]),
                source_files=self._string_by_alias(row, header_map, ["source_files"]),
                extra=extra,
            )
            record = self._complete_device_record(record)
            if record.enabled:
                enabled_count += 1
            records.append(record)

        if records:
            self._push_message(report, ValidationStatus.PASS, "记录数量", f"当前导入 {len(records)} 条设备记录")
        else:
            self._push_message(report, ValidationStatus.ERROR, "记录数量", "未导入到任何设备记录")

        if enabled_count > 0:
            self._push_message(report, ValidationStatus.PASS, "启用记录", f"当前启用设备数：{enabled_count}")
        else:
            self._push_message(report, ValidationStatus.WARNING, "启用记录", "当前没有启用设备")

        enabled_missing_price = sum(1 for r in records if r.enabled and r.energy_unit_price_yuan_per_kwh is None and r.price_yuan_per_wh is None)
        if enabled_missing_price:
            self._push_message(report, ValidationStatus.WARNING, "价格字段", f"存在 {enabled_missing_price} 条启用设备未填写价格字段")
        report.parsed_preview = {
            "record_count": len(records),
            "enabled_count": enabled_count,
            "sample_records": [f"{item.vendor}/{item.model}" for item in records[:5]],
        }
        self._finalize_report(report)
        return report, records


    def _complete_device_record(self, record: DeviceRecord) -> DeviceRecord:
        """Fill conservative defaults so the generated engineering model is closer to the original optimization inputs."""
        if record.usable_energy_kwh_at_fat is None and record.rated_energy_kwh is not None:
            record.usable_energy_kwh_at_fat = record.rated_energy_kwh

        if record.price_yuan_per_wh is None and record.energy_unit_price_yuan_per_kwh is not None:
            record.price_yuan_per_wh = round(float(record.energy_unit_price_yuan_per_kwh) / 1000.0, 6)
        if record.energy_unit_price_yuan_per_kwh is None and record.price_yuan_per_wh is not None:
            record.energy_unit_price_yuan_per_kwh = round(float(record.price_yuan_per_wh) * 1000.0, 3)

        if record.duration_hour is None and record.rated_power_kw not in (None, 0) and record.rated_energy_kwh is not None:
            record.duration_hour = round(float(record.rated_energy_kwh) / float(record.rated_power_kw), 4)
        if record.rated_power_kw is None and record.rated_energy_kwh is not None and record.duration_hour not in (None, 0):
            record.rated_power_kw = round(float(record.rated_energy_kwh) / float(record.duration_hour), 4)
        if record.rated_energy_kwh is None and record.rated_power_kw is not None and record.duration_hour not in (None, 0):
            record.rated_energy_kwh = round(float(record.rated_power_kw) * float(record.duration_hour), 4)

        if record.efficiency_pct is not None and 0 < float(record.efficiency_pct) <= 1.5:
            record.efficiency_pct = round(float(record.efficiency_pct) * 100.0, 4)

        # Conservative topology-based fallback when only energy is known.
        if record.rated_power_kw is None and record.rated_energy_kwh is not None:
            inferred_duration = self._infer_default_duration(record)
            if inferred_duration is not None and inferred_duration > 0:
                record.duration_hour = record.duration_hour or inferred_duration
                record.rated_power_kw = round(float(record.rated_energy_kwh) / inferred_duration, 4)
                record.extra["inferred_rated_power_kw"] = True
                record.extra["inference_basis"] = f"duration={inferred_duration}h"

        if record.rated_energy_kwh is None and record.rated_power_kw is not None:
            inferred_duration = self._infer_default_duration(record)
            if inferred_duration is not None and inferred_duration > 0:
                record.duration_hour = record.duration_hour or inferred_duration
                record.rated_energy_kwh = round(float(record.rated_power_kw) * inferred_duration, 4)
                record.extra["inferred_rated_energy_kwh"] = True
                record.extra["inference_basis"] = f"duration={inferred_duration}h"

        if record.cycle_life is None and record.enabled:
            record.cycle_life = self._infer_default_cycle_life(record)
            record.extra["default_cycle_life_applied"] = True
        if record.soc_min is None and record.enabled and record.system_topology_type != "dc_dc_coupled_unit":
            record.soc_min = 0.10
            record.extra["default_soc_min_applied"] = True
        if record.soc_max is None and record.enabled and record.system_topology_type != "dc_dc_coupled_unit":
            record.soc_max = 0.90
            record.extra["default_soc_max_applied"] = True
        if record.efficiency_pct is None and record.enabled:
            record.efficiency_pct = self._infer_default_efficiency(record)
            record.extra["default_efficiency_pct_applied"] = True
        return record

    def _infer_default_duration(self, record: DeviceRecord) -> Optional[float]:
        topology = (record.system_topology_type or "").strip().lower()
        family = (record.device_family or "").strip().lower()
        model = (record.model or "").strip().lower()
        if "4h" in model:
            return 4.0
        if "2h" in model or "233kwh" in model:
            return 2.0
        if topology in {"ac_all_in_one", "outdoor_liquid_cooled_cabinet"}:
            return 2.0
        if topology in {"containerized_large_scale", "dc_side_battery_cabinet"}:
            return 4.0 if (record.rated_energy_kwh or 0) >= 4000 else 2.0
        if "一体机" in family or "柜式" in family:
            return 2.0
        return None

    def _infer_default_cycle_life(self, record: DeviceRecord) -> int:
        chemistry = (record.battery_chemistry or "").strip().upper()
        topology = (record.system_topology_type or "").strip().lower()
        if "LFP" in chemistry and topology in {"containerized_large_scale", "dc_side_battery_cabinet"}:
            return 10000
        if "LFP" in chemistry:
            return 8000
        return 6000

    def _infer_default_efficiency(self, record: DeviceRecord) -> Optional[float]:
        topology = (record.system_topology_type or "").strip().lower()
        if topology in {"ac_all_in_one", "outdoor_liquid_cooled_cabinet"}:
            return 88.0
        if topology in {"containerized_large_scale", "dc_side_battery_cabinet"}:
            return 90.0
        if topology == "dc_dc_coupled_unit":
            return 99.0
        return None

    def _new_report(self, asset_kind: str) -> AssetValidationReport:
        return AssetValidationReport(ok=False, asset_kind=asset_kind)

    def _push_message(self, report: AssetValidationReport, status: ValidationStatus, title: str, message: str, detail: str | None = None) -> None:
        report.messages.append(AssetValidationMessage(status=status, title=title, message=message, detail=detail))

    def _finalize_report(self, report: AssetValidationReport) -> None:
        report.pass_count = sum(1 for item in report.messages if item.status == ValidationStatus.PASS)
        report.warning_count = sum(1 for item in report.messages if item.status == ValidationStatus.WARNING)
        report.error_count = sum(1 for item in report.messages if item.status == ValidationStatus.ERROR)
        report.ok = report.error_count == 0

    def _read_table(self, file_path: str | Path) -> ParsedTable:
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._read_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx(path)
        raise ValueError(f"暂不支持的文件格式：{path.suffix}")

    def _read_device_library_table(self, file_path: str | Path) -> ParsedTable:
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx(path, preferred_sheets=("设备库",))
        raise ValueError("设备策略库必须使用 v2 模板 .xlsx/.xlsm 文件")

    def _read_csv(self, path: Path) -> ParsedTable:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            rows = [dict(row) for row in reader]
        return ParsedTable(headers=headers, rows=rows)

    def _read_xlsx(self, path: Path, preferred_sheets: Sequence[str] | None = None) -> ParsedTable:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        for sheet_name in preferred_sheets or ():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break

        values = list(ws.iter_rows(values_only=True))
        if not values:
            return ParsedTable(headers=[], rows=[])

        header_idx = self._find_device_header_row(values) if preferred_sheets else 0
        header_row = values[header_idx]
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: List[Dict[str, Any]] = []
        for row_values in values[header_idx + 1:]:
            if row_values is None:
                continue
            row = {headers[idx]: row_values[idx] if idx < len(row_values) else None for idx in range(len(headers))}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return ParsedTable(headers=headers, rows=rows)

    def _find_device_header_row(self, values: Sequence[Sequence[Any]]) -> int:
        for idx, row in enumerate(values[:25]):
            normalized = {self._normalize_header_name(cell) for cell in row if cell not in (None, "")}
            if {"vendor", "model"}.issubset(normalized):
                return idx
        return 0

    def _read_device_library_schema_version(self, file_path: str | Path) -> str:
        path = Path(file_path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return ""
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return ""
        if "元数据" not in wb.sheetnames:
            return ""
        rows = list(wb["元数据"].iter_rows(values_only=True))
        if not rows:
            return ""
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        try:
            key_idx = headers.index("key")
            value_idx = headers.index("value")
        except ValueError:
            return ""
        for row in rows[1:]:
            key = str(row[key_idx] if key_idx < len(row) else "").strip()
            if key == "schema_version":
                return str(row[value_idx] if value_idx < len(row) else "").strip()
        return ""

    def _looks_like_device_description_row(self, row: Dict[str, Any], header_map: Dict[str, str]) -> bool:
        vendor_header = header_map.get("vendor")
        model_header = header_map.get("model")
        if not vendor_header or not model_header:
            return False
        vendor = str(row.get(vendor_header, "")).strip()
        model = str(row.get(model_header, "")).strip()
        return vendor in {"厂家", "供应商", "制造商"} and model in {"型号", "设备型号", "产品型号"}

    def _normalize_header_map(self, headers: Sequence[str]) -> Dict[str, str]:
        mapping: Dict[str, str] = {}
        for original in headers:
            key = self._normalize_header_name(original)
            if key:
                mapping[key] = original
        return mapping

    def _normalize_header_name(self, header: Any) -> str:
        text = str(header or "").strip().lower()
        alias_map = {
            "日期": "date",
            "date": "date",
            "day": "date",
            "internal_model_id": "internal_model_id",
            "model_id": "internal_model_id",
            "内部模型id": "internal_model_id",
            "厂家": "vendor",
            "供应商": "vendor",
            "manufacturer": "vendor",
            "vendor": "vendor",
            "型号": "model",
            "device_model": "model",
            "model": "model",
            "启用": "enabled",
            "enabled": "enabled",
            "时长_h": "duration_hour",
            "duration_h": "duration_hour",
            "duration_hour": "duration_hour",
            "额定功率_kw": "rated_power_kw",
            "rated_power_kw": "rated_power_kw",
            "额定容量_kwh": "rated_energy_kwh",
            "rated_energy_kwh": "rated_energy_kwh",
            "可用容量_kwh": "usable_energy_kwh_at_fat",
            "usable_energy_kwh_at_fat": "usable_energy_kwh_at_fat",
            "安全等级": "safety_level",
            "safety_level": "safety_level",
            "manual_safety_grade": "manual_safety_grade",
            "ems包": "ems_package",
            "ems_package": "ems_package",
            "ems_package_name": "ems_package_name",
            "价格_元每wh": "price_yuan_per_wh",
            "价格元每wh": "price_yuan_per_wh",
            "price_yuan_per_wh": "price_yuan_per_wh",
            "energy_unit_price_yuan_per_kwh": "energy_unit_price_yuan_per_kwh",
            "cycle_life": "cycle_life",
            "循环寿命": "cycle_life",
            "设计循环寿命": "cycle_life",
            "soc下限": "soc_min",
            "soc_min": "soc_min",
            "最小soc": "soc_min",
            "最小soc比例": "soc_min",
            "soc上限": "soc_max",
            "soc_max": "soc_max",
            "最大soc": "soc_max",
            "最大soc比例": "soc_max",
        }
        if text in alias_map:
            return alias_map[text]
        if text.startswith("电价_"):
            return text.replace(" ", "")
        if text.startswith("h") and len(text) in {2, 3}:
            digits = text[1:]
            if digits.isdigit():
                return f"h{int(digits):02d}"
        return text

    def _parse_date(self, value: Any) -> Optional[date]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _is_contiguous(self, dates: Sequence[date]) -> bool:
        if not dates:
            return False
        return all(curr - prev == timedelta(days=1) for prev, curr in zip(dates, dates[1:]))

    def _coerce_float(self, value: Any) -> Optional[float]:
        if value in (None, "", "nan"):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _value_by_alias(self, row: Dict[str, Any], header_map: Dict[str, str], aliases: Sequence[str]) -> Optional[float]:
        for alias in aliases:
            if alias in header_map:
                return self._coerce_float(row.get(header_map[alias]))
        return None

    def _string_by_alias(self, row: Dict[str, Any], header_map: Dict[str, str], aliases: Sequence[str]) -> Optional[str]:
        for alias in aliases:
            if alias in header_map:
                value = row.get(header_map[alias])
                if value in (None, ""):
                    return None
                return str(value).strip()
        return None

    def _int_by_alias(self, row: Dict[str, Any], header_map: Dict[str, str], aliases: Sequence[str]) -> Optional[int]:
        value = self._value_by_alias(row, header_map, aliases)
        return int(value) if value is not None else None

    def _bool_by_alias(self, row: Dict[str, Any], header_map: Dict[str, str], aliases: Sequence[str], default: bool) -> bool:
        for alias in aliases:
            if alias in header_map:
                value = row.get(header_map[alias])
                if isinstance(value, bool):
                    return value
                text = str(value).strip().lower()
                if text in {"1", "true", "yes", "y", "启用", "是"}:
                    return True
                if text in {"0", "false", "no", "n", "停用", "否"}:
                    return False
        return default
