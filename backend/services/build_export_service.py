from __future__ import annotations

import csv
import json
import math
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from services.build_signature import (
    asset_signature,
    build_input_hash,
    build_input_signature,
    extract_topology,
    stable_hash,
    topology_hash,
)
from services.dss_builder_service import DssBuilderService
from services.search_space_inference_service import SearchSpaceInferenceService


class BuildExportService:
    """
    Build/export service used by the frontend workspace.

    Responsibilities:
    - validate whether the visual topology is buildable
    - emit build manifest
    - compile OpenDSS inputs under build/inputs/dss/visual_model
    - prepare a lightweight solver handoff view
    """

    def __init__(
        self,
        project_service: Any = None,
        validation_service: Any = None,
        data_root: str | Path | None = None,
    ) -> None:
        # compatibility with older positional misuse: BuildExportService(project_service)
        if project_service is not None and not hasattr(project_service, "load_project") and data_root is None:
            if isinstance(project_service, (str, Path)):
                data_root = project_service
                project_service = None

        backend_root = Path(__file__).resolve().parent.parent
        self.project_service = project_service
        self.validation_service = validation_service
        self.data_root = Path(data_root) if data_root is not None else backend_root / "data" / "projects"
        self.dss_builder = DssBuilderService()
        self.search_space_inference = SearchSpaceInferenceService()

    def get_build_preview(self, project_id: str) -> dict[str, Any]:
        return self.preview_build(project_id)

    def preview_build(self, project_id: str) -> dict[str, Any]:
        project = self._load_project(project_id)
        topology = self._extract_topology(project)
        topology_hash = self._topology_hash(topology)
        build_input_hash = self._build_input_hash(project)
        validation = self._validate_topology(topology, project)
        warnings = validation["warnings"]
        errors = validation["errors"]
        summary = {
            "project_id": project_id,
            "project_name": str(project.get("project_name") or project.get("name") or project_id),
            "topology_hash": topology_hash,
            "build_input_hash": build_input_hash,
            "ready_for_build": validation["ready_for_build"],
            "warnings": warnings,
            "errors": errors,
            "node_count": len(topology["nodes"]),
            "edge_count": len(topology["edges"]),
            "grid_count": validation["grid_count"],
            "transformer_count": validation["transformer_count"],
            "load_count": validation["load_count"],
            "active_edge_count": validation["active_edge_count"],
            "disconnected_count": validation["disconnected_count"],
        }
        return {
            "success": True,
            "summary": summary,
            "validation": validation,
            "preview": {
                "nodes": topology["nodes"],
                "edges": topology["edges"],
                "economic_parameters": topology.get("economic_parameters", {}),
            },
        }

    def build_project(self, project_id: str) -> dict[str, Any]:
        return self.generate_build(project_id)

    def generate_build(self, project_id: str) -> dict[str, Any]:
        project = self._load_project(project_id)
        topology = self._extract_topology(project)
        topology_hash = self._topology_hash(topology)
        build_input_hash = self._build_input_hash(project)
        preview = self.preview_build(project_id)
        if not bool(preview["summary"]["ready_for_build"]):
            errors = self._dedupe(preview["summary"].get("errors") or [])
            message = "拓扑预览校验未通过，请先修正 Errors 后再生成 Solver Workspace。"
            if errors:
                message = f"{message} {'；'.join(errors)}"
            raise ValueError(message)

        project_dir = self._project_dir(project_id)
        build_dir = project_dir / "build"
        inputs_dir = build_dir / "inputs"
        dss_dir = inputs_dir / "dss" / "visual_model"
        manifest_dir = build_dir / "manifest"

        dss_dir.mkdir(parents=True, exist_ok=True)
        manifest_dir.mkdir(parents=True, exist_ok=True)

        dss_payload = self.dss_builder.compile_topology(project_id, topology, dss_dir)
        build_gate = self._build_gate_status(preview, dss_payload)

        handoff = self._prepare_solver_handoff(project_id, build_dir, inputs_dir, dss_payload, build_gate)
        solver_workspace = self._prepare_solver_workspace(project_id, project, build_dir, dss_payload, build_gate)
        all_warnings = self._dedupe([*preview["summary"]["warnings"], *build_gate["warnings"]])
        all_errors = self._dedupe([*preview["summary"]["errors"], *build_gate["errors"]])

        manifest = {
            "success": True,
            "project_id": project_id,
            "project_name": str(project.get("project_name") or project.get("name") or project_id),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "topology_hash": topology_hash,
            "topology_hash_current": topology_hash,
            "build_input_hash": build_input_hash,
            "build_input_hash_current": build_input_hash,
            "manifest_stale": False,
            "build_dir": str(build_dir),
            "inputs_dir": str(inputs_dir),
            "ready_for_build": preview["summary"]["ready_for_build"],
            "ready_for_solver": bool(solver_workspace.get("ready_for_solver", False)),
            "build_gate": build_gate,
            "warnings": all_warnings,
            "errors": all_errors,
            "topology_summary": {
                "node_count": len(topology["nodes"]),
                "edge_count": len(topology["edges"]),
            },
            **dss_payload,
            "solver_handoff": handoff,
            "solver_workspace": solver_workspace,
        }

        manifest_path = manifest_dir / "build_manifest.json"
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

        return {
            "success": True,
            "project_id": project_id,
            "manifest_path": str(manifest_path),
            "manifest": manifest,
        }

    def generate_project_build(self, project_id: str) -> dict[str, Any]:
        return self.generate_build(project_id)

    def get_build_manifest(self, project_id: str) -> dict[str, Any]:
        return self.read_build_manifest(project_id)

    def read_build_manifest(self, project_id: str) -> dict[str, Any]:
        manifest_path = self._project_dir(project_id) / "build" / "manifest" / "build_manifest.json"
        if not manifest_path.exists():
            raise FileNotFoundError(f"构建 manifest 不存在，请先执行构建：{manifest_path}")

        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        project = self._load_project(project_id)
        current_topology_hash = self._topology_hash(self._extract_topology(project))
        current_build_input_hash = self._build_input_hash(project)
        manifest_hash = str(manifest.get("topology_hash") or "")
        manifest_build_input_hash = str(manifest.get("build_input_hash") or "")
        topology_stale = not manifest_hash or manifest_hash != current_topology_hash
        build_input_stale = not manifest_build_input_hash or manifest_build_input_hash != current_build_input_hash
        stale = topology_stale or build_input_stale
        manifest["topology_hash_current"] = current_topology_hash
        manifest["build_input_hash_current"] = current_build_input_hash
        manifest["manifest_stale"] = stale
        if stale:
            stale_message = "当前拓扑或构建输入已变化，现有 build manifest 已过期，请重新生成 Solver Workspace。"
            manifest["warnings"] = self._dedupe([*(manifest.get("warnings") or []), stale_message])
            manifest["ready_for_solver"] = False
            gate = manifest.get("build_gate")
            if isinstance(gate, dict):
                gate["ready_for_solver_gate"] = False
                gate["warnings"] = self._dedupe([*(gate.get("warnings") or []), stale_message])
            handoff = manifest.get("solver_handoff")
            if isinstance(handoff, dict):
                handoff["status"] = "blocked"
                handoff["warnings"] = self._dedupe([*(handoff.get("warnings") or []), stale_message])
            workspace = manifest.get("solver_workspace")
            if isinstance(workspace, dict):
                workspace["ready_for_solver"] = False
                workspace["warnings"] = self._dedupe([*(workspace.get("warnings") or []), stale_message])
        return manifest

    def _topology_hash(self, topology: dict[str, Any]) -> str:
        return topology_hash(topology)

    def _build_input_hash(self, project: dict[str, Any]) -> str:
        return build_input_hash(project)

    def _stable_hash(self, payload: Any) -> str:
        return stable_hash(payload)

    def _build_input_signature(self, project: dict[str, Any]) -> dict[str, Any]:
        return build_input_signature(project)

    def _asset_signature(self, asset: Any) -> dict[str, Any] | None:
        return asset_signature(asset)

    def _build_gate_status(self, preview: dict[str, Any], dss_payload: dict[str, Any]) -> dict[str, Any]:
        warnings: list[str] = []
        errors: list[str] = []

        summary = preview.get("summary") if isinstance(preview.get("summary"), dict) else {}
        topology_ready = bool(summary.get("ready_for_build", False))
        if not topology_ready:
            preview_errors = [str(item) for item in (summary.get("errors") or []) if str(item).strip()]
            errors.extend(preview_errors or ["拓扑预览校验未通过，不能进入求解。"])

        compile_summary = dss_payload.get("dss_compile_summary") if isinstance(dss_payload.get("dss_compile_summary"), dict) else {}
        structural = compile_summary.get("structural_checks") if isinstance(compile_summary.get("structural_checks"), dict) else {}
        structural_passed = bool(structural.get("passed", False))
        warnings.extend(str(item) for item in (structural.get("warnings") or []) if str(item).strip())
        if not structural_passed:
            structural_errors = [str(item) for item in (structural.get("errors") or []) if str(item).strip()]
            errors.extend(structural_errors or ["DSS 结构自检未通过。"])

        probe = compile_summary.get("opendss_probe") if isinstance(compile_summary.get("opendss_probe"), dict) else {}
        probe_status = str(probe.get("status") or "").strip().lower()
        compile_ok = bool(probe.get("compile_succeeded", False))
        solve_ok = bool(probe.get("solve_converged", False))
        probe_passed = probe_status == "passed" and compile_ok and solve_ok
        if not probe_passed:
            message = str(probe.get("message") or "").strip()
            if probe_status == "skipped":
                errors.append(message or "OpenDSS 实编译探测被跳过，无法确认模型可被真实 OpenDSS 求解。")
            elif probe_status == "failed":
                errors.append(message or "OpenDSS 实编译探测失败。")
            else:
                errors.append(message or "OpenDSS 实编译探测未返回通过状态。")

        errors = self._dedupe(errors)
        warnings = self._dedupe(warnings)
        return {
            "ready_for_solver_gate": len(errors) == 0,
            "topology_ready": topology_ready,
            "dss_structural_passed": structural_passed,
            "opendss_probe_passed": probe_passed,
            "opendss_probe_status": probe_status or "unknown",
            "errors": errors,
            "warnings": warnings,
        }

    @staticmethod
    def _dedupe(items: list[Any]) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        for item in items:
            text = str(item).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            out.append(text)
        return out

    def _prepare_solver_handoff(
        self,
        project_id: str,
        build_dir: Path,
        inputs_dir: Path,
        dss_payload: dict[str, Any],
        build_gate: dict[str, Any],
    ) -> dict[str, Any]:
        handoff_dir = build_dir / "solver_handoff"
        handoff_dir.mkdir(parents=True, exist_ok=True)

        target_dss_dir = handoff_dir / "dss"
        shutil.copytree(Path(dss_payload["dss_dir"]), target_dss_dir, dirs_exist_ok=True)

        summary = {
            "project_id": project_id,
            "handoff_dir": str(handoff_dir),
            "dss_dir": str(target_dss_dir),
            "dss_master_path": str(target_dss_dir / "Master.dss"),
            "status": "ready" if build_gate.get("ready_for_solver_gate") else "blocked",
            "warnings": list(build_gate.get("warnings") or []),
            "errors": list(build_gate.get("errors") or []),
            "notes": [
                "该目录用于 solver 阶段读取可视化拓扑编译后的 OpenDSS 输入。",
                "只有拓扑、DSS 结构、OpenDSS 探测和运行输入均通过时，solver 才会读取该目录。",
            ],
        }
        (handoff_dir / "handoff_summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return summary

    def _prepare_solver_workspace(
        self,
        project_id: str,
        project: dict[str, Any],
        build_dir: Path,
        dss_payload: dict[str, Any],
        build_gate: dict[str, Any],
    ) -> dict[str, Any]:
        workspace_dir = build_dir / "solver_workspace"
        inputs_dir = workspace_dir / "inputs"
        registry_dir = inputs_dir / "registry"
        node_loads_dir = inputs_dir / "node_loads"
        storage_dir = inputs_dir / "storage"
        tariff_dir = inputs_dir / "tariff"
        dss_dir = inputs_dir / "dss" / "visual_model"
        outputs_dir = workspace_dir / "outputs" / "integrated_optimization"

        for path in [registry_dir, node_loads_dir, storage_dir, tariff_dir, dss_dir.parent, outputs_dir]:
            path.mkdir(parents=True, exist_ok=True)

        shutil.copytree(Path(dss_payload["dss_dir"]), dss_dir, dirs_exist_ok=True)

        warnings: list[str] = list(build_gate.get("warnings") or [])
        errors: list[str] = list(build_gate.get("errors") or [])

        tariff_rel_path, tariff_abs_path = self._prepare_tariff_input(project, tariff_dir, errors)
        strategy_rel_path, strategy_abs_path = self._prepare_strategy_library(project, storage_dir, errors, warnings)
        registry_rows = self._build_registry_rows(
            project=project,
            node_loads_dir=node_loads_dir,
            workspace_dir=workspace_dir,
            tariff_rel_path=tariff_rel_path,
            errors=errors,
            warnings=warnings,
        )
        runtime_manifest_rel_path = self._write_network_runtime_manifest(
            registry_dir=registry_dir,
            rows=registry_rows,
        )
        if runtime_manifest_rel_path:
            for row in registry_rows:
                row["network_runtime_manifest_path"] = runtime_manifest_rel_path

        selected_target_id = self._resolve_default_storage_target(registry_rows, errors, warnings)

        registry_path = registry_dir / "node_registry.xlsx"
        self._write_registry_xlsx(registry_path, registry_rows)

        command = self._build_solver_command(
            registry_path=registry_path,
            strategy_path=strategy_abs_path,
            output_dir=outputs_dir,
            dss_master_path=dss_dir / "Master.dss",
            project=project,
            selected_target_id=selected_target_id,
        )

        command_path = workspace_dir / "solver_command.json"
        command_path.write_text(json.dumps(command, ensure_ascii=False, indent=2), encoding="utf-8")

        summary = {
            "project_id": project_id,
            "workspace_dir": str(workspace_dir.resolve()),
            "inputs_dir": str(inputs_dir.resolve()),
            "registry_path": str(registry_path.resolve()),
            "registry_relative_path": "inputs/registry/node_registry.xlsx",
            "strategy_library_path": str(strategy_abs_path.resolve()) if strategy_abs_path else None,
            "strategy_library_relative_path": strategy_rel_path,
            "tariff_path": str(tariff_abs_path.resolve()) if tariff_abs_path else None,
            "tariff_relative_path": tariff_rel_path,
            "dss_master_path": str((dss_dir / "Master.dss").resolve()),
            "network_runtime_manifest_path": str((registry_dir / "network_runtime_manifest.json").resolve()) if runtime_manifest_rel_path else None,
            "network_runtime_manifest_relative_path": runtime_manifest_rel_path,
            "selected_target_id": selected_target_id,
            "outputs_dir": str(outputs_dir.resolve()),
            "command_path": str(command_path.resolve()),
            "solver_command": command,
            "registry_row_count": len(registry_rows),
            "warnings": self._dedupe(warnings),
            "errors": self._dedupe(errors),
            "ready_for_solver": bool(build_gate.get("ready_for_solver_gate")) and len(errors) == 0 and len(registry_rows) > 0,
        }
        (workspace_dir / "workspace_summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return summary

    def _prepare_tariff_input(
        self,
        project: dict[str, Any],
        tariff_dir: Path,
        errors: list[str],
    ) -> tuple[str | None, Path | None]:
        asset = (project.get("tariff") or {}).get("asset") if isinstance(project.get("tariff"), dict) else None
        source_path = self._asset_path(asset)
        if source_path is None:
            errors.append("未绑定电价表，无法生成求解器 tariff 输入。")
            return None, None

        suffix = source_path.suffix.lower() or ".xlsx"
        target_name = "tariff_annual.csv" if suffix == ".csv" else "tariff_annual.xlsx"
        target_path = tariff_dir / target_name
        shutil.copy2(source_path, target_path)
        return f"inputs/tariff/{target_name}", target_path

    def _prepare_strategy_library(
        self,
        project: dict[str, Any],
        storage_dir: Path,
        errors: list[str],
        warnings: list[str],
    ) -> tuple[str | None, Path | None]:
        _ = warnings
        target_path = storage_dir / "工商业储能设备策略库.xlsx"
        asset = (project.get("device_library") or {}).get("asset") if isinstance(project.get("device_library"), dict) else None
        source_path = self._asset_path(asset)
        if source_path is not None and source_path.suffix.lower() in {".xlsx", ".xlsm"}:
            if not self._strategy_library_has_v2_schema(source_path):
                errors.append("设备策略库必须使用 device_library_v2 模板，不能生成求解器 strategy-library 输入。")
                return None, None
            shutil.copy2(source_path, target_path)
            return "inputs/storage/工商业储能设备策略库.xlsx", target_path

        if source_path is None:
            errors.append("未绑定 device_library_v2 设备策略库，无法生成求解器 strategy-library 输入。")
        else:
            errors.append("设备策略库必须是 device_library_v2 模板 .xlsx/.xlsm 文件。")
        return None, None

    def _load_runtime_stats(self, year_path: Path, model_path: Path) -> dict[str, Any]:
        if not year_path.exists() or not model_path.exists():
            return {}

        weights: dict[str, int] = {}
        try:
            with year_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    model_id = str(row.get("internal_model_id") or "").strip()
                    if model_id:
                        weights[model_id] = weights.get(model_id, 0) + 1
        except Exception:
            return {}

        peak_kw = None
        valley_kw = None
        weighted_mean = 0.0
        total_days = 0
        daily_energy_acc = 0.0
        try:
            with model_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    model_id = str(row.get("internal_model_id") or "").strip()
                    w = weights.get(model_id, 0)
                    if w <= 0:
                        continue
                    vals = []
                    for i in range(24):
                        for key in (f"h{i:02d}", f"H{i:02d}", f"load_{i:02d}", f"p_{i:02d}"):
                            if key in row:
                                try:
                                    vals.append(float(row[key]))
                                except Exception:
                                    vals.append(0.0)
                                break
                    if not vals:
                        continue
                    local_peak = max(vals)
                    local_valley = min(vals)
                    local_mean = sum(vals) / len(vals)
                    local_daily_energy = sum(vals)
                    peak_kw = local_peak if peak_kw is None else max(peak_kw, local_peak)
                    valley_kw = local_valley if valley_kw is None else min(valley_kw, local_valley)
                    weighted_mean += local_mean * w
                    daily_energy_acc += local_daily_energy * w
                    total_days += w
        except Exception:
            return {}

        return {
            "peak_kw": peak_kw,
            "valley_kw": valley_kw,
            "annual_mean_kw": (weighted_mean / total_days) if total_days else None,
            "mean_daily_energy_kwh": (daily_energy_acc / total_days) if total_days else None,
        }

    def _build_registry_rows(
        self,
        project: dict[str, Any],
        node_loads_dir: Path,
        workspace_dir: Path,
        tariff_rel_path: str | None,
        errors: list[str],
        warnings: list[str],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        network = project.get("network") if isinstance(project.get("network"), dict) else {}
        nodes = network.get("nodes") if isinstance(network.get("nodes"), list) else []
        edges = network.get("edges") if isinstance(network.get("edges"), list) else []
        node_map = {str(node.get("id") or ""): node for node in nodes if isinstance(node, dict)}
        economic_params = network.get("economic_parameters") if isinstance(network.get("economic_parameters"), dict) else {}
        assets = project.get("assets") if isinstance(project.get("assets"), dict) else {}
        device_records = ((project.get("device_library") or {}).get("records") or []) if isinstance(project.get("device_library"), dict) else []

        load_index = 0
        for node in nodes:
            if str(node.get("type")) != "load":
                continue
            load_index += 1
            params = node.get("params") if isinstance(node.get("params"), dict) else {}
            node_id = self._safe_int(params.get("node_id"), load_index)
            raw_category = str(params.get("category") or "industrial").strip() or "industrial"
            normalized_category = self._normalize_load_category(raw_category)
            category = self._path_segment(raw_category)
            if category != raw_category:
                warnings.append(
                    f"负荷节点 {node.get('name') or node.get('id')} 的 category 含有不安全字符，已用于路径时清洗为 {category}。"
                )
            internal_id = self._path_segment(str(node.get("id") or f"load_{load_index:03d}"))
            dss_bus_name = self._dss_bus_name(node, node_id)
            dss_load_name = self._dss_load_name(node, node_id)
            dss_phases = 3
            allow_grid_export = params.get(
                "allow_grid_export",
                params.get("allow_reverse_power_to_grid", params.get("allow_export_to_grid", False)),
            )
            node_dir_rel = f"inputs/node_loads/{category}/{internal_id}"
            node_dir = workspace_dir / node_dir_rel
            node_dir.mkdir(parents=True, exist_ok=True)

            binding = node.get("runtime_binding") if isinstance(node.get("runtime_binding"), dict) else {}
            year_asset = assets.get(str(binding.get("year_map_file_id") or ""))
            model_asset = assets.get(str(binding.get("model_library_file_id") or ""))

            year_path = self._asset_path(year_asset)
            model_path = self._asset_path(model_asset)
            if year_path is None or model_path is None:
                errors.append(f"负荷节点 {node.get('name') or node.get('id')} 未完整绑定 runtime_year_model_map/runtime_model_library。")
                continue

            shutil.copy2(year_path, node_dir / "runtime_year_model_map.csv")
            shutil.copy2(model_path, node_dir / "runtime_model_library.csv")

            load_transformer_capacity_kva = self._safe_float(params.get("transformer_capacity_kva"), 0.0)
            connected_tx = self._connected_distribution_transformer(str(node.get("id") or ""), node_map, edges)
            connected_tx_kva = self._safe_float(
                (connected_tx.get("params") if isinstance(connected_tx, dict) and isinstance(connected_tx.get("params"), dict) else {}).get("rated_kva"),
                0.0,
            )
            transformer_capacity_kva = connected_tx_kva or load_transformer_capacity_kva or None
            if normalized_category == "residential" and connected_tx_kva > 0 and load_transformer_capacity_kva > 0:
                denominator = max(connected_tx_kva, load_transformer_capacity_kva, 1.0)
                if abs(connected_tx_kva - load_transformer_capacity_kva) / denominator > 0.02:
                    warnings.append(
                        f"负荷节点 {node.get('name') or internal_id} 的 transformer_capacity_kva={load_transformer_capacity_kva:.1f} kVA "
                        f"与相连用户配变 {connected_tx.get('name') or connected_tx.get('id')} rated_kva={connected_tx_kva:.1f} kVA 不一致；"
                        "求解器已优先采用用户配变 rated_kva。"
                    )
            transformer_pf_limit = self._safe_float(params.get("transformer_pf_limit"), self._safe_float(params.get("pf"), 0.95))
            transformer_reserve_ratio = self._safe_float(params.get("transformer_reserve_ratio"), 0.15)
            grid_interconnection_limit_kw = None
            inference = self.search_space_inference.infer(
                node_params=params,
                runtime_stats=self._load_runtime_stats(year_path, model_path),
                device_records=device_records,
                transformer_capacity_kva=transformer_capacity_kva,
                transformer_pf_limit=transformer_pf_limit,
                transformer_reserve_ratio=transformer_reserve_ratio,
                grid_interconnection_limit_kw=grid_interconnection_limit_kw,
            )

            rows.append(
                {
                    "internal_model_id": internal_id,
                    "enabled": 1 if self._safe_bool(params.get("enabled"), True) else 0,
                    "optimize_storage": 1 if self._safe_bool(params.get("optimize_storage"), False) else 0,
                    "allow_grid_export": 1 if self._safe_bool(allow_grid_export, False) else 0,
                    "node_id": node_id,
                    "scenario_name": str(node.get("name") or internal_id),
                    "category": category,
                    "node_dir": node_dir_rel,
                    "year_model_map_file": "runtime_year_model_map.csv",
                    "model_library_file": "runtime_model_library.csv",
                    "tariff_path": tariff_rel_path or "",
                    "service_calendar_path": "",
                    "pv_capacity_kw": 0.0,
                    "q_to_p_ratio": self._safe_float(params.get("q_to_p_ratio"), 0.25),
                    "description": "",
                    "remarks": str(params.get("remarks") or params.get("description") or ""),
                    "target_bus_name": dss_bus_name,
                    "target_load_name": dss_load_name,
                    "target_kv_ln": self._normalize_distribution_base_kv(
                        params.get("target_kv_ln"),
                        self._safe_int(params.get("phases"), 3),
                    ),
                    "static_load_reference_kw": self._safe_float(params.get("design_kw"), 0.0),
                    "dss_bus_name": dss_bus_name,
                    "dss_load_name": dss_load_name,
                    "dss_phases": dss_phases,
                    "transformer_capacity_kva": transformer_capacity_kva,
                    "transformer_pf_limit": transformer_pf_limit,
                    "transformer_reserve_ratio": transformer_reserve_ratio,
                    "grid_interconnection_limit_kw": grid_interconnection_limit_kw,
                    "device_power_max_kw": inference.device_power_max_kw,
                    "search_power_min_kw": inference.search_power_min_kw,
                    "search_duration_min_h": inference.search_duration_min_h,
                    "search_duration_max_h": inference.search_duration_max_h,
                    "include_aux_service_revenue": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_aux_service_revenue", False), False) else 0,
                    "include_capacity_revenue": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_capacity_revenue", False), False) else 0,
                    "include_loss_reduction_revenue": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_loss_reduction_revenue", False), False) else 0,
                    "include_degradation_cost": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_degradation_cost", True), True) else 0,
                    "include_government_subsidy": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_government_subsidy", False), False) else 0,
                    "include_replacement_cost": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_replacement_cost", True), True) else 0,
                    "include_demand_saving": 1 if self._safe_bool(self._economic_param(params, economic_params, "include_demand_saving", True), True) else 0,
                    "default_capacity_price_yuan_per_kw": self._safe_float(self._economic_param(params, economic_params, "default_capacity_price_yuan_per_kw", 0.05), 0.05),
                    "default_delivery_price_yuan_per_kwh": self._safe_float(self._economic_param(params, economic_params, "default_delivery_price_yuan_per_kwh", 0.10), 0.10),
                    "default_penalty_price_yuan_per_kwh": self._safe_float(self._economic_param(params, economic_params, "default_penalty_price_yuan_per_kwh", 0.20), 0.20),
                    "default_activation_factor": self._safe_float(self._economic_param(params, economic_params, "default_activation_factor", 0.15), 0.15),
                    "max_service_power_ratio": self._safe_float(self._economic_param(params, economic_params, "max_service_power_ratio", 0.30), 0.30),
                    "capacity_service_price_yuan_per_kw_day": self._safe_float(self._economic_param(params, economic_params, "capacity_service_price_yuan_per_kw_day", 0.0), 0.0),
                    "capacity_revenue_eligible_days": self._safe_float(self._economic_param(params, economic_params, "capacity_revenue_eligible_days", 365.0), 365.0),
                    "network_loss_price_yuan_per_kwh": self._safe_float(self._economic_param(params, economic_params, "network_loss_price_yuan_per_kwh", 0.30), 0.30),
                    "network_loss_proxy_rate": self._safe_float(self._economic_param(params, economic_params, "network_loss_proxy_rate", 0.02), 0.02),
                    "government_subsidy_rate_on_capex": self._safe_float(self._economic_param(params, economic_params, "government_subsidy_rate_on_capex", 0.0), 0.0),
                    "government_subsidy_yuan_per_kwh": self._safe_float(self._economic_param(params, economic_params, "government_subsidy_yuan_per_kwh", 0.0), 0.0),
                    "government_subsidy_yuan_per_kw": self._safe_float(self._economic_param(params, economic_params, "government_subsidy_yuan_per_kw", 0.0), 0.0),
                    "government_subsidy_cap_yuan": self._safe_float(self._economic_param(params, economic_params, "government_subsidy_cap_yuan", 0.0), 0.0),
                    "replacement_cost_ratio": self._safe_float(self._economic_param(params, economic_params, "replacement_cost_ratio", 0.60), 0.60),
                    "replacement_year_override": self._safe_float(self._economic_param(params, economic_params, "replacement_year_override", 0.0), 0.0),
                    "replacement_trigger_soh": self._safe_float(self._economic_param(params, economic_params, "replacement_trigger_soh", 0.70), 0.70),
                    "replacement_reset_soh": self._safe_float(self._economic_param(params, economic_params, "replacement_reset_soh", 0.95), 0.95),
                    "project_life_years": self._safe_int(self._economic_param(params, economic_params, "project_life_years", 20), 20),
                    "discount_rate": self._safe_float(self._economic_param(params, economic_params, "discount_rate", 0.06), 0.06),
                    "annual_revenue_growth_rate": self._safe_float(self._economic_param(params, economic_params, "annual_revenue_growth_rate", 0.0), 0.0),
                    "annual_om_growth_rate": self._safe_float(self._economic_param(params, economic_params, "annual_om_growth_rate", 0.02), 0.02),
                    "integration_markup_ratio": self._safe_float(self._economic_param(params, economic_params, "integration_markup_ratio", 0.15), 0.15),
                    "safety_markup_ratio": self._safe_float(self._economic_param(params, economic_params, "safety_markup_ratio", 0.02), 0.02),
                    "other_capex_yuan": self._safe_float(self._economic_param(params, economic_params, "other_capex_yuan", 0.0), 0.0),
                    "battery_capex_share": self._safe_float(self._economic_param(params, economic_params, "battery_capex_share", 0.60), 0.60),
                    "calendar_life_years": self._safe_float(self._economic_param(params, economic_params, "calendar_life_years", 20.0), 20.0),
                    "calendar_fade_share": self._safe_float(self._economic_param(params, economic_params, "calendar_fade_share", 0.15), 0.15),
                    "dispatch_mode": str(params.get("dispatch_mode") or "hybrid"),
                    "demand_charge_yuan_per_kw_month": self._safe_float(
                        self._economic_param(
                            params,
                            economic_params,
                            "demand_charge_yuan_per_kw_month",
                            self._economic_param(params, economic_params, "daily_demand_shadow_yuan_per_kw", 48.0),
                        ),
                        48.0,
                    ),
                    "daily_demand_shadow_yuan_per_kw": self._safe_float(
                        self._economic_param(
                            params,
                            economic_params,
                            "daily_demand_shadow_yuan_per_kw",
                            self._economic_param(params, economic_params, "demand_charge_yuan_per_kw_month", 48.0),
                        ),
                        48.0,
                    ),
                    "voltage_penalty_coeff_yuan": self._safe_float(self._economic_param(params, economic_params, "voltage_penalty_coeff_yuan", 0.0), 0.0),
                    "run_mode": str(params.get("run_mode") or "single_user"),
                    "model_year": self._safe_int(params.get("model_year"), 2025),
                }
            )

        if not rows and not errors:
            warnings.append("当前拓扑没有完整可运行的负荷节点，solver registry 为空。")
        return rows

    def _write_network_runtime_manifest(
        self,
        registry_dir: Path,
        rows: list[dict[str, Any]],
    ) -> str | None:
        if not rows:
            return None

        rel_path = "inputs/registry/network_runtime_manifest.json"
        manifest = {
            "version": 1,
            "strict_runtime_only": True,
            "runtime_node_count": len(rows),
            "active_node_ids": [int(row["node_id"]) for row in rows if self._safe_bool(row.get("enabled"), True)],
            "loads": [],
        }
        for row in rows:
            manifest["loads"].append(
                {
                    "internal_model_id": row.get("internal_model_id"),
                    "node_id": row.get("node_id"),
                    "scenario_name": row.get("scenario_name"),
                    "category": row.get("category"),
                    "enabled": row.get("enabled"),
                    "optimize_storage": row.get("optimize_storage"),
                    "node_dir": row.get("node_dir"),
                    "year_model_map_file": row.get("year_model_map_file"),
                    "model_library_file": row.get("model_library_file"),
                    "dss_bus_name": row.get("dss_bus_name"),
                    "dss_load_name": row.get("dss_load_name"),
                    "dss_phases": row.get("dss_phases"),
                    "target_kv_ln": row.get("target_kv_ln"),
                    "q_to_p_ratio": row.get("q_to_p_ratio"),
                    "pv_capacity_kw": row.get("pv_capacity_kw"),
                }
            )

        path = registry_dir / "network_runtime_manifest.json"
        path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        return rel_path

    def _resolve_default_storage_target(
        self,
        rows: list[dict[str, Any]],
        errors: list[str],
        warnings: list[str],
    ) -> str | None:
        if not rows:
            return None

        targets = [
            row
            for row in rows
            if self._safe_bool(row.get("enabled"), True) and self._safe_bool(row.get("optimize_storage"), False)
        ]
        if len(targets) == 1:
            return str(targets[0].get("internal_model_id") or "")
        if not targets:
            errors.append("未指定配储目标：请在一个负荷节点上设置 optimize_storage=1，其他负荷保持 optimize_storage=0 作为背景负荷。")
            return None

        target_ids = ", ".join(str(row.get("internal_model_id") or row.get("scenario_name") or "") for row in targets)
        warnings.append(
            "检测到多个候选配储目标；构建仍可用于求解，计算运行时需要从 target_id 下拉列表选择一个。"
            f"候选目标：{target_ids}。"
        )
        return None

    def _write_registry_xlsx(self, path: Path, rows: list[dict[str, Any]]) -> None:
        headers = [
            "internal_model_id", "enabled", "optimize_storage", "allow_grid_export", "node_id", "scenario_name", "category",
            "node_dir", "year_model_map_file", "model_library_file", "network_runtime_manifest_path",
            "tariff_path", "service_calendar_path",
            "pv_capacity_kw", "q_to_p_ratio", "description", "model_year", "remarks",
            "target_bus_name", "target_load_name", "target_kv_ln", "static_load_reference_kw",
            "dss_bus_name", "dss_load_name", "dss_phases",
            "transformer_capacity_kva", "transformer_pf_limit",
            "transformer_reserve_ratio", "grid_interconnection_limit_kw", "device_power_max_kw",
            "search_power_min_kw", "search_duration_min_h", "search_duration_max_h",
            "include_aux_service_revenue", "include_capacity_revenue", "include_loss_reduction_revenue",
            "include_degradation_cost", "include_government_subsidy", "include_replacement_cost",
            "include_demand_saving",
            "default_capacity_price_yuan_per_kw", "default_delivery_price_yuan_per_kwh",
            "default_penalty_price_yuan_per_kwh", "default_activation_factor", "max_service_power_ratio",
            "capacity_service_price_yuan_per_kw_day", "capacity_revenue_eligible_days",
            "network_loss_price_yuan_per_kwh", "network_loss_proxy_rate",
            "government_subsidy_rate_on_capex", "government_subsidy_yuan_per_kwh",
            "government_subsidy_yuan_per_kw", "government_subsidy_cap_yuan",
            "replacement_cost_ratio", "replacement_year_override", "replacement_trigger_soh", "replacement_reset_soh",
            "project_life_years", "discount_rate", "annual_revenue_growth_rate", "annual_om_growth_rate",
            "integration_markup_ratio", "safety_markup_ratio", "other_capex_yuan",
            "battery_capex_share",
            "calendar_life_years", "calendar_fade_share",
            "dispatch_mode", "demand_charge_yuan_per_kw_month", "daily_demand_shadow_yuan_per_kw", "voltage_penalty_coeff_yuan",
            "run_mode",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "node_registry"
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header) for header in headers])
        wb.save(path)

    def _build_solver_command(
        self,
        registry_path: Path,
        strategy_path: Path | None,
        output_dir: Path,
        dss_master_path: Path,
        project: dict[str, Any],
        selected_target_id: str | None = None,
    ) -> dict[str, Any]:
        solve_config = project.get("solve_config") if isinstance(project.get("solve_config"), dict) else {}
        extra = solve_config.get("extra") if isinstance(solve_config.get("extra"), dict) else {}
        population_size = self._safe_int(extra.get("population_size"), 16)
        generations = self._safe_int(extra.get("generations"), 8)

        args = [
            "--registry", str(registry_path.resolve()),
            "--strategy-library", str(strategy_path.resolve()) if strategy_path else "",
            "--output-dir", str(output_dir.resolve()),
            "--population-size", str(population_size),
            "--generations", str(generations),
        ]
        if selected_target_id:
            args.extend(["--target-id", selected_target_id])
        args.extend(["--enable-opendss-oracle", "--dss-master-path", str(dss_master_path.resolve())])
        return {
            "entry": "main.py",
            "args": args,
            "registry_path": str(registry_path.resolve()),
            "strategy_library_path": str(strategy_path.resolve()) if strategy_path else None,
            "output_dir": str(output_dir.resolve()),
            "dss_master_path": str(dss_master_path.resolve()),
        }

    def _asset_path(self, asset: dict[str, Any] | None) -> Path | None:
        if not isinstance(asset, dict):
            return None
        stored_path = asset.get("metadata", {}).get("stored_path") if isinstance(asset.get("metadata"), dict) else None
        if not stored_path:
            return None
        path = Path(str(stored_path))
        return path if path.exists() else None

    def _strategy_library_has_v2_schema(self, path: Path) -> bool:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return False
        if "元数据" not in wb.sheetnames or "设备库" not in wb.sheetnames:
            return False
        rows = list(wb["元数据"].iter_rows(values_only=True))
        if not rows:
            return False
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        try:
            key_idx = headers.index("key")
            value_idx = headers.index("value")
        except ValueError:
            return False
        for row in rows[1:]:
            key = str(row[key_idx] if key_idx < len(row) else "").strip()
            value = str(row[value_idx] if value_idx < len(row) else "").strip()
            if key == "schema_version":
                return value == "device_library_v2"
        return False

    def _safe_name(self, value: str) -> str:
        return self._path_segment(value)

    def _path_segment(self, value: str) -> str:
        chars = []
        for ch in str(value).strip():
            if ch.isascii() and (ch.isalnum() or ch == "_"):
                chars.append(ch)
            else:
                chars.append("_")
        return "".join(chars).strip("_") or "unnamed"

    def _dss_safe_name(self, value: str) -> str:
        chars = []
        for ch in value.strip():
            if ch.isascii() and (ch.isalnum() or ch == "_"):
                chars.append(ch)
            else:
                chars.append("_")
        return "".join(chars).strip("_") or "unnamed"

    def _dss_bus_name(self, node: dict[str, Any], node_id: int) -> str:
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        explicit = str(params.get("dss_bus_name") or params.get("bus_name") or "").strip()
        if explicit:
            return self._dss_safe_name(explicit)
        if node_id > 0:
            return self._dss_safe_name(f"n{node_id}")
        return self._dss_safe_name(str(node.get("id") or "load"))

    def _dss_load_name(self, node: dict[str, Any], node_id: int) -> str:
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        explicit = str(params.get("dss_load_name") or params.get("load_name") or "").strip()
        if explicit:
            return self._dss_safe_name(explicit)
        if node_id > 0:
            return self._dss_safe_name(f"LD{node_id:02d}")
        return self._dss_safe_name(str(node.get("id") or "load"))

    def _safe_float(self, value: Any, default: float) -> float:
        try:
            if value in (None, ""):
                return float(default)
            return float(value)
        except Exception:
            return float(default)

    def _normalize_distribution_base_kv(self, value: Any, phases: int = 3) -> float:
        return self.dss_builder._distribution_voltage_kv_for_opendss(value, phases)

    def _safe_int(self, value: Any, default: int) -> int:
        try:
            if value in (None, ""):
                return int(default)
            return int(float(value))
        except Exception:
            return int(default)

    def _safe_bool(self, value: Any, default: bool) -> bool:
        if value in (None, ""):
            return bool(default)
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(int(value))
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "on", "是", "启用"}:
            return True
        if text in {"0", "false", "no", "n", "off", "否", "停用"}:
            return False
        return bool(default)

    def _is_distribution_transformer_node(self, node: dict[str, Any]) -> bool:
        node_type = str(node.get("type") or "").strip().lower()
        if node_type == "distribution_transformer":
            return True
        if node_type != "transformer":
            return False
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        role = str(params.get("transformer_role") or params.get("role") or "").strip().lower()
        return role in {"distribution", "distribution_transformer", "customer_distribution"} or self._safe_bool(
            params.get("is_distribution_transformer"),
            False,
        )

    @staticmethod
    def _is_low_side_resource_node(node: dict[str, Any]) -> bool:
        return str(node.get("type") or "").strip().lower() in {"load", "storage", "pv", "wind", "capacitor"}

    def _is_transformer_connection_edge(self, from_node: dict[str, Any], to_node: dict[str, Any]) -> bool:
        from_type = str(from_node.get("type") or "").strip().lower()
        to_type = str(to_node.get("type") or "").strip().lower()
        if from_type in {"grid", "source"} and to_type in {"transformer", "distribution_transformer"}:
            return True
        if to_type in {"grid", "source"} and from_type in {"transformer", "distribution_transformer"}:
            return True
        if self._is_distribution_transformer_node(to_node) and not self._is_low_side_resource_node(from_node):
            return True
        if self._is_distribution_transformer_node(from_node) and not self._is_low_side_resource_node(to_node):
            return True
        return False

    def _economic_param(
        self,
        params: dict[str, Any],
        economic_params: dict[str, Any],
        key: str,
        default: Any,
    ) -> Any:
        if key in economic_params and economic_params.get(key) not in (None, ""):
            return economic_params.get(key)
        return params.get(key, default)

    @staticmethod
    def _validate_project_id(project_id: str) -> str:
        if not project_id or not project_id.strip():
            raise ValueError("项目 ID 不能为空")
        if ".." in project_id or "/" in project_id or "\\" in project_id:
            raise ValueError(f"无效的项目 ID：{project_id}")
        return project_id.strip()

    def _project_dir(self, project_id: str) -> Path:
        BuildExportService._validate_project_id(project_id)
        return self.data_root / project_id

    def _project_file(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "project.json"

    def _load_project(self, project_id: str) -> dict[str, Any]:
        if self.project_service is not None and hasattr(self.project_service, "load_project"):
            project = self.project_service.load_project(project_id)
            if isinstance(project, dict):
                return project
            if hasattr(project, "model_dump"):
                return project.model_dump(mode="json")
            if hasattr(project, "dict"):
                return project.dict()

        project_file = self._project_file(project_id)
        if not project_file.exists():
            return {
                "project_id": project_id,
                "project_name": project_id,
                "network": {"nodes": [], "edges": [], "economic_parameters": {}},
            }
        return json.loads(project_file.read_text(encoding="utf-8"))

    def _extract_topology(self, project: dict[str, Any]) -> dict[str, Any]:
        return extract_topology(project)

    def _build_warnings(self, topology: dict[str, Any]) -> list[str]:
        return list(self._validate_topology(topology)["warnings"])

    def _build_errors(self, topology: dict[str, Any]) -> list[str]:
        return list(self._validate_topology(topology)["errors"])

    def _validate_topology(self, topology: dict[str, Any], project: dict[str, Any] | None = None) -> dict[str, Any]:
        errors: list[str] = []
        warnings: list[str] = []
        nodes = topology["nodes"]
        edges = topology["edges"]

        if not nodes:
            warnings.append("当前项目暂无节点。")
        if not edges:
            warnings.append("当前项目暂无线路。")

        grid_nodes = [node for node in nodes if str(node.get("type")).strip().lower() in {"grid", "source"}]
        transformer_nodes = [
            node
            for node in nodes
            if str(node.get("type")).strip().lower() == "transformer" and not self._is_distribution_transformer_node(node)
        ]
        load_nodes = [node for node in nodes if str(node.get("type")).strip().lower() == "load"]

        if not grid_nodes:
            errors.append("拓扑缺少电源/电网节点，无法生成完整 OpenDSS Circuit/Source。")
        if not transformer_nodes:
            errors.append("拓扑缺少主变节点，无法建立 sourcebus 到低压侧母线的电气连接。")
        if not load_nodes:
            errors.append("拓扑缺少负荷节点，无法生成 node_registry.xlsx。")

        node_id_values = [str(node.get("id")) for node in nodes if node.get("id") is not None]
        duplicates = self._find_duplicates(node_id_values)
        for duplicate in duplicates:
            errors.append(f"节点 id 重复：{duplicate}。")

        node_ids = {str(node.get("id")) for node in nodes}
        node_map = {str(node.get("id")): node for node in nodes if node.get("id") is not None}

        load_node_ids: list[str] = []
        load_dss_names: list[str] = []
        load_internal_ids: list[str] = []
        bus_names: list[str] = []
        legacy_phase_seen = False
        load_sequence = 0
        for node in nodes:
            node_id = str(node.get("id") or "")
            node_type = str(node.get("type") or "").strip().lower()
            params = node.get("params") if isinstance(node.get("params"), dict) else {}
            phases = self._safe_int(params.get("phases"), 3)
            if phases != 3:
                legacy_phase_seen = True

            if node_type not in {"grid", "source"}:
                bus = self._dss_bus_name(node, self._safe_int(params.get("node_id"), 0))
                bus_names.append(bus)
                if not str(params.get("dss_bus_name") or params.get("bus_name") or "").strip():
                    warnings.append(f"节点 {node.get('name') or node_id} 未填写 dss_bus_name，Build 将按规则自动生成。")

            if node_type == "load":
                load_sequence += 1
                raw_category = str(params.get("category") or "industrial").strip() or "industrial"
                safe_category = self._path_segment(raw_category)
                if safe_category != raw_category:
                    warnings.append(f"负荷节点 {node.get('name') or node_id} 的 category 会在生成文件路径时清洗为 {safe_category}。")
                load_internal_ids.append(self._path_segment(str(node.get("id") or f"load_{load_sequence:03d}")))

                registry_node_id = str(params.get("node_id") or "").strip()
                if registry_node_id:
                    load_node_ids.append(registry_node_id)
                else:
                    warnings.append(f"负荷节点 {node.get('name') or node_id} 未填写 node_id，registry 将按负荷顺序补齐。")

                load_name = self._dss_load_name(node, self._safe_int(params.get("node_id"), 0))
                load_dss_names.append(load_name)
                if not str(params.get("dss_load_name") or params.get("load_name") or "").strip():
                    warnings.append(f"负荷节点 {node.get('name') or node_id} 未填写 dss_load_name，Build 将按 node_id 自动生成。")

                target_kv = self._safe_float(params.get("target_kv_ln"), 0.0)
                if target_kv <= 0:
                    errors.append(f"负荷节点 {node.get('name') or node_id} 缺少基准电压。")
                if self._safe_float(params.get("design_kw"), 0.0) <= 0:
                    warnings.append(f"负荷节点 {node.get('name') or node_id} 的设计负荷 design_kw 未设置或为 0。")
                self._append_load_runtime_capacity_warnings(
                    project=project,
                    node=node,
                    node_map=node_map,
                    edges=edges,
                    warnings=warnings,
                )

        for duplicate in self._find_duplicates(load_node_ids):
            errors.append(f"负荷 node_id 重复：{duplicate}。")
        for duplicate in self._find_duplicates(load_internal_ids):
            errors.append(f"负荷节点清洗后的 internal_model_id 重复：{duplicate}。请调整负荷节点 id，避免中文/特殊字符清洗后碰撞。")
        for duplicate in self._find_duplicates(load_dss_names):
            errors.append(f"OpenDSS 负荷名重复：{duplicate}。")
        for duplicate in self._find_duplicates(bus_names):
            warnings.append(f"多个可视化节点映射到同一 OpenDSS 母线：{duplicate}，请确认这不是误填。")

        active_adjacency: dict[str, set[str]] = {node_id: set() for node_id in node_ids}
        active_edge_count = 0
        for edge in edges:
            edge_id = str(edge.get("id") or "")
            from_node_id = str(edge.get("from_node_id", "")).strip()
            to_node_id = str(edge.get("to_node_id", "")).strip()
            if from_node_id not in node_ids:
                errors.append(f"线路 {edge_id} 起点无效。")
                continue
            if to_node_id not in node_ids:
                errors.append(f"线路 {edge_id} 终点无效。")
                continue

            params = edge.get("params") if isinstance(edge.get("params"), dict) else {}
            linecode = str(params.get("linecode") or params.get("line_code") or "").strip()
            from_node = node_map.get(from_node_id, {})
            to_node = node_map.get(to_node_id, {})
            is_transformer_connection = self._is_transformer_connection_edge(from_node, to_node)
            if not is_transformer_connection:
                if self.dss_builder._bus_name(from_node) == self.dss_builder._bus_name(to_node):
                    errors.append(
                        f"线路 {edge_id} 两端映射到同一 OpenDSS 母线 {self.dss_builder._bus_name(from_node)}。"
                        "如果这是同母线挂载关系，请删除线路；如果是实际线路，请设置不同 dss_bus_name。"
                    )
                    continue
                has_explicit_rx = bool(params.get("r_ohm_per_km") not in (None, "") and params.get("x_ohm_per_km") not in (None, ""))
                has_geometry = bool(params.get("geometry") or params.get("line_geometry"))
                if not (linecode or has_explicit_rx or has_geometry):
                    errors.append(f"线路 {edge_id} 缺少 LineCode、显式阻抗或 Geometry/WireData，不能使用后端默认线路模板。")
                if linecode:
                    missing_linecode_fields = [
                        field
                        for field in ("r_ohm_per_km", "x_ohm_per_km", "r0_ohm_per_km", "x0_ohm_per_km", "c1_nf_per_km", "c0_nf_per_km")
                        if params.get(field) in (None, "") and not self.dss_builder._linecode_library_has(linecode, field)
                    ]
                    if missing_linecode_fields:
                        errors.append(f"线路 {edge_id} 的 LineCode 参数不完整：{', '.join(missing_linecode_fields)}。")
                elif has_explicit_rx:
                    missing_sequence_fields = [field for field in ("r0_ohm_per_km", "x0_ohm_per_km") if params.get(field) in (None, "")]
                    if missing_sequence_fields:
                        errors.append(f"线路 {edge_id} 显式阻抗缺少零序参数：{', '.join(missing_sequence_fields)}。")

                phases = self._safe_int(params.get("phases"), 0)
                if phases not in {1, 2, 3}:
                    errors.append(f"线路 {edge_id} 缺少有效相数 phases。")
                elif phases != 3:
                    legacy_phase_seen = True

                if self._safe_float(params.get("length_km"), 0.0) <= 0:
                    errors.append(f"线路 {edge_id} 长度 length_km 必须大于 0。")
                if not self.dss_builder._line_length_unit_is_explicit(params):
                    errors.append(f"线路 {edge_id} 缺少长度单位 units。")
                if self._safe_float(params.get("rated_current_a") or params.get("normamps"), 0.0) <= 0:
                    errors.append(f"线路 {edge_id} 缺少额定电流 rated_current_a/normamps。")
                if self._safe_float(params.get("emerg_current_a") or params.get("emergamps"), 0.0) <= 0:
                    errors.append(f"线路 {edge_id} 缺少应急电流 emerg_current_a/emergamps。")

                from_phases = self._safe_int((node_map[from_node_id].get("params") or {}).get("phases") if isinstance(node_map[from_node_id].get("params"), dict) else None, 3)
                to_phases = self._safe_int((node_map[to_node_id].get("params") or {}).get("phases") if isinstance(node_map[to_node_id].get("params"), dict) else None, 3)
                if phases in {1, 3} and (from_phases in {1, 3} and to_phases in {1, 3}) and (phases != from_phases or phases != to_phases):
                    warnings.append(f"线路 {edge_id} 相数与端点节点相数不一致。")

            active = self._safe_bool(params.get("enabled"), True) and not self._safe_bool(params.get("normally_open"), False)
            if active:
                active_edge_count += 1
                active_adjacency[from_node_id].add(to_node_id)
                active_adjacency[to_node_id].add(from_node_id)

        nodes_by_bus: dict[str, list[str]] = {}
        for node in nodes:
            node_id = str(node.get("id") or "")
            if not node_id:
                continue
            nodes_by_bus.setdefault(self.dss_builder._bus_name(node), []).append(node_id)
        for same_bus_node_ids in nodes_by_bus.values():
            if len(same_bus_node_ids) < 2:
                continue
            anchor = same_bus_node_ids[0]
            for node_id in same_bus_node_ids[1:]:
                active_adjacency.setdefault(anchor, set()).add(node_id)
                active_adjacency.setdefault(node_id, set()).add(anchor)

        if legacy_phase_seen:
            warnings.append("检测到拓扑中存在非三相相数设置，请确认相别、相序与端点节点一致。")

        roots = {str(node.get("id")) for node in [*grid_nodes, *transformer_nodes] if node.get("id") is not None}
        visited = self._walk_graph(active_adjacency, roots)
        disconnected = sorted(node_id for node_id in node_ids if node_id not in visited)
        for node_id in disconnected[:12]:
            node = node_map.get(node_id, {})
            errors.append(f"节点 {node.get('name') or node_id} 未通过闭合线路连接到电源/主变。")
        if len(disconnected) > 12:
            errors.append(f"另有 {len(disconnected) - 12} 个节点未连接到电源/主变。")

        return {
            "ready_for_build": len(nodes) > 0 and len(errors) == 0,
            "warnings": warnings,
            "errors": errors,
            "grid_count": len(grid_nodes),
            "transformer_count": len(transformer_nodes),
            "load_count": len(load_nodes),
            "active_edge_count": active_edge_count,
            "disconnected_count": len(disconnected),
        }

    def _append_load_runtime_capacity_warnings(
        self,
        *,
        project: dict[str, Any] | None,
        node: dict[str, Any],
        node_map: dict[str, dict[str, Any]],
        edges: list[dict[str, Any]],
        warnings: list[str],
    ) -> None:
        if not isinstance(project, dict):
            return

        binding = node.get("runtime_binding") if isinstance(node.get("runtime_binding"), dict) else {}
        if not binding:
            return
        assets = project.get("assets") if isinstance(project.get("assets"), dict) else {}
        year_path = self._asset_path(assets.get(str(binding.get("year_map_file_id") or "")))
        model_path = self._asset_path(assets.get(str(binding.get("model_library_file_id") or "")))
        if year_path is None or model_path is None:
            return

        stats = self._load_runtime_stats(year_path, model_path)
        peak_kw = self._safe_float(stats.get("peak_kw"), 0.0)
        if peak_kw <= 0:
            return

        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        if self._normalize_load_category(params.get("category")) != "residential":
            return

        label = str(node.get("name") or node.get("id") or "负荷节点")
        design_kw = self._safe_float(params.get("design_kw"), 0.0)
        connected_tx = self._connected_distribution_transformer(str(node.get("id") or ""), node_map, edges)
        connected_tx_params = connected_tx.get("params") if isinstance(connected_tx, dict) and isinstance(connected_tx.get("params"), dict) else {}
        connected_tx_kva = self._safe_float(connected_tx_params.get("rated_kva"), 0.0)
        load_tx_kva = self._safe_float(params.get("transformer_capacity_kva"), 0.0)

        if connected_tx_kva > 0 and load_tx_kva > 0:
            denominator = max(connected_tx_kva, load_tx_kva, 1.0)
            if abs(connected_tx_kva - load_tx_kva) / denominator > 0.02:
                warnings.append(
                    f"负荷节点 {label} 的 transformer_capacity_kva={load_tx_kva:.1f} kVA 与相连用户配变 "
                    f"{connected_tx.get('name') or connected_tx.get('id')} rated_kva={connected_tx_kva:.1f} kVA 不一致；"
                    "OpenDSS 建模以用户配变节点为准。"
                )

        capacity_kva = connected_tx_kva or load_tx_kva

        q_to_p_ratio = self._safe_float(params.get("q_to_p_ratio"), 0.0)
        if q_to_p_ratio <= 0 and design_kw > 0:
            q_to_p_ratio = self._safe_float(params.get("kvar"), 0.0) / max(design_kw, 1e-9)
        peak_kva = peak_kw * math.sqrt(1.0 + max(q_to_p_ratio, 0.0) ** 2)
        pf_limit = min(max(self._safe_float(params.get("transformer_pf_limit"), self._safe_float(params.get("pf"), 0.95)), 0.0), 1.0)
        reserve_ratio = min(max(self._safe_float(params.get("transformer_reserve_ratio"), 0.15), 0.0), 0.95)
        operating_denominator = pf_limit * (1.0 - reserve_ratio)
        required_kva = peak_kva
        if operating_denominator > 0:
            required_kva = max(required_kva, peak_kw / operating_denominator)
        recommended_kva = self._next_standard_transformer_kva(required_kva)

        if capacity_kva <= 0:
            warnings.append(
                f"居民负荷节点 {label} 未设置相连用户配变容量；按导入曲线峰值 {peak_kw:.1f} kW "
                f"估算，建议用户配变 rated_kva 不低于 {recommended_kva:.0f} kVA。"
            )
            return

        if required_kva > capacity_kva * 1.02:
            warnings.append(
                f"居民负荷节点 {label} 当前相连/配置配变容量 {capacity_kva:.1f} kVA；按导入曲线峰值 {peak_kw:.1f} kW、"
                f"无功比例和 {reserve_ratio:.0%} 备用率估算，建议容量不低于 {recommended_kva:.0f} kVA。"
                "若已有真实台账容量，请以台账为准；若没有真实容量，建议将该值回填到相连用户配变 rated_kva。"
            )

    def _connected_distribution_transformer(
        self,
        node_id: str,
        node_map: dict[str, dict[str, Any]],
        edges: list[dict[str, Any]],
    ) -> dict[str, Any] | None:
        if not node_id:
            return None
        for edge in edges:
            if not isinstance(edge, dict):
                continue
            from_node_id = str(edge.get("from_node_id") or "").strip()
            to_node_id = str(edge.get("to_node_id") or "").strip()
            if from_node_id != node_id and to_node_id != node_id:
                continue
            other_id = to_node_id if from_node_id == node_id else from_node_id
            other_node = node_map.get(other_id)
            if other_node and self._is_distribution_transformer_node(other_node):
                return other_node
        return None

    @staticmethod
    def _normalize_load_category(value: Any) -> str:
        text = str(value or "industrial").strip().lower()
        if text in {"residential", "resident", "居民", "居民负荷"}:
            return "residential"
        if text in {"commercial", "commerce", "business", "商业", "商业负荷"}:
            return "commercial"
        return "industrial"

    @staticmethod
    def _next_standard_transformer_kva(required_kva: float) -> float:
        standards = (50, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000)
        if required_kva <= 0:
            return 0.0
        for value in standards:
            if required_kva <= value:
                return float(value)
        return math.ceil(required_kva / 1000.0) * 1000.0

    def _find_duplicates(self, values: list[str]) -> list[str]:
        seen: set[str] = set()
        duplicates: set[str] = set()
        for value in values:
            key = str(value).strip()
            if not key:
                continue
            if key in seen:
                duplicates.add(key)
            seen.add(key)
        return sorted(duplicates)

    def _walk_graph(self, adjacency: dict[str, set[str]], roots: set[str]) -> set[str]:
        visited: set[str] = set()
        stack = [root for root in roots if root in adjacency]
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            stack.extend(sorted(adjacency.get(current, set()) - visited))
        return visited
