from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import signal
import subprocess
import sys
import threading
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from services.project_model_service import ProjectModelService


class SolverExecutionService:
    """读取 solver 任务状态、日志和结果文件。

    兼容两种构造方式：
    1. SolverExecutionService(project_service=self.project_service)
    2. 旧代码误传 SolverExecutionService(self.project_service)
    3. SolverExecutionService(data_root="data/projects")
    """

    NETWORK_TOPOLOGY_SUMMARY_CACHE_FILE = "network_topology_summary_cache.json"
    NETWORK_TOPOLOGY_SUMMARY_CACHE_VERSION = 1

    def __init__(
        self,
        data_root: str | Path | ProjectModelService | None = None,
        project_service: Optional[ProjectModelService] = None,
    ) -> None:
        backend_root = Path(__file__).resolve().parents[1]

        if isinstance(data_root, ProjectModelService):
            project_service = data_root
            data_root = None

        self.project_service = project_service

        if data_root is not None:
            self.data_root = Path(data_root)
        elif project_service is not None and hasattr(project_service, "base_dir"):
            self.data_root = Path(project_service.base_dir)
        else:
            self.data_root = backend_root / "data" / "projects"
        self._last_network_topology_cache_diagnostics: Dict[str, Any] | None = None

    # ---------------------------------------------------------------------
    # 基础路径
    # ---------------------------------------------------------------------
    @staticmethod
    def _validate_project_id(project_id: str) -> str:
        if not project_id or not project_id.strip():
            raise ValueError("项目 ID 不能为空")
        if ".." in project_id or "/" in project_id or "\\" in project_id:
            raise ValueError(f"无效的项目 ID：{project_id}")
        return project_id.strip()

    def _project_dir(self, project_id: str) -> Path:
        self._validate_project_id(project_id)
        if self.project_service is not None and hasattr(self.project_service, "_project_dir"):
            return Path(self.project_service._project_dir(project_id))
        return self.data_root / project_id

    def _solver_runs_dir(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "solver_runs"

    def _build_dir(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "build"

    def _task_dir(self, project_id: str, task_id: str) -> Path:
        return self._solver_runs_dir(project_id) / f"task_{task_id}"

    @staticmethod
    def _model_dump(model: Any) -> Dict[str, Any]:
        if hasattr(model, "model_dump"):
            return model.model_dump(mode="json")
        if hasattr(model, "dict"):
            return model.dict()
        if isinstance(model, dict):
            return model
        return {}

    def configure_solver(self, project_id: str, solver_binding: Dict[str, Any]) -> Dict[str, Any]:
        if self.project_service is None:
            raise RuntimeError("未配置 ProjectModelService，无法保存 solver 配置。")
        project = self.project_service.load_project(project_id)
        current = self._model_dump(getattr(project, "solver_binding", {}))
        current.update(solver_binding or {})
        from models.project_model import SolverBindingConfig

        project.solver_binding = SolverBindingConfig(**current)
        self.project_service.save_project(project)
        return self._model_dump(project.solver_binding)

    def get_solver_config(self, project_id: str) -> Dict[str, Any]:
        if self.project_service is not None:
            project = self.project_service.load_project(project_id)
            return self._model_dump(getattr(project, "solver_binding", {}))
        return {}

    def run_solver(self, project_id: str, request: Dict[str, Any] | None = None) -> Dict[str, Any]:
        project_dir = self._project_dir(project_id)
        if not project_dir.exists():
            raise FileNotFoundError(f"项目不存在：{project_id}")

        request = request or {}
        manifest_path = self._build_dir(project_id) / "manifest" / "build_manifest.json"
        manifest = self._safe_load_json(manifest_path, default={})
        workspace_info = manifest.get("solver_workspace") if isinstance(manifest, dict) else None
        if not isinstance(workspace_info, dict):
            raise FileNotFoundError("未找到 solver workspace，请先在构建校验页执行构建。")
        if not bool(workspace_info.get("ready_for_solver", False)):
            errors = workspace_info.get("errors") or []
            raise ValueError("solver workspace 尚未就绪：" + "；".join(map(str, errors)))

        task_id = uuid.uuid4().hex[:12]
        task_dir = self._task_dir(project_id, task_id)
        task_dir.mkdir(parents=True, exist_ok=True)
        stdout_log = task_dir / "stdout.log"
        stderr_log = task_dir / "stderr.log"
        state_dir = task_dir / "state"
        state_dir.mkdir(parents=True, exist_ok=True)

        build_workspace = Path(str(workspace_info["workspace_dir"]))
        task_workspace = task_dir / "solver_workspace"
        if task_workspace.exists():
            shutil.rmtree(task_workspace)
        shutil.copytree(build_workspace, task_workspace)

        solver_project_root = self._resolve_solver_project_root(project_id)
        python_executable = self._resolve_python_executable(project_id, solver_project_root)
        output_subdir = str(request.get("output_subdir_name") or "integrated_optimization").strip() or "integrated_optimization"
        output_dir = task_workspace / "outputs" / output_subdir
        output_dir.mkdir(parents=True, exist_ok=True)

        command = self._build_runtime_command(
            python_executable=python_executable,
            solver_project_root=solver_project_root,
            task_workspace=task_workspace,
            output_dir=output_dir,
            request=request,
        )

        task = {
            "task_id": task_id,
            "project_id": project_id,
            "task_name": request.get("task_name") or "storage_solver_run",
            "status": "running",
            "message": "求解器已启动。",
            "return_code": None,
            "started_at": self._now(),
            "completed_at": None,
            "stdout_log": str(stdout_log.resolve()),
            "stderr_log": str(stderr_log.resolve()),
            "command": command,
            "solver_workspace": str(task_workspace.resolve()),
            "outputs_dir": str(output_dir.resolve()),
            "progress_hint": {
                "percent": 1,
                "label": "求解器已启动",
                "detail": "正在等待求解器输出进度日志。",
                "source": "backend_log_parser",
            },
            "metadata": {
                "build_manifest": str(manifest_path.resolve()) if manifest_path.exists() else None,
                "solver_workspace": workspace_info,
                "summary_rows": [],
                "run_request": request,
                "progress_hint": {
                    "percent": 1,
                    "label": "求解器已启动",
                    "detail": "正在等待求解器输出进度日志。",
                    "source": "backend_log_parser",
                },
            },
        }

        stdout_handle = stdout_log.open("w", encoding="utf-8", errors="replace")
        stderr_handle = stderr_log.open("w", encoding="utf-8", errors="replace")
        stdout_handle.write("[solver] command: " + " ".join(command) + "\n")
        stdout_handle.flush()

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUNBUFFERED"] = "1"
        env["PYTHONUTF8"] = "1"
        try:
            process = subprocess.Popen(
                command,
                cwd=str(solver_project_root),
                stdout=stdout_handle,
                stderr=stderr_handle,
                text=True,
                env=env,
            )
        except Exception as exc:
            stderr_handle.write(f"[solver] failed to start: {exc}\n")
            stdout_handle.close()
            stderr_handle.close()
            task["status"] = "failed"
            task["message"] = f"求解器启动失败：{exc}"
            task["completed_at"] = self._now()
            self._write_task_files(task_dir, task)
            return task
        task["pid"] = process.pid
        self._write_task_files(task_dir, task)

        watcher = threading.Thread(
            target=self._watch_solver_process,
            args=(process, stdout_handle, stderr_handle, task_dir, task),
            daemon=True,
        )
        watcher.start()
        return task

    def _watch_solver_process(
        self,
        process: subprocess.Popen,
        stdout_handle,
        stderr_handle,
        task_dir: Path,
        task: Dict[str, Any],
    ) -> None:
        return_code = process.wait()
        stdout_handle.close()
        stderr_handle.close()
        current = (
            self._safe_load_json(task_dir / "state" / "task.json", default={})
            or self._safe_load_json(task_dir / "task_meta.json", default={})
            or task
        )
        final_task = dict(current)
        status = str(final_task.get("status") or "").lower()
        cancel_requested = status in {"cancelling", "cancelled", "canceling", "canceled"}
        final_task["return_code"] = return_code
        final_task["completed_at"] = self._now()
        if cancel_requested:
            final_task["status"] = "cancelled"
            final_task["message"] = "求解已终止。"
        else:
            final_task["status"] = "completed" if return_code == 0 else "failed"
            final_task["message"] = "求解完成。" if return_code == 0 else f"求解失败，返回码 {return_code}。"
        final_task["metadata"] = dict(final_task.get("metadata") or {})
        final_task["metadata"]["summary_rows"] = self._load_task_summary_rows(final_task)
        stdout_path = Path(str(final_task.get("stdout_log") or ""))
        stdout_info = self._read_text_file(stdout_path)
        final_task = self._with_progress_hint(final_task, stdout_info["text"])
        self._write_task_files(task_dir, final_task)

    def cancel_task(self, task_id: str, project_id: Optional[str] = None) -> Dict[str, Any]:
        task = self.get_task(task_id=task_id, project_id=project_id)
        if not task:
            raise FileNotFoundError(f"未找到任务：{task_id}")

        task_project_id = str(task.get("project_id") or project_id or "").strip()
        if not task_project_id:
            raise ValueError("任务缺少 project_id，无法定位状态文件。")
        status = str(task.get("status") or "").lower()
        if status not in {"running", "cancelling", "canceling"}:
            return task

        pid = self._safe_int(task.get("pid"), 0)
        if pid <= 0:
            task_dir = self._task_dir(task_project_id, task_id)
            task = self._finalize_stale_active_task(task_dir, task, reason="missing_pid", requested_cancel=True)
            return task

        task_dir = self._task_dir(task_project_id, task_id)
        metadata = dict(task.get("metadata") or {})
        metadata["cancel_requested_at"] = self._now()
        task["metadata"] = metadata
        task["status"] = "cancelling"
        task["message"] = "已请求终止求解器进程。"
        self._write_task_files(task_dir, task)

        result = self._terminate_process_tree(pid)
        metadata = dict(task.get("metadata") or {})
        metadata["cancel_result"] = result
        task["metadata"] = metadata
        if result.get("return_code") == 0 or result.get("not_found") is True:
            task["status"] = "cancelled"
            task["message"] = "求解已终止。" if result.get("return_code") == 0 else "求解器进程已不存在，任务已收敛为终止状态。"
            task["completed_at"] = self._now()
        elif result.get("return_code") is not None:
            task["message"] = "已请求终止求解器进程，等待后台确认进程状态。"
        self._write_task_files(task_dir, task)
        return task

    def _terminate_process_tree(self, pid: int) -> Dict[str, Any]:
        try:
            if os.name == "nt":
                result = subprocess.run(
                    ["taskkill", "/PID", str(pid), "/T", "/F"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=10,
                )
                return {
                    "return_code": result.returncode,
                    "stdout": result.stdout[-2000:],
                    "stderr": result.stderr[-2000:],
                    "not_found": self._taskkill_not_found(result.stdout, result.stderr),
                }
            os.kill(pid, signal.SIGTERM)
            return {"return_code": 0, "stdout": "", "stderr": ""}
        except Exception as exc:
            return {"return_code": None, "stdout": "", "stderr": str(exc)}

    @staticmethod
    def _taskkill_not_found(stdout: str, stderr: str) -> bool:
        text = f"{stdout}\n{stderr}".lower()
        return "not found" in text or "没有找到" in text or "找不到" in text

    @staticmethod
    def _is_active_task_status(status: Any) -> bool:
        return str(status or "").strip().lower() in {"running", "cancelling", "canceling"}

    def _settle_stale_active_task(self, task: Dict[str, Any], project_id: str) -> Dict[str, Any]:
        if not self._is_active_task_status(task.get("status")):
            return task

        task_id = str(task.get("task_id") or "").strip()
        normalized_project_id = str(project_id or task.get("project_id") or "").strip()
        if not task_id or not normalized_project_id:
            return task

        metadata = task.get("metadata") if isinstance(task.get("metadata"), dict) else {}
        cancel_result = metadata.get("cancel_result") if isinstance(metadata, dict) else None
        status = str(task.get("status") or "").strip().lower()
        if status in {"cancelling", "canceling"} and isinstance(cancel_result, dict):
            if cancel_result.get("not_found") is True or self._taskkill_not_found(
                str(cancel_result.get("stdout") or ""),
                str(cancel_result.get("stderr") or ""),
            ):
                return self._finalize_stale_active_task(
                    self._task_dir(normalized_project_id, task_id),
                    task,
                    reason="cancel_process_not_found",
                    requested_cancel=True,
                )

        pid = self._safe_int(task.get("pid"), 0)
        if pid > 0 and self._pid_matches_task_process(pid, task):
            return task

        reason = "missing_pid" if pid <= 0 else "process_not_found"
        return self._finalize_stale_active_task(
            self._task_dir(normalized_project_id, task_id),
            task,
            reason=reason,
            requested_cancel=status in {"cancelling", "canceling"},
        )

    def _finalize_stale_active_task(
        self,
        task_dir: Path,
        task: Dict[str, Any],
        *,
        reason: str,
        requested_cancel: bool = False,
    ) -> Dict[str, Any]:
        final_task = dict(task)
        metadata = dict(final_task.get("metadata") or {})
        now = self._now()
        previous_progress = final_task.get("progress_hint")
        previous_percent = 1
        if isinstance(previous_progress, dict):
            previous_percent = max(1, self._safe_int(previous_progress.get("percent"), 1))

        metadata["auto_settled_at"] = now
        metadata["auto_settle_reason"] = reason
        status = str(final_task.get("status") or "").strip().lower()

        if requested_cancel or status in {"cancelling", "canceling"}:
            final_task["status"] = "cancelled"
            final_task["message"] = "求解器进程已不存在，任务已收敛为终止状态。"
            progress_hint = {
                "percent": previous_percent,
                "label": "运行已终止",
                "detail": "求解器进程已不存在，任务已收敛为终止状态。",
                "source": "backend_stale_task_settlement",
            }
        elif self._task_has_complete_result_outputs(final_task):
            final_task["status"] = "completed"
            final_task["message"] = "检测到结果文件已生成，任务已自动收敛为完成状态。"
            metadata["summary_rows"] = self._load_task_summary_rows(final_task)
            progress_hint = {
                "percent": 100,
                "label": "求解完成",
                "detail": "检测到结果文件已生成，任务已自动收敛为完成状态。",
                "source": "backend_stale_task_settlement",
            }
        else:
            final_task["status"] = "failed"
            final_task["message"] = "求解器进程已不存在，任务已自动收敛为中断状态。"
            progress_hint = {
                "percent": previous_percent,
                "label": "运行中断",
                "detail": "求解器进程已不存在，任务已自动收敛为中断状态。",
                "source": "backend_stale_task_settlement",
            }

        final_task["completed_at"] = final_task.get("completed_at") or now
        final_task["metadata"] = metadata
        final_task["progress_hint"] = progress_hint
        metadata["progress_hint"] = progress_hint
        self._write_task_files(task_dir, final_task)
        return final_task

    def _task_has_complete_result_outputs(self, task: Dict[str, Any]) -> bool:
        raw_output_dir = str(task.get("outputs_dir") or "").strip()
        if not raw_output_dir:
            return False
        output_dir = Path(raw_output_dir)
        if not output_dir.exists() or not output_dir.is_dir():
            return False

        overall_best = self._safe_load_json(output_dir / "overall_best_schemes.json", default=None)
        if isinstance(overall_best, list) and len(overall_best) > 0:
            return True

        required_case_files = {
            "best_result_summary.json",
            "best_annual_hourly_operation.csv",
            "best_financial_summary.csv",
        }
        for case_dir in [output_dir, *[path for path in output_dir.iterdir() if path.is_dir()]]:
            if all((case_dir / name).exists() for name in required_case_files):
                return True
        return False

    def _pid_is_running(self, pid: int) -> bool:
        if pid <= 0:
            return False
        if os.name == "nt":
            try:
                result = subprocess.run(
                    ["tasklist", "/FI", f"PID eq {pid}", "/FO", "CSV", "/NH"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=5,
                )
            except Exception:
                return True
            text = f"{result.stdout}\n{result.stderr}".lower()
            if result.returncode != 0:
                return True
            if str(pid) not in text:
                return False
            if "no tasks are running" in text or "没有运行" in text:
                return False
            return True

        try:
            os.kill(pid, 0)
        except ProcessLookupError:
            return False
        except PermissionError:
            return True
        except Exception:
            return True
        return True

    def _pid_matches_task_process(self, pid: int, task: Dict[str, Any]) -> bool:
        if pid <= 0:
            return False
        if os.name != "nt":
            return self._pid_is_running(pid)

        exists, command_line = self._read_windows_process_info(pid)
        if exists is False:
            return False
        if exists is True:
            if not command_line:
                return True
            return self._command_line_matches_task(command_line, task)
        return self._pid_is_running(pid)

    def _read_windows_process_info(self, pid: int) -> Tuple[Optional[bool], Optional[str]]:
        try:
            result = subprocess.run(
                ["wmic", "process", "where", f"processid={pid}", "get", "ProcessId,CommandLine", "/format:list"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=5,
            )
        except Exception:
            return None, None

        text = f"{result.stdout}\n{result.stderr}"
        if result.returncode != 0:
            return None, None
        if str(pid) not in text:
            return False, ""

        command_line = ""
        for line in text.splitlines():
            if line.lower().startswith("commandline="):
                command_line = line.split("=", 1)[1].strip()
                break
        return True, command_line

    @staticmethod
    def _normalize_process_text(value: str) -> str:
        return str(value or "").strip().strip('"').replace("/", "\\").lower()

    def _command_line_matches_task(self, command_line: str, task: Dict[str, Any]) -> bool:
        text = self._normalize_process_text(command_line)
        task_id = str(task.get("task_id") or "").strip()
        specific_anchors = [
            str(task.get("solver_workspace") or ""),
            str(task.get("outputs_dir") or ""),
            f"task_{task_id}" if task_id else "",
        ]
        anchors = [self._normalize_process_text(value) for value in specific_anchors if str(value or "").strip()]
        if anchors:
            return any(anchor in text for anchor in anchors)

        command = task.get("command")
        if isinstance(command, list):
            main_anchors = [
                self._normalize_process_text(str(token))
                for token in command
                if str(token).replace("/", "\\").lower().endswith("\\main.py")
            ]
            if main_anchors:
                return any(anchor in text for anchor in main_anchors)
        return True

    def _write_task_files(self, task_dir: Path, task: Dict[str, Any]) -> None:
        data = json.dumps(task, ensure_ascii=False, indent=2)
        (task_dir / "task_meta.json").write_text(data, encoding="utf-8")
        state_dir = task_dir / "state"
        state_dir.mkdir(parents=True, exist_ok=True)
        (state_dir / "task.json").write_text(data, encoding="utf-8")

    def _load_task_summary_rows(self, task: Dict[str, Any]) -> List[Dict[str, Any]]:
        output_dir = Path(str(task.get("outputs_dir") or ""))
        rows = self._safe_load_json(output_dir / "overall_best_schemes.json", default=[])
        return rows if isinstance(rows, list) else []

    def _build_runtime_command(
        self,
        python_executable: str,
        solver_project_root: Path,
        task_workspace: Path,
        output_dir: Path,
        request: Dict[str, Any],
    ) -> List[str]:
        registry_path = task_workspace / "inputs" / "registry" / "node_registry.xlsx"
        strategy_path = task_workspace / "inputs" / "storage" / "工商业储能设备策略库.xlsx"
        command = [
            python_executable,
            "-u",
            str((solver_project_root / "main.py").resolve()),
            "--registry", str(registry_path.resolve()),
            "--strategy-library", str(strategy_path.resolve()),
            "--output-dir", str(output_dir.resolve()),
            "--population-size", str(self._safe_int(request.get("population_size"), 16)),
            "--generations", str(self._safe_int(request.get("generations"), 8)),
        ]
        target_id = str(request.get("target_id") or "").strip()
        if not target_id:
            target_id = self._resolve_single_registry_target_id(registry_path)
        if target_id:
            command.extend(["--target-id", target_id])

        initial_soc = self._safe_float(request.get("initial_soc"), math.nan)
        if math.isfinite(initial_soc):
            command.extend(["--initial-soc", self._format_float(min(max(initial_soc, 0.0), 1.0))])

        terminal_soc_mode = str(request.get("terminal_soc_mode") or "").strip().lower()
        fixed_terminal_soc_target = self._safe_float(request.get("fixed_terminal_soc_target"), math.nan)
        if not terminal_soc_mode and math.isfinite(fixed_terminal_soc_target):
            terminal_soc_mode = "fixed"
        if terminal_soc_mode:
            command.extend(["--terminal-soc-mode", terminal_soc_mode])
        if math.isfinite(fixed_terminal_soc_target):
            command.extend(["--fixed-terminal-soc-target", self._format_float(min(max(fixed_terminal_soc_target, 0.0), 1.0))])

        daily_terminal_soc_tolerance = self._safe_float(request.get("daily_terminal_soc_tolerance"), math.nan)
        if math.isfinite(daily_terminal_soc_tolerance):
            command.extend(["--daily-terminal-soc-tolerance", self._format_float(min(max(daily_terminal_soc_tolerance, 0.0), 0.20))])

        safety_tradeoff = self._safe_float(request.get("safety_economy_tradeoff"), math.nan)
        if math.isfinite(safety_tradeoff):
            command.extend(["--safety-economy-tradeoff", self._format_float(min(max(safety_tradeoff, 0.0), 1.0))])

        dss_master = task_workspace / "inputs" / "dss" / "visual_model" / "Master.dss"
        voltage_penalty_coeff = self._resolve_registry_target_float(
            registry_path=registry_path,
            target_id=target_id,
            column_name="voltage_penalty_coeff_yuan",
            default=0.0,
        )
        command.extend([
            "--enable-opendss-oracle",
            "--dss-master-path", str(dss_master.resolve()),
            "--opendss-voltage-penalty-coeff", self._format_float(voltage_penalty_coeff),
        ])
        return command

    def _resolve_single_registry_target_id(self, registry_path: Path) -> str | None:
        if not registry_path.exists():
            return None

        from openpyxl import load_workbook

        wb = load_workbook(registry_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        index = {name: idx for idx, name in enumerate(headers) if name}
        required = {"enabled", "optimize_storage", "internal_model_id"}
        if not required <= set(index):
            return None

        targets: list[str] = []
        for row in rows[1:]:
            if not row or all(value is None for value in row):
                continue
            enabled = self._safe_bool(row[index["enabled"]], False)
            optimize = self._safe_bool(row[index["optimize_storage"]], False)
            if enabled and optimize:
                target_id = str(row[index["internal_model_id"]] or "").strip()
                if target_id:
                    targets.append(target_id)

        if len(targets) == 1:
            return targets[0]
        if not targets:
            raise ValueError("未指定配储目标：请在一个负荷节点上设置 optimize_storage=1，其他负荷作为背景负荷。")
        raise ValueError(
            "当前工程按单个工商业用户单独配储，target_id 留空时只能有一个 enabled=1 且 optimize_storage=1 的负荷节点；"
            f"当前发现 {len(targets)} 个：{', '.join(targets)}。"
        )

    def _resolve_registry_target_float(
        self,
        registry_path: Path,
        target_id: str | None,
        column_name: str,
        default: float,
    ) -> float:
        if not registry_path.exists():
            return float(default)

        from openpyxl import load_workbook

        wb = load_workbook(registry_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        if not rows:
            return float(default)

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        index = {name: idx for idx, name in enumerate(headers) if name}
        value_idx = index.get(column_name)
        if value_idx is None:
            return float(default)

        target_idx = index.get("internal_model_id")
        fallback: float | None = None
        wanted = str(target_id or "").strip()
        for row in rows[1:]:
            if not row or all(value is None for value in row):
                continue
            raw_value = row[value_idx] if value_idx < len(row) else None
            parsed = self._safe_float(raw_value, math.nan)
            if math.isfinite(parsed) and fallback is None:
                fallback = parsed
            if wanted and target_idx is not None and target_idx < len(row):
                if str(row[target_idx] or "").strip() == wanted:
                    return max(0.0, parsed if math.isfinite(parsed) else float(default))

        return max(0.0, fallback if fallback is not None else float(default))

    def _local_solver_project_root(self) -> Path:
        return Path(__file__).resolve().parents[2] / "storage_engine_project"

    def _is_stale_bundled_solver_path(self, path: Path) -> bool:
        local_root = self._local_solver_project_root().resolve()
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        if resolved == local_root:
            return False
        return resolved.name == "storage_engine_project" and resolved.parent.name.startswith("storage_web_platform")

    def _resolve_solver_project_root(self, project_id: str) -> Path:
        config = self.get_solver_config(project_id)
        local_root = self._local_solver_project_root()
        configured = config.get("solver_project_root")
        if configured:
            configured_path = Path(str(configured))
            if local_root.exists() and self._is_stale_bundled_solver_path(configured_path):
                return local_root.resolve()
            return configured_path.resolve()
        return local_root

    def _resolve_python_executable(self, project_id: str, solver_project_root: Path | None = None) -> str:
        config = self.get_solver_config(project_id)
        configured = str(config.get("python_executable") or "").strip()
        if configured and configured.lower() not in {"python", "python.exe", "python3", "python3.exe"}:
            local_root = self._local_solver_project_root()
            configured_path = Path(configured)
            configured_python_root = configured_path
            for parent in configured_path.parents:
                if parent.name == "storage_engine_project":
                    configured_python_root = parent
                    break
            if local_root.exists() and self._is_stale_bundled_solver_path(configured_python_root):
                local_python = local_root / ".venv" / "Scripts" / "python.exe"
                if local_python.exists():
                    return str(local_python.resolve())
            return configured

        root = solver_project_root or self._resolve_solver_project_root(project_id)
        for candidate in [
            root / ".venv" / "Scripts" / "python.exe",
            root / ".venv" / "bin" / "python",
            root / "venv" / "Scripts" / "python.exe",
            root / "venv" / "bin" / "python",
        ]:
            if candidate.exists():
                return str(candidate.resolve())
        return sys.executable

    def _safe_int(self, value: Any, default: int) -> int:
        try:
            if value in (None, ""):
                return int(default)
            return int(float(value))
        except Exception:
            return int(default)

    def _safe_float(self, value: Any, default: float) -> float:
        try:
            if value in (None, ""):
                return float(default)
            parsed = float(value)
        except Exception:
            return float(default)
        return parsed if math.isfinite(parsed) else float(default)

    def _format_float(self, value: float) -> str:
        number = self._safe_float(value, 0.0)
        return f"{number:.12g}"

    def _safe_bool(self, value: Any, default: bool) -> bool:
        if value in (None, ""):
            return bool(default)
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(int(value))
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "on", "是", "启用"}:
            return True
        if text in {"0", "false", "no", "n", "off", "否", "停用"}:
            return False
        return bool(default)

    def _now(self) -> str:
        return datetime.now().isoformat(timespec="seconds")

    # ---------------------------------------------------------------------
    # 通用读文件
    # ---------------------------------------------------------------------
    @staticmethod
    def _safe_load_json(path: Path, default: Any = None) -> Any:
        if not path.exists():
            return default
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            try:
                return json.loads(path.read_text(encoding="utf-8-sig"))
            except Exception:
                return default

    @staticmethod
    def _decode_bytes(raw: bytes) -> Tuple[str, str]:
        candidates = [
            "utf-8",
            "utf-8-sig",
            "gb18030",
            "gbk",
            "cp936",
            "utf-16",
            "utf-16-le",
            "utf-16-be",
        ]
        for enc in candidates:
            try:
                return raw.decode(enc), enc
            except Exception:
                continue
        return raw.decode("latin-1", errors="replace"), "latin-1"

    def _read_text_file(self, path: Path) -> Dict[str, Any]:
        if not path.exists() or not path.is_file():
            return {
                "exists": False,
                "path": str(path),
                "text": "",
                "encoding": None,
                "size": 0,
            }
        raw = path.read_bytes()
        text, enc = self._decode_bytes(raw)
        return {
            "exists": True,
            "path": str(path),
            "text": text,
            "encoding": enc,
            "size": len(raw),
        }

    def _requested_generations(self, task: Dict[str, Any]) -> int:
        metadata = task.get("metadata")
        if isinstance(metadata, dict):
            request = metadata.get("run_request")
            if isinstance(request, dict):
                generations = self._safe_int(request.get("generations"), 0)
                if generations > 0:
                    return generations

        command = task.get("command")
        if isinstance(command, list):
            for index, token in enumerate(command):
                if str(token) == "--generations" and index + 1 < len(command):
                    generations = self._safe_int(command[index + 1], 0)
                    if generations > 0:
                        return generations
        return 0

    def _estimate_progress_from_task(self, task: Dict[str, Any], stdout_text: str) -> Dict[str, Any]:
        parsed = self._parse_stdout_progress(stdout_text, self._requested_generations(task))
        status = str(task.get("status") or "").lower()
        metadata = task.get("metadata") if isinstance(task.get("metadata"), dict) else {}
        stale_reason = metadata.get("auto_settle_reason") if isinstance(metadata, dict) else None
        if status == "completed":
            return {
                "percent": 100,
                "label": "求解完成",
                "detail": "结果文件已生成，可进入结果展示页查看。",
                "source": "backend_log_parser",
            }
        if status == "failed":
            if stale_reason:
                return {
                    "percent": max(parsed["percent"], 1),
                    "label": "运行中断",
                    "detail": str(task.get("message") or "求解器进程已不存在，任务已自动收敛为中断状态。"),
                    "source": "backend_stale_task_settlement",
                }
            return {
                "percent": max(parsed["percent"], 1),
                "label": "求解失败",
                "detail": parsed["detail"] or "请查看 stderr 日志定位失败原因。",
                "source": "backend_log_parser",
            }
        if status in {"cancelled", "canceled"}:
            if stale_reason:
                return {
                    "percent": max(parsed["percent"], 1),
                    "label": "运行已终止",
                    "detail": str(task.get("message") or "求解器进程已不存在，任务已收敛为终止状态。"),
                    "source": "backend_stale_task_settlement",
                }
            return {
                "percent": max(parsed["percent"], 1),
                "label": "运行已终止",
                "detail": "用户已终止求解进程。",
                "source": "backend_log_parser",
            }
        if status in {"cancelling", "canceling"}:
            return {
                "percent": max(parsed["percent"], 1),
                "label": "正在终止",
                "detail": "已发送终止请求，等待求解器进程退出。",
                "source": "backend_log_parser",
            }
        if status == "running":
            if parsed["percent"] > 0:
                return parsed
            return {
                "percent": 3,
                "label": "求解器已启动",
                "detail": "正在等待求解器输出进度日志。",
                "source": "backend_log_parser",
            }
        if parsed["percent"] > 0:
            return parsed
        return {
            "percent": 0,
            "label": str(task.get("status") or "等待运行"),
            "detail": str(task.get("message") or "暂无可解析进度。"),
            "source": "backend_log_parser",
        }

    def _parse_stdout_progress(self, stdout_text: str, requested_generations: int) -> Dict[str, Any]:
        """解析 stdout 日志估算求解进度。

        求解器实际执行阶段与进度权重分配：
          阶段 1 — 初始化（加载数据、构建 OpenDSS oracle）:  0% ~  5%
          阶段 2 — GA 迭代（fast_proxy 代表日评估）:          5% ~ 30%
          阶段 3 — GA 最终种群评估（额外一轮 evaluate）:     30% ~ 35%
          阶段 4 — full_recheck（365 天全年 OpenDSS 潮流）:  35% ~ 90%
          阶段 5 — 结果导出（写 CSV/JSON/图表）:             90% ~100%
        """
        if not stdout_text.strip():
            return {
                "percent": 0,
                "label": "等待日志",
                "detail": "stdout 暂无进度输出。",
                "source": "backend_log_parser",
            }
        if "已导出总体最优方案汇总" in stdout_text:
            return {
                "percent": 100,
                "label": "结果汇总已导出",
                "detail": "求解流程已完成。",
                "source": "backend_log_parser",
            }

        total_cases = self._last_number(stdout_text, r"共加载\s+(\d+)\s+个待优化场景")
        case_match = self._last_match(stdout_text, r"开始场景优化\s+\[(\d+)\/(\d+)\]")
        completed_cases = self._count_matches(stdout_text, r"场景完成：")
        generations_from_log = self._last_number(stdout_text, r"优化参数：总代数=(\d+)")
        generations = max(generations_from_log or requested_generations or 0, 1)
        iteration = min(self._last_number(stdout_text, r"优化迭代\s+(\d+)") or 0, generations)

        # full_recheck 阶段：匹配 "[年度运行] 进度 N/365"
        annual_match = self._last_match(stdout_text, r"年度运行[^\n]*进度\s+(\d+)\/365")
        annual_day = min(self._safe_int(annual_match[0], 0), 365) if annual_match else 0

        # fast_proxy 阶段：匹配 "[年度运行] 代表日 N/M"
        proxy_match = self._last_match(stdout_text, r"年度运行[^\n]*代表日\s+(\d+)\/(\d+)")
        proxy_current = self._safe_int(proxy_match[0], 0) if proxy_match else 0
        proxy_total = self._safe_int(proxy_match[1], 0) if proxy_match else 0

        # 检测是否已进入 full_recheck 阶段
        in_full_recheck = "full_recheck" in stdout_text and annual_day > 0
        # 检测是否已进入最终重校核（"对最终折中解执行全年重校核" 或 "调用 OpenDSS oracle 对最终折中解"）
        in_final_recheck = bool(re.search(r"对最终折中解执行全年重校核|调用 OpenDSS oracle 对最终折中解", stdout_text))
        # 检测结果导出阶段
        in_export = "场景完成：" in stdout_text and completed_cases > 0

        total = self._safe_int(case_match[1], 0) if case_match else total_cases
        current = self._safe_int(case_match[0], 0) if case_match else min(completed_cases + 1, total or 1)

        # --- 阶段 5：结果导出 (90-100%) ---
        if in_export and total and completed_cases >= total:
            return {
                "percent": 95,
                "label": "正在导出结果",
                "detail": f"全部 {total} 个场景已完成，正在写入结果文件。",
                "source": "backend_log_parser",
            }

        # --- 阶段 4：full_recheck (35-90%) ---
        if in_final_recheck and annual_day > 0:
            recheck_fraction = annual_day / 365.0
            percent = self._clamp_progress(35 + recheck_fraction * 55)
            return {
                "percent": percent,
                "label": "全年重校核（最耗时阶段）",
                "detail": f"OpenDSS 全年逐日潮流重校核 {annual_day}/365 天。",
                "source": "backend_log_parser",
            }

        if in_full_recheck and annual_day > 0 and iteration >= generations:
            recheck_fraction = annual_day / 365.0
            percent = self._clamp_progress(35 + recheck_fraction * 55)
            return {
                "percent": percent,
                "label": "全年重校核（最耗时阶段）",
                "detail": f"全年逐日重校核 {annual_day}/365 天。",
                "source": "backend_log_parser",
            }

        # --- 阶段 3：GA 最终种群评估 (30-35%) ---
        if iteration >= generations and not in_full_recheck and not in_final_recheck:
            # 迭代已完成但还没进入 full_recheck，说明在做最终种群评估
            detail = f"GA 迭代已完成 {generations} 代，正在评估最终种群。"
            if proxy_current > 0 and proxy_total > 0:
                proxy_fraction = proxy_current / proxy_total
                percent = self._clamp_progress(30 + proxy_fraction * 5)
                detail = f"最终种群评估，代表日 {proxy_current}/{proxy_total}。"
            else:
                percent = 32
            return {
                "percent": percent,
                "label": "评估最终种群",
                "detail": detail,
                "source": "backend_log_parser",
            }

        # --- 阶段 2：GA 迭代 (5-30%) ---
        if total and current:
            iteration_fraction = iteration / generations if iteration > 0 else 0
            # fast_proxy 代表日进度作为迭代内的细粒度指标
            proxy_fraction = (proxy_current / proxy_total) if proxy_current > 0 and proxy_total > 0 else 0
            # 迭代内进度：迭代完成比例 + 当前迭代内的代表日进度
            in_iteration_fraction = (iteration_fraction + proxy_fraction / generations) if iteration < generations else iteration_fraction
            in_case_fraction = min(in_iteration_fraction, 1.0)
            completed_fraction = completed_cases / total
            running_fraction = (max(current - 1, 0) + in_case_fraction) / total
            overall_fraction = max(completed_fraction, running_fraction)
            percent = self._clamp_progress(5 + overall_fraction * 25)
            if proxy_current > 0 and proxy_total > 0 and iteration > 0:
                detail = f"第 {current}/{total} 个场景，迭代 {iteration}/{generations}，代表日 {proxy_current}/{proxy_total}。"
            elif iteration > 0:
                detail = f"第 {current}/{total} 个场景，优化迭代 {iteration}/{generations}。"
            else:
                detail = f"第 {current}/{total} 个场景正在初始化。"
            return {
                "percent": percent,
                "label": "正在运行 GA 优化",
                "detail": detail,
                "source": "backend_log_parser",
            }

        if iteration > 0:
            iteration_fraction = iteration / generations
            return {
                "percent": self._clamp_progress(5 + iteration_fraction * 25),
                "label": "正在运行 GA 优化",
                "detail": f"优化迭代 {iteration}/{generations}。",
                "source": "backend_log_parser",
            }

        # --- 阶段 1：初始化 (0-5%) ---
        return {
            "percent": 5,
            "label": "求解器已启动",
            "detail": "已捕获 stdout 日志，正在解析后续进度。",
            "source": "backend_log_parser",
        }

    def _with_progress_hint(self, task: Dict[str, Any], stdout_text: str) -> Dict[str, Any]:
        task_copy = dict(task)
        metadata = dict(task_copy.get("metadata") or {})
        previous = task_copy.get("progress_hint")
        if not isinstance(previous, dict):
            previous = metadata.get("progress_hint") if isinstance(metadata.get("progress_hint"), dict) else {}
        estimated = self._estimate_progress_from_task(task_copy, stdout_text)
        previous_percent = self._safe_int(previous.get("percent"), 0) if isinstance(previous, dict) else 0
        if str(task_copy.get("status") or "").lower() == "completed":
            percent = 100
        else:
            percent = max(previous_percent, self._safe_int(estimated.get("percent"), 0))
        progress_hint = {
            "percent": percent,
            "label": str(estimated.get("label") or previous.get("label") or "等待运行"),
            "detail": str(estimated.get("detail") or previous.get("detail") or "暂无可解析进度。"),
            "source": str(estimated.get("source") or previous.get("source") or "backend_log_parser"),
        }
        task_copy["progress_hint"] = progress_hint
        metadata["progress_hint"] = progress_hint
        task_copy["metadata"] = metadata
        return task_copy

    def _persist_progress_hint(self, task: Dict[str, Any], stdout_text: str) -> Dict[str, Any]:
        updated = self._with_progress_hint(task, stdout_text)
        previous = task.get("progress_hint")
        if not isinstance(previous, dict):
            metadata = task.get("metadata")
            previous = metadata.get("progress_hint") if isinstance(metadata, dict) and isinstance(metadata.get("progress_hint"), dict) else {}
        if previous == updated.get("progress_hint"):
            return updated

        project_id = str(updated.get("project_id") or "").strip()
        task_id = str(updated.get("task_id") or "").strip()
        if project_id and task_id:
            self._write_task_files(self._task_dir(project_id, task_id), updated)
        return updated

    @staticmethod
    def _last_match(text: str, pattern: str) -> Optional[Tuple[str, ...]]:
        matches = list(re.finditer(pattern, text))
        if not matches:
            return None
        return matches[-1].groups()

    def _last_number(self, text: str, pattern: str) -> int:
        match = self._last_match(text, pattern)
        if not match:
            return 0
        return self._safe_int(match[0], 0)

    @staticmethod
    def _count_matches(text: str, pattern: str) -> int:
        return len(list(re.finditer(pattern, text)))

    @staticmethod
    def _clamp_progress(value: float) -> int:
        if not math.isfinite(value):
            return 0
        return max(0, min(int(round(value)), 99))

    # ---------------------------------------------------------------------
    # 任务状态
    # ---------------------------------------------------------------------
    def get_latest_task(self, project_id: str) -> Optional[Dict[str, Any]]:
        solver_runs_dir = self._solver_runs_dir(project_id)
        if not solver_runs_dir.exists():
            return None

        candidates: List[Path] = []
        candidates.extend(solver_runs_dir.glob("task_*/task_meta.json"))
        candidates.extend(solver_runs_dir.glob("task_*/state/task.json"))
        candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)

        for meta_file in candidates:
            data = self._safe_load_json(meta_file)
            if data:
                project_id_from_task = str(data.get("project_id") or project_id or "").strip()
                data = self._settle_stale_active_task(data, project_id_from_task)
                stdout_path = Path(str(data.get("stdout_log") or ""))
                stdout_info = self._read_text_file(stdout_path)
                return self._persist_progress_hint(data, stdout_info["text"])
        return None

    @staticmethod
    def _task_time_value(task: Dict[str, Any]) -> float:
        value = task.get("completed_at") or task.get("started_at") or 0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str) and value.strip():
            text = value.strip().replace("Z", "+00:00")
            try:
                return datetime.fromisoformat(text).timestamp()
            except ValueError:
                return 0.0
        return 0.0

    def _task_sort_key(self, task: Dict[str, Any]) -> Tuple[int, float]:
        has_completed_time = 1 if task.get("completed_at") else 0
        return has_completed_time, self._task_time_value(task)

    def _task_has_result_outputs(self, project_id: str, task: Dict[str, Any]) -> bool:
        task_id = task.get("task_id")
        if not task_id:
            return False
        return (self._task_dir(project_id, str(task_id)) / "solver_workspace" / "outputs" / "integrated_optimization").exists()

    def _read_task_metadata(self, task_id: str, project_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
        if project_id:
            for candidate in [
                self._task_dir(project_id, task_id) / "task_meta.json",
                self._task_dir(project_id, task_id) / "state" / "task.json",
            ]:
                data = self._safe_load_json(candidate)
                if data:
                    return self._settle_stale_active_task(data, project_id)
            return None

        if not self.data_root.exists():
            return None
        for project_dir in self.data_root.iterdir():
            if not project_dir.is_dir():
                continue
            for candidate in [
                project_dir / "solver_runs" / f"task_{task_id}" / "task_meta.json",
                project_dir / "solver_runs" / f"task_{task_id}" / "state" / "task.json",
            ]:
                data = self._safe_load_json(candidate)
                if data:
                    return self._settle_stale_active_task(data, str(data.get("project_id") or project_dir.name))
        return None

    def get_latest_result_task(self, project_id: str) -> Optional[Dict[str, Any]]:
        for task in self.list_tasks(project_id).get("tasks", []):
            if self._task_has_result_outputs(project_id, task):
                full_task = self._read_task_metadata(str(task.get("task_id")), project_id)
                return full_task or task
        return self.get_latest_task(project_id)

    def list_tasks(self, project_id: str) -> Dict[str, Any]:
        solver_runs_dir = self._solver_runs_dir(project_id)
        tasks_by_id: Dict[str, Dict[str, Any]] = {}
        if solver_runs_dir.exists():
            for task_dir in solver_runs_dir.glob("task_*"):
                if not task_dir.is_dir():
                    continue
                task_id = task_dir.name.removeprefix("task_")
                data = self._read_task_metadata(task_id=task_id, project_id=project_id)
                if data:
                    tasks_by_id[str(data.get("task_id") or task_id)] = data

        tasks = sorted(tasks_by_id.values(), key=self._task_sort_key, reverse=True)
        return {
            "project_id": project_id,
            "tasks": [self._task_brief(task, project_id=project_id) for task in tasks],
        }

    def get_task(self, task_id: str, project_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
        if project_id:
            for candidate in [
                self._task_dir(project_id, task_id) / "task_meta.json",
                self._task_dir(project_id, task_id) / "state" / "task.json",
            ]:
                data = self._safe_load_json(candidate)
                if data:
                    data = self._settle_stale_active_task(data, project_id)
                    stdout_path = Path(str(data.get("stdout_log") or ""))
                    stdout_info = self._read_text_file(stdout_path)
                    return self._persist_progress_hint(data, stdout_info["text"])
            return None

        if not self.data_root.exists():
            return None

        for project_dir in self.data_root.iterdir():
            if not project_dir.is_dir():
                continue
            for candidate in [
                project_dir / "solver_runs" / f"task_{task_id}" / "task_meta.json",
                project_dir / "solver_runs" / f"task_{task_id}" / "state" / "task.json",
            ]:
                data = self._safe_load_json(candidate)
                if data:
                    data = self._settle_stale_active_task(data, str(data.get("project_id") or project_dir.name))
                    stdout_path = Path(str(data.get("stdout_log") or ""))
                    stdout_info = self._read_text_file(stdout_path)
                    return self._persist_progress_hint(data, stdout_info["text"])
        return None

    # ---------------------------------------------------------------------
    # 日志
    # ---------------------------------------------------------------------
    def get_task_logs(self, task_id: str, project_id: Optional[str] = None) -> Dict[str, Any]:
        task = self.get_task(task_id=task_id, project_id=project_id)
        if not task:
            raise FileNotFoundError(f"未找到任务：{task_id}")

        stdout_path = Path(task.get("stdout_log") or "")
        stderr_path = Path(task.get("stderr_log") or "")

        stdout_info = self._read_text_file(stdout_path)
        stderr_info = self._read_text_file(stderr_path)

        task_copy = self._persist_progress_hint(task, stdout_info["text"])
        task_copy["stdout_text"] = stdout_info["text"]
        task_copy["stderr_text"] = stderr_info["text"]
        task_copy["stdout_encoding"] = stdout_info["encoding"]
        task_copy["stderr_encoding"] = stderr_info["encoding"]
        task_copy["stdout_size"] = stdout_info["size"]
        task_copy["stderr_size"] = stderr_info["size"]
        return task_copy

    # ---------------------------------------------------------------------
    # 结果目录与摘要
    # ---------------------------------------------------------------------
    @staticmethod
    def _flatten_file_tree(root: Path) -> List[Dict[str, Any]]:
        if not root.exists():
            return []
        items: List[Dict[str, Any]] = []
        for file_path in sorted([p for p in root.rglob("*") if p.is_file()]):
            items.append(
                {
                    "name": file_path.name,
                    "relative_path": file_path.relative_to(root).as_posix(),
                    "absolute_path": str(file_path),
                    "size_bytes": file_path.stat().st_size,
                    "suffix": file_path.suffix.lower(),
                }
            )
        return items

    def _resolve_result_roots(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        selected_task = self._read_task_metadata(task_id, project_id) if task_id else self.get_latest_result_task(project_id)
        if task_id and not selected_task:
            raise FileNotFoundError(f"未找到任务：{task_id}")

        roots: Dict[str, Any] = {}
        if selected_task:
            resolved_task_id = selected_task.get("task_id")
            if resolved_task_id:
                roots["selected_task"] = selected_task
                task_dir = self._task_dir(project_id, str(resolved_task_id))
                roots["task_root"] = task_dir
                roots["task_outputs"] = task_dir / "outputs"
                roots["solver_workspace"] = task_dir / "solver_workspace"
                roots["integrated_optimization"] = task_dir / "solver_workspace" / "outputs" / "integrated_optimization"
                roots["adapted_results"] = task_dir / "outputs" / "adapted_results"
        build_dir = self._build_dir(project_id)
        roots["build_root"] = build_dir
        roots["build_manifest"] = build_dir / "manifest"
        return roots

    def list_result_files(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        roots = self._resolve_result_roots(project_id, task_id)
        result_groups: Dict[str, List[Dict[str, Any]]] = {}
        flattened: List[Dict[str, Any]] = []

        for key in ["integrated_optimization", "adapted_results", "task_outputs"]:
            root = roots.get(key)
            if not root:
                continue
            files = self._flatten_file_tree(root)
            result_groups[key] = files
            for item in files:
                flattened.append({**item, "group": key})

        return {
            "project_id": project_id,
            "groups": result_groups,
            "files": flattened,
            "counts": {k: len(v) for k, v in result_groups.items()},
        }

    def get_summary(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        roots = self._resolve_result_roots(project_id, task_id)
        adapted_root = roots.get("adapted_results")
        integrated_root = roots.get("integrated_optimization")

        summary_rows = self._safe_load_json((adapted_root / "summary_rows.json") if adapted_root else Path("__missing__"), default=[])
        best_summary = self._safe_load_json((integrated_root / "best_result_summary.json") if integrated_root else Path("__missing__"), default={})
        overall_best = self._safe_load_json((integrated_root / "overall_best_schemes.json") if integrated_root else Path("__missing__"), default=[])

        return {
            "project_id": project_id,
            "summary_rows": summary_rows or [],
            "best_result_summary": best_summary or {},
            "overall_best_schemes": overall_best or [],
        }

    def get_project_summary(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        return {"success": True, **self.get_summary(project_id, task_id)}

    def get_result_charts(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        roots = self._resolve_result_roots(project_id, task_id)
        integrated_root = roots.get("integrated_optimization")
        latest_task = roots.get("selected_task") or {}
        warnings: List[str] = []
        self._last_network_topology_cache_diagnostics = None

        if integrated_root is None or not integrated_root.exists():
            return {
                "success": True,
                "project_id": project_id,
                "latest_task": self._task_brief(latest_task, project_id=project_id),
                "selected_case": None,
                "warnings": ["未找到 integrated_optimization 输出目录。"],
                "source_files": {},
                "charts": {},
                "diagnostics": {
                    "network_topology_cache": {
                        "status": "missing",
                        "reason": "integrated_optimization_not_found",
                    }
                },
            }

        summary = self.get_summary(project_id, task_id)
        case_dir = self._select_case_dir(integrated_root, summary)
        if case_dir is None:
            return {
                "success": True,
                "project_id": project_id,
                "latest_task": self._task_brief(latest_task, project_id=project_id),
                "selected_case": None,
                "warnings": ["未找到可解析的节点结果目录。"],
                "source_files": {},
                "charts": {},
                "diagnostics": {
                    "network_topology_cache": {
                        "status": "missing",
                        "reason": "case_dir_not_found",
                    }
                },
            }

        source_files = {
            "monthly_summary": str(case_dir / "best_monthly_summary.csv"),
            "hourly_operation": str(case_dir / "best_annual_hourly_operation.csv"),
            "cashflow": str(case_dir / "best_cashflow_table.csv"),
            "annual_summary": str(case_dir / "best_annual_summary.csv"),
            "financial_summary": str(case_dir / "best_financial_summary.csv"),
            "population_results": str(case_dir / "population_results.csv"),
            "archive_results": str(case_dir / "archive_results.csv"),
            "optimization_history": str(case_dir / "optimization_history.csv"),
            "best_result_summary": str(case_dir / "best_result_summary.json"),
            "configuration_report": str(case_dir / "configuration_report.json"),
            "operation_report": str(case_dir / "operation_report.json"),
            "financial_report": str(case_dir / "financial_report.json"),
            "network_impact_report": str(case_dir / "network_impact_report.json"),
            "run_health_report": str(case_dir / "run_health_report.json"),
            "bus_voltage_trace": str(case_dir / "best_bus_voltage_trace.csv"),
            "line_loading_trace": str(case_dir / "best_line_loading_trace.csv"),
            "network_loss_trace": str(case_dir / "best_network_loss_trace.csv"),
        }
        line_capacity = self._build_line_capacity_chart(project_id, latest_task)
        if line_capacity:
            source_files["line_capacity"] = "build/manifest/build_manifest.json"

        monthly_rows = self._read_csv_dicts(case_dir / "best_monthly_summary.csv")
        hourly_rows = self._read_csv_dicts(case_dir / "best_annual_hourly_operation.csv")
        cashflow_rows = self._read_csv_dicts(case_dir / "best_cashflow_table.csv")
        annual_summary = self._first_row(case_dir / "best_annual_summary.csv")
        financial_summary = self._first_row(case_dir / "best_financial_summary.csv")
        best_result_summary = self._safe_load_json(case_dir / "best_result_summary.json", default={})
        if not isinstance(best_result_summary, dict):
            best_result_summary = {}
        deliverables = self._load_result_deliverables(case_dir)
        population_rows = self._read_csv_dicts(case_dir / "population_results.csv")
        archive_rows = self._read_csv_dicts(case_dir / "archive_results.csv")
        history_rows = self._read_csv_dicts(case_dir / "optimization_history.csv")

        if not monthly_rows:
            warnings.append("未解析到月度收益表。")
        if not hourly_rows:
            warnings.append("未解析到年度逐时运行表。")
        if not cashflow_rows:
            warnings.append("未解析到现金流表。")

        operation = self._build_operation_charts(hourly_rows)
        feasibility = self._build_feasibility_diagnostics(best_result_summary, population_rows or archive_rows, history_rows)
        if feasibility["summary"].get("status") == "infeasible":
            warnings.append("当前推荐方案不是严格可行解，请重点查看可行性诊断。")
        charts = {
            "feasibility_diagnostics": feasibility,
            "monthly_revenue": self._build_monthly_revenue_chart(monthly_rows),
            "representative_day": operation["representative_day"],
            "daily_operation": operation["daily_operation"],
            "yearly_soc": operation["yearly_soc"],
            "cashflow": self._build_cashflow_chart(cashflow_rows, financial_summary),
            "capital_breakdown": self._build_capital_breakdown(financial_summary),
            "annual_value_breakdown": self._build_annual_value_breakdown(financial_summary),
            "financial_metrics": self._build_financial_metrics(financial_summary, annual_summary),
            "pareto": self._build_pareto_chart(population_rows or archive_rows),
            "optimization_history": self._build_history_chart(history_rows),
            "storage_impact": self._build_storage_impact_chart(operation["representative_day"].get("rows", [])),
            "network_constraints": self._build_network_constraint_charts(hourly_rows, monthly_rows, annual_summary),
            "line_capacity": line_capacity,
            "network_topology": self._build_network_topology_view(
                project_id=project_id,
                latest_task=latest_task,
                selected_case=case_dir.name,
                case_dir=case_dir,
                hourly_rows=hourly_rows,
                annual_summary=annual_summary,
                best_result_summary=best_result_summary,
            ),
            "deliverables": deliverables,
        }

        return {
            "success": True,
            "project_id": project_id,
            "latest_task": self._task_brief(latest_task, project_id=project_id),
            "selected_case": case_dir.name,
            "warnings": warnings,
            "source_files": source_files,
            "charts": charts,
            "diagnostics": self._build_result_diagnostics(case_dir, deliverables),
        }

    def get_project_results(self, project_id: str, task_id: Optional[str] = None) -> Dict[str, Any]:
        return {"success": True, **self.get_summary(project_id, task_id), **self.list_result_files(project_id, task_id)}

    def resolve_result_file_path(self, project_id: str, relative_path: str, group: Optional[str] = None, task_id: Optional[str] = None) -> Tuple[Path, str]:
        roots = self._resolve_result_roots(project_id, task_id)
        candidate_groups = [group] if group else ["integrated_optimization", "adapted_results", "task_outputs"]
        for g in candidate_groups:
            if not g:
                continue
            root = roots.get(g)
            if not root:
                continue
            root_resolved = root.resolve()
            candidate = (root_resolved / relative_path).resolve()
            if candidate.exists() and candidate.is_file() and (
                candidate == root_resolved or root_resolved in candidate.parents
            ):
                return candidate, g
        raise FileNotFoundError(f"结果文件不存在：{relative_path}")

    def get_result_file_preview(self, project_id: str, relative_path: str, group: Optional[str] = None, max_rows: int = 50, task_id: Optional[str] = None) -> Dict[str, Any]:
        file_path, group_name = self.resolve_result_file_path(project_id, relative_path, group, task_id)
        suffix = file_path.suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"}:
            return {
                "success": True,
                "project_id": project_id,
                "group": group_name,
                "relative_path": relative_path,
                "file_name": file_path.name,
                "type": "image",
            }
        if suffix == ".csv":
            header, body, total_rows = self._read_csv_preview(file_path, max_rows=max_rows)
            return {
                "success": True,
                "project_id": project_id,
                "group": group_name,
                "relative_path": relative_path,
                "file_name": file_path.name,
                "type": "csv",
                "header": header,
                "rows": body,
                "row_count": len(body),
                "total_rows": total_rows,
            }

        text_info = self._read_text_file(file_path)
        return {
            "success": True,
            "project_id": project_id,
            "group": group_name,
            "relative_path": relative_path,
            "file_name": file_path.name,
            "type": "text",
            "content": text_info["text"],
            "encoding": text_info["encoding"],
        }

    def _task_brief(self, task: Dict[str, Any], project_id: Optional[str] = None) -> Dict[str, Any]:
        brief = {
            "task_id": task.get("task_id"),
            "status": task.get("status"),
            "started_at": task.get("started_at"),
            "completed_at": task.get("completed_at"),
        }
        if project_id and brief.get("task_id"):
            brief.update(self._task_health_brief(project_id, str(brief["task_id"])))
        return brief

    def _task_health_brief(self, project_id: str, task_id: str) -> Dict[str, Any]:
        case_dir = self._first_task_case_dir(project_id, task_id)
        if case_dir is None:
            return {}
        report = self._safe_load_json(case_dir / "run_health_report.json", default={})
        if not isinstance(report, dict) or not report:
            return {}
        normalized = self._normalize_run_health_report(report)
        summary = normalized.get("summary") if isinstance(normalized.get("summary"), dict) else {}
        return {
            "health_status": normalized.get("status"),
            "health_issue_count": int(summary.get("issue_count") or 0),
            "health_warning_count": int(summary.get("warning_count") or 0),
            "health_critical_count": int(summary.get("critical_count") or summary.get("error_count") or 0),
        }

    def _first_task_case_dir(self, project_id: str, task_id: str) -> Optional[Path]:
        integrated_root = self._task_dir(project_id, task_id) / "solver_workspace" / "outputs" / "integrated_optimization"
        if not integrated_root.exists() or not integrated_root.is_dir():
            return None
        direct = integrated_root / "run_health_report.json"
        if direct.exists():
            return integrated_root
        for child in sorted(integrated_root.iterdir()):
            if child.is_dir() and (child / "run_health_report.json").exists():
                return child
        return None

    def _select_case_dir(self, integrated_root: Path, summary: Dict[str, Any]) -> Optional[Path]:
        candidates: List[str] = []
        for row in summary.get("overall_best_schemes") or []:
            if isinstance(row, dict):
                value = row.get("internal_model_id") or row.get("model_id") or row.get("load_id")
                if value:
                    candidates.append(str(value))
        best = summary.get("best_result_summary")
        if isinstance(best, dict):
            value = best.get("internal_model_id") or best.get("model_id") or best.get("load_id")
            if value:
                candidates.append(str(value))

        for name in candidates:
            path = integrated_root / name
            if path.exists() and path.is_dir():
                return path

        for path in sorted(integrated_root.iterdir()):
            if path.is_dir() and (path / "best_annual_hourly_operation.csv").exists():
                return path
        return None

    def _read_csv_dicts(self, path: Path) -> List[Dict[str, Any]]:
        return list(self._iter_csv_dicts(path))

    def _iter_csv_dicts(self, path: Path) -> Iterable[Dict[str, Any]]:
        if not path.exists():
            return
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                yield {str(k).strip().lstrip("\ufeff"): self._normalize_csv_value(v) for k, v in row.items() if k is not None}

    def _read_csv_preview(self, path: Path, max_rows: int = 50) -> Tuple[List[str], List[List[str]], int | None]:
        if not path.exists():
            return [], [], 0
        header: List[str] = []
        body: List[List[str]] = []
        total_rows = 0
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fp:
            reader = csv.reader(fp)
            for index, row in enumerate(reader):
                if index == 0:
                    header = row
                    continue
                total_rows += 1
                if len(body) < max_rows:
                    body.append(row)
                elif path.stat().st_size > 10 * 1024 * 1024:
                    return header, body, None
                else:
                    continue
        return header, body, total_rows

    def _load_result_deliverables(self, case_dir: Path) -> Dict[str, Any]:
        files = {
            "configuration": "configuration_report.json",
            "operation": "operation_report.json",
            "financial": "financial_report.json",
            "network_impact": "network_impact_report.json",
            "run_health": "run_health_report.json",
        }
        deliverables: Dict[str, Any] = {}
        for key, filename in files.items():
            data = self._safe_load_json(case_dir / filename, default={})
            if not isinstance(data, dict):
                deliverables[key] = {}
            elif key == "run_health":
                deliverables[key] = self._normalize_run_health_report(data)
            elif key == "network_impact":
                deliverables[key] = self._normalize_network_impact_report(data)
            elif key == "financial":
                deliverables[key] = self._normalize_financial_report(data)
            else:
                deliverables[key] = data
        return deliverables

    def _build_result_diagnostics(self, case_dir: Path, deliverables: Dict[str, Any]) -> Dict[str, Any]:
        run_health = deliverables.get("run_health") if isinstance(deliverables.get("run_health"), dict) else {}
        network_impact = deliverables.get("network_impact") if isinstance(deliverables.get("network_impact"), dict) else {}
        return {
            "network_topology_cache": self._last_network_topology_cache_diagnostics
            or {
                "status": "missing",
                "reason": "network_topology_not_built",
                "cache_file": str(case_dir / self.NETWORK_TOPOLOGY_SUMMARY_CACHE_FILE),
            },
            "run_health": {
                "status": run_health.get("status"),
                "issue_counts": run_health.get("issue_counts") or run_health.get("summary") or {},
            },
            "network_impact": {
                "data_quality": network_impact.get("data_quality") or {},
                "risk_classification_summary": network_impact.get("risk_classification_summary") or {},
            },
        }

    def _normalize_run_health_report(self, report: Dict[str, Any]) -> Dict[str, Any]:
        normalized = dict(report)
        raw_issues = report.get("issues") if isinstance(report.get("issues"), list) else []
        issues = [
            self._normalize_health_issue(item)
            for item in raw_issues
            if isinstance(item, dict)
        ]
        summary = dict(report.get("summary") or {}) if isinstance(report.get("summary"), dict) else {}
        critical_count = sum(1 for item in issues if item.get("level") == "critical")
        warning_count = sum(1 for item in issues if item.get("level") == "warning")
        if issues:
            issue_count = len(issues)
            error_count = critical_count
        else:
            issue_count = int(summary.get("issue_count") or summary.get("total") or 0)
            error_count = int(summary.get("error_count") or summary.get("critical_count") or 0)
            warning_count = int(summary.get("warning_count") or warning_count)
            critical_count = int(summary.get("critical_count") or error_count)
        status = self._health_status_level(report.get("status"), issues, summary)
        old_status = str(report.get("status") or "").strip().lower()
        summary.update({
            "issue_count": issue_count,
            "error_count": error_count,
            "warning_count": warning_count,
            "critical_count": critical_count,
        })
        normalized["status"] = status
        if old_status and old_status != status:
            normalized["legacy_status"] = old_status
        normalized["summary"] = summary
        normalized["issue_counts"] = {
            "total": issue_count,
            "warning": warning_count,
            "critical": critical_count,
            "error": error_count,
        }
        normalized["issues"] = issues
        return normalized

    def _normalize_health_issue(self, issue: Dict[str, Any]) -> Dict[str, Any]:
        item = dict(issue)
        code = str(item.get("code") or "").strip()
        severity = str(item.get("severity") or item.get("level") or "").strip().lower()
        if severity in {"critical", "error", "failed"}:
            level = "critical"
        elif severity == "warning":
            level = "warning"
        else:
            level = "warning"
        defaults = self._health_issue_defaults(code, str(item.get("message") or ""))
        item["level"] = level
        item.setdefault("reason", defaults["reason"])
        item.setdefault("impact", defaults["impact"])
        item.setdefault("suggestion", defaults["suggestion"])
        item.setdefault("related_section", defaults["related_section"])
        return item

    def _health_issue_defaults(self, code: str, message: str) -> Dict[str, str]:
        text = f"{code} {message}".lower()
        if "feasib" in text or "infeasible" in text:
            return {
                "reason": "候选方案可行性或最优方案可行性异常。",
                "impact": "储能配置方案可能不是严格可执行方案，经济性和运行曲线可信度下降。",
                "suggestion": "优先查看可行性诊断、约束罚分和候选方案状态，必要时收紧搜索空间后重跑。",
                "related_section": "feasibility",
            }
        if "soc" in text:
            return {
                "reason": "SOC 越界或能量守恒检查异常。",
                "impact": "年运行情况中的充放电轨迹和循环统计可能失真。",
                "suggestion": "检查储能容量、效率、初始 SOC、SOC 上下限和逐时调度约束。",
                "related_section": "operation",
            }
        if "opendss" in text or "network" in text or "voltage" in text or "line" in text:
            return {
                "reason": "OpenDSS 潮流、配网越限或网络约束检查异常。",
                "impact": "配网承载力变化结论可能不完整，需关注电压、线路和配变风险。",
                "suggestion": "检查 DSS 编译、目标接入母线、线路额定电流和潮流收敛小时。",
                "related_section": "network_impact",
            }
        if "economic" in text or "economics" in text or "npv" in text or "irr" in text:
            return {
                "reason": "经济性核心字段缺失、非有限值或异常值。",
                "impact": "经济性结论和审计账本不能直接作为可信收益判断。",
                "suggestion": "检查电价、收益开关、成本参数、退化和更换成本输入。",
                "related_section": "financial",
            }
        return {
            "reason": message or "健康检查发现异常。",
            "impact": "结果可信性需要结合相关模块进一步核对。",
            "suggestion": "查看 issue details 和关联结果文件，必要时重新运行求解。",
            "related_section": "run_health",
        }

    def _health_status_level(
        self,
        raw_status: Any,
        issues: List[Dict[str, Any]],
        summary: Dict[str, Any],
    ) -> str:
        text = str(raw_status or "").strip().lower()
        if text in {"critical", "failed", "error"}:
            return "critical"
        if text in {"warning", "passed"}:
            return text
        if any(item.get("level") == "critical" for item in issues) or int(summary.get("error_count") or 0) > 0:
            return "critical"
        if any(item.get("level") == "warning" for item in issues) or int(summary.get("warning_count") or 0) > 0:
            return "warning"
        return "passed"

    def _normalize_network_impact_report(self, report: Dict[str, Any]) -> Dict[str, Any]:
        normalized = dict(report)
        risk_details = dict(report.get("risk_details") or {}) if isinstance(report.get("risk_details"), dict) else {}
        normalized["risk_classification_summary"] = self._network_risk_classification_summary(risk_details)
        transformer_top_risks = self._network_transformer_top_risks(report, risk_details)
        normalized["transformer_top_risks"] = transformer_top_risks
        if transformer_top_risks:
            risk_details["transformer_top_risks"] = transformer_top_risks
        normalized["risk_details"] = risk_details
        normalized["target_area_conclusion"] = self._network_target_area_conclusion(report)
        normalized["attribution_summary"] = self._network_attribution_summary(report)
        return normalized

    def _network_risk_classification_summary(self, risk_details: Dict[str, Any]) -> Dict[str, Any]:
        classes = ["existing_background", "storage_induced", "worsened_by_storage", "improved_by_storage", "cleared_by_storage", "normal"]
        summary = {
            key: {
                "total": 0,
                "voltage": 0,
                "line": 0,
                "transformer": 0,
            }
            for key in classes
        }
        for source_key, bucket in [("voltage_classification_counts", "voltage"), ("line_classification_counts", "line")]:
            counts = risk_details.get(source_key) if isinstance(risk_details.get(source_key), dict) else {}
            for key, value in counts.items():
                class_key = str(key or "normal")
                if class_key not in summary:
                    summary[class_key] = {"total": 0, "voltage": 0, "line": 0, "transformer": 0}
                count = int(value or 0)
                summary[class_key][bucket] += count
                summary[class_key]["total"] += count
        transformer = risk_details.get("transformer") if isinstance(risk_details.get("transformer"), dict) else {}
        transformer_class = str(transformer.get("classification") or "normal")
        if transformer_class:
            summary.setdefault(transformer_class, {"total": 0, "voltage": 0, "line": 0, "transformer": 0})
            summary[transformer_class]["transformer"] += 1
            summary[transformer_class]["total"] += 1
        return {
            "items": [
                {"classification": key, **value}
                for key, value in summary.items()
                if value["total"] > 0 or key in classes
            ],
            "total_risks": sum(value["total"] for value in summary.values()),
        }

    def _network_transformer_top_risks(self, report: Dict[str, Any], risk_details: Dict[str, Any]) -> List[Dict[str, Any]]:
        transformer = risk_details.get("transformer") if isinstance(risk_details.get("transformer"), dict) else {}
        if not transformer:
            return []
        baseline_hours = self._finite_number(transformer.get("baseline_overload_hours"), 0.0)
        with_storage_hours = self._finite_number(transformer.get("with_storage_overload_hours"), 0.0)
        baseline = report.get("baseline") if isinstance(report.get("baseline"), dict) else {}
        with_storage = report.get("with_storage") if isinstance(report.get("with_storage"), dict) else {}
        return [
            {
                "transformer": transformer.get("transformer") or "目标上级配变",
                "classification": transformer.get("classification") or "normal",
                "baseline_overload_hours": baseline_hours,
                "with_storage_overload_hours": with_storage_hours,
                "overload_hour_delta": with_storage_hours - baseline_hours,
                "max_baseline_loading_pct": baseline.get("max_transformer_loading_pct"),
                "max_with_storage_loading_pct": with_storage.get("max_transformer_loading_pct"),
                "source": "network_impact_report.risk_details.transformer",
            }
        ]

    def _network_target_area_conclusion(self, report: Dict[str, Any]) -> Dict[str, Any]:
        target = report.get("target_connection") if isinstance(report.get("target_connection"), dict) else {}
        delta = target.get("delta") if isinstance(target.get("delta"), dict) else {}
        safety_delta = self._finite_number(delta.get("safety_violation_hours"), 0.0)
        voltage_delta = self._finite_number(delta.get("max_voltage_violation_pu"), 0.0)
        line_delta = self._finite_number(delta.get("max_line_loading_pct"), 0.0)
        if safety_delta < -1e-6 or voltage_delta < -1e-6 or line_delta < -1e-6:
            status = "worsened"
            conclusion = "该储能方案加剧目标接入区域风险，需重点复核目标节点电压、接入线路和上游配变。"
        elif safety_delta > 1e-6 or voltage_delta > 1e-6 or line_delta > 1e-6:
            status = "improved"
            conclusion = "该储能方案未加剧目标接入区域风险，并对部分安全越限指标有改善。"
        else:
            status = "neutral"
            conclusion = "该储能方案对目标接入区域风险影响不明显，主要风险来自原网背景或数据缺口。"
        return {
            "status": status,
            "conclusion": conclusion,
            "target_node": target.get("target_node"),
            "target_transformer": target.get("target_transformer"),
            "access_line": target.get("access_line"),
            "upstream_feeder": target.get("upstream_feeder"),
            "metrics": {
                "safety_violation_hours_delta": safety_delta,
                "max_voltage_violation_pu_delta": voltage_delta,
                "max_line_loading_pct_delta": line_delta,
            },
        }

    def _network_attribution_summary(self, report: Dict[str, Any]) -> Dict[str, Any]:
        delta = report.get("delta") if isinstance(report.get("delta"), dict) else {}
        target = report.get("target_connection") if isinstance(report.get("target_connection"), dict) else {}
        target_delta = target.get("delta") if isinstance(target.get("delta"), dict) else {}
        safety_delta = self._finite_number(delta.get("safety_violation_hours"), 0.0)
        loss_delta = self._finite_number(delta.get("loss_reduction_kwh"), 0.0)
        voltage_delta = self._finite_number(delta.get("max_voltage_violation_pu"), 0.0)
        line_delta = self._finite_number(delta.get("max_line_loading_pct"), 0.0)
        target_safety_delta = self._finite_number(target_delta.get("safety_violation_hours"), 0.0)
        return {
            "voltage": self._delta_text(voltage_delta, "电压越限幅度"),
            "line_loading": self._delta_text(line_delta, "线路最大负载率"),
            "target_area": self._delta_text(target_safety_delta, "目标区域安全越限小时"),
            "losses": self._delta_text(loss_delta, "网损"),
            "primary_drivers": [
                text
                for text in [
                    "削峰降低安全越限小时" if safety_delta > 1e-6 else "",
                    "局部反送或储能充放电加重越限" if safety_delta < -1e-6 else "",
                    "网损下降贡献经济收益" if loss_delta > 1e-6 else "",
                    "网损上升抵消部分收益" if loss_delta < -1e-6 else "",
                ]
                if text
            ],
        }

    @staticmethod
    def _delta_text(delta: float, metric_name: str) -> str:
        if delta > 1e-6:
            return f"{metric_name}改善，储后指标较储前降低。"
        if delta < -1e-6:
            return f"{metric_name}恶化，储后指标较储前升高。"
        return f"{metric_name}基本持平。"

    def _normalize_financial_report(self, report: Dict[str, Any]) -> Dict[str, Any]:
        normalized = dict(report)
        ledger = dict(report.get("annual_audit_ledger") or {}) if isinstance(report.get("annual_audit_ledger"), dict) else {}
        rows = ledger.get("items") if isinstance(ledger.get("items"), list) else []
        anomalies: List[Dict[str, Any]] = []
        for index, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            name = str(row.get("name") or f"item_{index}")
            category = str(row.get("category") or "")
            amount = self._finite_number(row.get("amount_yuan"), None)
            quantity = self._finite_number(row.get("quantity"), None)
            unit_price = self._finite_number(row.get("unit_price"), None)
            for field, value in [("amount_yuan", amount), ("quantity", quantity), ("unit_price", unit_price)]:
                if row.get(field) not in (None, "") and value is None:
                    anomalies.append({"item": name, "field": field, "level": "critical", "message": "字段不是有限数值。"})
            if amount is not None and abs(amount) > 1e11:
                anomalies.append({"item": name, "field": "amount_yuan", "level": "warning", "message": "年度金额异常大。"})
            if category in {"revenue", "subsidy"} and amount is not None and amount < 0:
                anomalies.append({"item": name, "field": "amount_yuan", "level": "warning", "message": "收益项为负，需确认是否代表收益损失。"})
            if amount is not None and quantity is not None and unit_price is not None:
                expected = quantity * unit_price
                tolerance = max(1.0, abs(amount) * 0.01)
                if abs(expected - amount) > tolerance:
                    anomalies.append({"item": name, "field": "amount_yuan", "level": "warning", "message": "年度金额与单价×数量偏差超过 1%。"})
        normalized["audit_ledger_summary"] = {
            "item_count": sum(1 for row in rows if isinstance(row, dict)),
            "anomaly_count": len(anomalies),
            "anomalies": anomalies,
        }
        return normalized

    @staticmethod
    def _finite_number(value: Any, default: float | None = 0.0) -> float | None:
        try:
            if value in (None, ""):
                return default
            number = float(value)
        except (TypeError, ValueError):
            return default
        return number if math.isfinite(number) else default

    def _first_row(self, path: Path) -> Dict[str, Any]:
        rows = self._read_csv_dicts(path)
        return rows[0] if rows else {}

    def _normalize_csv_value(self, value: Any) -> Any:
        if value is None:
            return None
        if not isinstance(value, str):
            return value
        text = value.strip()
        if text == "":
            return None
        if text.lower() in {"true", "false"}:
            return text.lower() == "true"
        try:
            number = float(text)
            if not math.isfinite(number):
                return None
            if number.is_integer():
                return int(number)
            return number
        except Exception:
            return text

    def _number(self, row: Dict[str, Any], key: str, default: float = 0.0) -> float:
        value = row.get(key)
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        try:
            number = float(value)
        except Exception:
            return default
        return number if math.isfinite(number) else default

    def _number_any(self, row: Dict[str, Any], keys: List[str], default: float = 0.0) -> float:
        for key in keys:
            if key in row and row.get(key) not in (None, ""):
                return self._number(row, key, default)
        return default

    def _optional_number(self, row: Dict[str, Any], key: str) -> Optional[float]:
        value = row.get(key)
        if value in (None, ""):
            return None
        try:
            number = float(value)
        except Exception:
            return None
        return number if math.isfinite(number) else None

    def _bool_value(self, value: Any) -> Optional[bool]:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            return bool(value)
        if isinstance(value, str):
            text = value.strip().lower()
            if text in {"true", "1", "yes", "y"}:
                return True
            if text in {"false", "0", "no", "n"}:
                return False
        return None

    def _build_feasibility_diagnostics(
        self,
        best_result: Dict[str, Any],
        candidate_rows: List[Dict[str, Any]],
        history_rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        best_feasible = self._bool_value(best_result.get("feasible"))
        best_valid = self._bool_value(best_result.get("is_valid"))

        feasible_count = 0
        infeasible_count = 0
        candidate_violations: List[Dict[str, Any]] = []
        for idx, row in enumerate(candidate_rows):
            feasible = self._bool_value(row.get("feasible"))
            if feasible is True:
                feasible_count += 1
            elif feasible is False:
                infeasible_count += 1
            candidate_violations.append(
                {
                    "index": idx + 1,
                    "feasible": bool(feasible),
                    "totalViolation": self._number(row, "total_violation"),
                    "cycleViolation": self._number(row, "cycle_violation"),
                    "durationViolationH": self._number(row, "duration_violation_h"),
                    "transformerViolationHours": self._number(row, "transformer_violation_hours"),
                    "negativeCashflowViolation": self._number(row, "negative_cashflow_violation"),
                    "paybackViolationYears": self._number(row, "payback_violation_years"),
                }
            )

        if not candidate_rows and best_feasible is not None:
            feasible_count = 1 if best_feasible else 0
            infeasible_count = 0 if best_feasible else 1

        latest_history = history_rows[-1] if history_rows else {}
        history_feasible_count = int(self._number(latest_history, "feasible_count", feasible_count))
        population_size = int(self._number(latest_history, "population_size", len(candidate_rows)))
        archive_size = int(self._number(latest_history, "archive_size", 0))

        constraint_items = [
            ("循环次数超限", "cycle_violation", "cycleViolation"),
            ("配置时长超限", "duration_violation_h", "durationViolationH"),
            ("变压器越限小时", "transformer_violation_hours", "transformerViolationHours"),
            ("负现金流违反", "negative_cashflow_violation", "negativeCashflowViolation"),
            ("回收期超限", "payback_violation_years", "paybackViolationYears"),
            ("最大违反量", "max_violation", "maxViolation"),
            ("总违反量", "total_violation", "totalViolation"),
        ]
        violations = [
            {"name": label, "key": camel_key, "value": self._number(best_result, source_key)}
            for label, source_key, camel_key in constraint_items
        ]

        if best_feasible is True:
            status = "feasible"
            message = "当前推荐方案满足可行性约束。"
        elif best_feasible is False:
            status = "infeasible"
            message = "当前推荐方案为最佳折中方案，但未满足全部可行性约束。"
        else:
            status = "unknown"
            message = "未解析到推荐方案可行性状态。"

        return {
            "summary": {
                "status": status,
                "message": message,
                "feasible": best_feasible,
                "isValid": best_valid,
                "feasibleCount": history_feasible_count,
                "populationSize": population_size,
                "archiveSize": archive_size,
                "candidateFeasibleCount": feasible_count,
                "candidateInfeasibleCount": infeasible_count,
                "bestTotalViolation": self._number(best_result, "total_violation"),
                "bestCycleViolation": self._number(best_result, "cycle_violation"),
            },
            "violations": violations,
            "candidate_status": [
                {"name": "可行方案", "count": feasible_count},
                {"name": "不可行方案", "count": infeasible_count},
            ],
            "candidate_violations": candidate_violations,
        }

    def _build_monthly_revenue_chart(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chart: List[Dict[str, Any]] = []
        for row in rows:
            service_net = (
                self._number(row, "service_capacity_revenue_yuan")
                + self._number(row, "service_delivery_revenue_yuan")
                - self._number(row, "service_penalty_yuan")
            )
            grid_cost = (
                self._number(row, "degradation_cost_yuan")
                + self._number(row, "transformer_penalty_yuan")
                + self._number(row, "voltage_penalty_yuan")
            )
            chart.append(
                {
                    "month": int(self._number(row, "month")),
                    "arbitrageRevenueWan": self._number(row, "arbitrage_revenue_yuan") / 10000.0,
                    "demandSavingWan": self._number(row, "demand_saving_yuan") / 10000.0,
                    "serviceNetRevenueWan": service_net / 10000.0,
                    "capacityRevenueWan": self._number(row, "capacity_revenue_yuan") / 10000.0,
                    "lossReductionRevenueWan": self._number(row, "loss_reduction_revenue_yuan") / 10000.0,
                    "penaltyCostWan": -grid_cost / 10000.0,
                    "netCashflowWan": self._number(row, "net_operating_cashflow_yuan") / 10000.0,
                }
            )
        return chart

    def _build_operation_charts(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        days: Dict[int, Dict[str, Any]] = {}
        for row in rows:
            day = int(self._number_any(row, ["day_index", "dayIndex", "day"], 0.0))
            hour = int(self._number_any(row, ["hour", "hour_index", "hourIndex"], 0.0))
            if day <= 0:
                day = int(hour // 24) + 1 if hour >= 24 else 1
                hour = int(hour % 24)
            hourly_cashflow = (
                self._number(row, "arbitrage_revenue_yuan")
                + self._number(row, "service_capacity_revenue_yuan")
                + self._number(row, "service_delivery_revenue_yuan")
                - self._number(row, "service_penalty_yuan")
                - self._number(row, "degradation_cost_yuan")
                - self._number(row, "transformer_penalty_yuan")
                - self._number(row, "voltage_penalty_yuan")
            )
            charge = self._number_any(row, ["exec_charge_kw", "charge_kw", "chargeKw"])
            discharge = self._number_any(row, ["exec_discharge_kw", "discharge_kw", "dischargeKw"])
            entry = {
                "hour": hour,
                "tariffYuanPerKwh": self._number_any(row, ["tariff_yuan_per_kwh", "tariffYuanPerKwh"]),
                "actualNetLoadKw": self._number_any(row, ["actual_net_load_kw", "actualNetLoadKw"]),
                "gridExchangeKw": self._number_any(row, ["grid_exchange_kw", "gridExchangeKw"]),
                "chargeKw": -charge,
                "dischargeKw": discharge,
                "socOpen": self._number_any(row, ["soc_open", "socOpen", "SOC_open"]),
                "socClose": self._number_any(row, ["soc_close", "socClose", "SOC_close", "soc"]),
                "netCashflowYuan": hourly_cashflow,
            }
            day_bucket = days.setdefault(
                day,
                {
                    "rows": [],
                    "throughputKwh": 0.0,
                    "netCashflowYuan": 0.0,
                    "socOpen": entry["socOpen"],
                    "socClose": entry["socClose"],
                },
            )
            day_bucket["rows"].append(entry)
            day_bucket["throughputKwh"] += charge + discharge
            day_bucket["netCashflowYuan"] += hourly_cashflow
            day_bucket["socClose"] = entry["socClose"]

        if not days:
            return {"representative_day": {"dayIndex": None, "rows": []}, "daily_operation": [], "yearly_soc": []}

        representative_day = max(days.items(), key=lambda item: item[1]["throughputKwh"])[0]
        daily_rows = [
            {
                "dayIndex": day,
                "throughputKwh": bucket["throughputKwh"],
                "netCashflowYuan": bucket["netCashflowYuan"],
                "socOpen": bucket["socOpen"],
                "socClose": bucket["socClose"],
            }
            for day, bucket in sorted(days.items())
        ]
        self._add_moving_average(daily_rows, "throughputKwh", "throughputMa7Kwh")
        self._add_moving_average(daily_rows, "netCashflowYuan", "netCashflowMa7Yuan")

        representative_rows = sorted(days[representative_day]["rows"], key=lambda item: item["hour"])
        yearly_soc_rows = [
            {
                "hourOfYear": (day - 1) * 24 + int(row["hour"]),
                "dayIndex": day,
                "hour": int(row["hour"]),
                "socOpen": row["socOpen"],
                "socClose": row["socClose"],
                "chargeKw": row["chargeKw"],
                "dischargeKw": row["dischargeKw"],
            }
            for day, bucket in sorted(days.items())
            for row in sorted(bucket["rows"], key=lambda item: item["hour"])
        ]
        return {
            "representative_day": {"dayIndex": representative_day, "rows": representative_rows},
            "daily_operation": daily_rows,
            "yearly_soc": yearly_soc_rows,
        }

    def _build_storage_impact_chart(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chart: List[Dict[str, Any]] = []
        for row in rows:
            actual = self._number(row, "actualNetLoadKw")
            grid = self._number(row, "gridExchangeKw")
            charge = abs(self._number(row, "chargeKw"))
            discharge = self._number(row, "dischargeKw")
            chart.append(
                {
                    "hour": int(self._number(row, "hour")),
                    "actualNetLoadKw": actual,
                    "gridExchangeKw": grid,
                    "chargeKw": -charge,
                    "dischargeKw": discharge,
                    "storageNetKw": grid - actual,
                    "peakShavingKw": actual - grid,
                    "tariffYuanPerKwh": self._number(row, "tariffYuanPerKwh"),
                }
            )
        return chart

    def _build_network_constraint_charts(
        self,
        hourly_rows: List[Dict[str, Any]],
        monthly_rows: List[Dict[str, Any]],
        annual_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        by_day: Dict[int, Dict[str, Any]] = {}
        has_opendss_loss = False
        annual_opendss_loss_reduction_kwh = 0.0
        for row in hourly_rows:
            day = int(self._number(row, "day_index"))
            bucket = by_day.setdefault(
                day,
                {
                    "dayIndex": day,
                    "transformerPenaltyWan": 0.0,
                    "voltagePenaltyWan": 0.0,
                    "penaltyHours": 0,
                    "maxGridExchangeKw": 0.0,
                    "maxActualNetLoadKw": 0.0,
                    "opendssLossReductionKwh": 0.0,
                },
            )
            transformer_penalty = self._number(row, "transformer_penalty_yuan")
            voltage_penalty = self._number(row, "voltage_penalty_yuan")
            if row.get("opendss_loss_reduction_kwh") not in (None, ""):
                loss_reduction_kwh = self._number(row, "opendss_loss_reduction_kwh")
                bucket["opendssLossReductionKwh"] += loss_reduction_kwh
                annual_opendss_loss_reduction_kwh += loss_reduction_kwh
                has_opendss_loss = True
            bucket["transformerPenaltyWan"] += transformer_penalty / 10000.0
            bucket["voltagePenaltyWan"] += voltage_penalty / 10000.0
            if transformer_penalty > 0 or voltage_penalty > 0:
                bucket["penaltyHours"] += 1
            bucket["maxGridExchangeKw"] = max(bucket["maxGridExchangeKw"], self._number(row, "grid_exchange_kw"))
            bucket["maxActualNetLoadKw"] = max(bucket["maxActualNetLoadKw"], self._number(row, "actual_net_load_kw"))

        monthly = [
            {
                "month": int(self._number(row, "month")),
                "transformerPenaltyWan": self._number(row, "transformer_penalty_yuan") / 10000.0,
                "voltagePenaltyWan": self._number(row, "voltage_penalty_yuan") / 10000.0,
                "totalPenaltyWan": (
                    self._number(row, "transformer_penalty_yuan")
                    + self._number(row, "voltage_penalty_yuan")
                ) / 10000.0,
            }
            for row in monthly_rows
        ]
        summary = [
            {"name": "储前安全违约小时", "value": self._number(annual_summary, "baseline_safety_violation_hours"), "unit": "h"},
            {"name": "储后安全违约小时", "value": self._number(annual_summary, "storage_safety_violation_hours"), "unit": "h"},
            {"name": "安全违约小时改善", "value": self._number(annual_summary, "delta_safety_violation_hours"), "unit": "h"},
            {"name": "储前电压越限小时", "value": self._number(annual_summary, "baseline_hours_with_voltage_violation"), "unit": "h"},
            {"name": "储后电压越限小时", "value": self._number(annual_summary, "hours_with_voltage_violation"), "unit": "h"},
            {"name": "储前线路过载小时", "value": self._number(annual_summary, "baseline_hours_with_line_overload"), "unit": "h"},
            {"name": "储后线路过载小时", "value": self._number(annual_summary, "hours_with_line_overload"), "unit": "h"},
            {"name": "变压器越限小时", "value": self._number(annual_summary, "transformer_violation_hours"), "unit": "h"},
            {"name": "最大变压器裕度缺口", "value": self._number(annual_summary, "max_transformer_slack_kw"), "unit": "kW"},
            {"name": "年变压器罚金", "value": self._number(annual_summary, "annual_transformer_penalty_yuan") / 10000.0, "unit": "万元"},
            {"name": "年电压罚金", "value": self._number(annual_summary, "annual_voltage_penalty_yuan") / 10000.0, "unit": "万元"},
        ]
        if has_opendss_loss:
            summary.append(
                {
                    "name": "OpenDSS年网损差",
                    "value": annual_opendss_loss_reduction_kwh,
                    "unit": "kWh",
                }
            )
        return {
            "daily": [by_day[key] for key in sorted(by_day)],
            "monthly": monthly,
            "summary": summary,
        }

    def _build_line_capacity_chart(self, project_id: str, latest_task: Dict[str, Any]) -> List[Dict[str, Any]]:
        rows = self._line_summary_from_manifest(project_id, latest_task)
        if not rows:
            rows = self._line_summary_from_dss_files(project_id, latest_task)
        chart: List[Dict[str, Any]] = []
        for row in rows:
            normamps = self._number(row, "normamps")
            emergamps = self._number(row, "emergamps", normamps * 1.25)
            chart.append(
                {
                    "name": str(row.get("name") or row.get("id") or ""),
                    "lineId": str(row.get("id") or row.get("name") or ""),
                    "linecode": str(row.get("linecode") or ""),
                    "fromBus": str(row.get("from_bus") or row.get("bus1") or ""),
                    "toBus": str(row.get("to_bus") or row.get("bus2") or ""),
                    "lengthKm": self._number(row, "length_km"),
                    "normamps": normamps,
                    "emergamps": emergamps,
                    "enabled": bool(row.get("enabled", True)),
                    "autoServiceLine": bool(row.get("auto_service_line")),
                    "serviceSecondaryKv": self._number(row, "service_secondary_kv"),
                    "serviceTransformerKva": self._number(row, "service_transformer_kva"),
                    "serviceResourceKva": self._number(row, "service_resource_kva"),
                    "serviceTransformerCurrentA": self._number(row, "service_transformer_current_a"),
                    "serviceResourceCurrentA": self._number(row, "service_resource_current_a"),
                    "serviceEquivalentMode": str(row.get("service_equivalent_mode") or ""),
                    "serviceCableName": str(row.get("service_cable_name") or ""),
                    "serviceCableParallel": self._number(row, "service_cable_parallel"),
                    "serviceEquivalentR1OhmPerKm": self._number(row, "service_equivalent_r1_ohm_per_km"),
                    "serviceEquivalentX1OhmPerKm": self._number(row, "service_equivalent_x1_ohm_per_km"),
                    "lineVoltageKv": self._number(row, "line_voltage_kv"),
                    "downstreamTransformerKva": self._number(row, "downstream_transformer_kva"),
                    "downstreamLoadKva": self._number(row, "downstream_load_kva"),
                    "downstreamApparentKva": self._number(row, "downstream_apparent_kva"),
                    "estimatedRequiredCurrentA": self._number(row, "estimated_required_current_a"),
                    "recommendedCurrentA": self._number(row, "recommended_current_a"),
                    "recommendedLinecode": str(row.get("recommended_linecode") or ""),
                    "capacityCheckStatus": str(row.get("capacity_check_status") or ""),
                    "capacityCheckMessage": str(row.get("capacity_check_message") or ""),
                }
            )
        return chart

    def _line_summary_from_manifest(self, project_id: str, latest_task: Dict[str, Any]) -> List[Dict[str, Any]]:
        candidates: List[Path] = [self._build_dir(project_id) / "manifest" / "build_manifest.json"]
        metadata = latest_task.get("metadata") if isinstance(latest_task.get("metadata"), dict) else {}
        build_manifest = metadata.get("build_manifest") if isinstance(metadata, dict) else None
        if build_manifest:
            candidates.append(Path(str(build_manifest)))

        for manifest_path in candidates:
            manifest = self._safe_load_json(manifest_path, default={})
            if not isinstance(manifest, dict):
                continue
            summary = manifest.get("dss_compile_summary")
            if isinstance(summary, dict) and isinstance(summary.get("line_summary"), list):
                return [row for row in summary["line_summary"] if isinstance(row, dict)]
        return []

    def _line_summary_from_dss_files(self, project_id: str, latest_task: Dict[str, Any]) -> List[Dict[str, Any]]:
        candidates: List[Path] = []
        manifest = self._safe_load_json(self._build_dir(project_id) / "manifest" / "build_manifest.json", default={})
        if isinstance(manifest, dict) and manifest.get("dss_dir"):
            candidates.append(Path(str(manifest["dss_dir"])) / "Lines_Main.dss")
            candidates.append(Path(str(manifest["dss_dir"])) / "Lines.dss")

        workspace = latest_task.get("solver_workspace")
        if workspace:
            root = Path(str(workspace))
            candidates.extend(
                [
                    root / "inputs" / "dss" / "visual_model" / "Lines_Main.dss",
                    root / "inputs" / "dss" / "visual_model" / "Lines.dss",
                    root / "inputs" / "dss" / "ieee33" / "Lines_Main.dss",
                    root / "inputs" / "dss" / "ieee33" / "Lines.dss",
                ]
            )

        for path in candidates:
            rows = self._parse_dss_lines(path)
            if rows:
                return rows
        return []

    def _parse_dss_lines(self, path: Path) -> List[Dict[str, Any]]:
        if not path.exists():
            return []
        text_info = self._read_text_file(path)
        rows: List[Dict[str, Any]] = []
        for raw_line in text_info["text"].splitlines():
            line = raw_line.strip()
            if not line.lower().startswith("new line."):
                continue
            match = re.match(r"New\s+Line\.([^\s]+)\s+(.*)$", line, flags=re.IGNORECASE)
            if not match:
                continue
            line_id = match.group(1)
            attrs: Dict[str, Any] = {"id": line_id, "name": line_id}
            for token in match.group(2).split():
                if "=" not in token:
                    continue
                key, value = token.split("=", 1)
                key = key.strip().lower()
                value = value.strip()
                if key in {"length", "normamps", "emergamps", "phases"}:
                    attrs[{"length": "length_km"}.get(key, key)] = self._normalize_csv_value(value)
                elif key == "bus1":
                    attrs["bus1"] = value.split(".")[0]
                elif key == "bus2":
                    attrs["bus2"] = value.split(".")[0]
                elif key == "enabled":
                    attrs["enabled"] = value.lower() not in {"no", "false", "0"}
                else:
                    attrs[key] = value
            rows.append(attrs)
        return rows

    def _metric_key(self, value: Any) -> str:
        text = str(value or "").strip().lower()
        if text.startswith("line."):
            text = text[5:]
        return self._safe_name(text).lower()

    def _trace_voltage_pu(self, value: Optional[float], bus: str, bus_voltage_base_kv: Dict[str, float]) -> Optional[float]:
        if value is None:
            return None
        number = float(value)
        if not math.isfinite(number):
            return None
        abs_value = abs(number)
        if 1.45 <= abs_value <= 1.90:
            return number / math.sqrt(3.0)
        if abs_value <= 2.0:
            return number
        base_kv = bus_voltage_base_kv.get(self._metric_key(bus))
        if base_kv and base_kv > 0:
            normalized = number / (base_kv * 1000.0)
            if math.isfinite(normalized) and abs(normalized) <= 2.0:
                return normalized
        return None

    @staticmethod
    def _voltage_violation_amount_pu(v_min: Optional[float], v_max: Optional[float]) -> float:
        undervoltage = 0.0 if v_min is None else max(0.0, 0.93 - float(v_min))
        overvoltage = 0.0 if v_max is None else max(0.0, float(v_max) - 1.07)
        return float(undervoltage + overvoltage)

    def _summarize_bus_voltage_trace(
        self,
        rows: Iterable[Dict[str, Any]],
        bus_voltage_base_kv: Optional[Dict[str, float]] = None,
    ) -> Dict[str, Dict[str, Any]]:
        base_kv_by_bus = bus_voltage_base_kv or {}
        summary: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            bus = str(row.get("bus") or "").strip()
            if not bus:
                continue
            base_min = self._trace_voltage_pu(self._optional_number(row, "baseline_voltage_pu_min"), bus, base_kv_by_bus)
            base_max = self._trace_voltage_pu(self._optional_number(row, "baseline_voltage_pu_max"), bus, base_kv_by_bus)
            v_min = self._trace_voltage_pu(self._optional_number(row, "voltage_pu_min"), bus, base_kv_by_bus)
            v_max = self._trace_voltage_pu(self._optional_number(row, "voltage_pu_max"), bus, base_kv_by_bus)
            incremental_violation = None
            if v_min is not None or v_max is not None or base_min is not None or base_max is not None:
                incremental_violation = max(
                    0.0,
                    self._voltage_violation_amount_pu(v_min, v_max)
                    - self._voltage_violation_amount_pu(base_min, base_max),
                )
            elif self._optional_number(row, "storage_voltage_violation_increment_pu") is not None:
                incremental_violation = self._optional_number(row, "storage_voltage_violation_increment_pu")
            if v_min is None and v_max is None and base_min is None and base_max is None:
                continue
            item = summary.setdefault(
                self._metric_key(bus),
                {
                    "bus": bus,
                    "baselineVoltagePuMin": None,
                    "baselineVoltagePuMax": None,
                    "voltagePuMin": None,
                    "voltagePuMax": None,
                    "storageVoltageViolationIncrementPu": None,
                    "hourCount": 0,
                    "baselineUndervoltageHours": 0,
                    "baselineOvervoltageHours": 0,
                    "storageWorsenedVoltageHours": 0,
                    "undervoltageHours": 0,
                    "overvoltageHours": 0,
                },
            )
            item["hourCount"] += 1
            if base_min is not None:
                item["baselineVoltagePuMin"] = base_min if item["baselineVoltagePuMin"] is None else min(float(item["baselineVoltagePuMin"]), base_min)
                if base_min < 0.93:
                    item["baselineUndervoltageHours"] += 1
            if base_max is not None:
                item["baselineVoltagePuMax"] = base_max if item["baselineVoltagePuMax"] is None else max(float(item["baselineVoltagePuMax"]), base_max)
                if base_max > 1.07:
                    item["baselineOvervoltageHours"] += 1
            if v_min is not None:
                item["voltagePuMin"] = v_min if item["voltagePuMin"] is None else min(float(item["voltagePuMin"]), v_min)
                if v_min < 0.93:
                    item["undervoltageHours"] += 1
            if v_max is not None:
                item["voltagePuMax"] = v_max if item["voltagePuMax"] is None else max(float(item["voltagePuMax"]), v_max)
                if v_max > 1.07:
                    item["overvoltageHours"] += 1
            if incremental_violation is not None:
                item["storageVoltageViolationIncrementPu"] = (
                    incremental_violation
                    if item["storageVoltageViolationIncrementPu"] is None
                    else max(float(item["storageVoltageViolationIncrementPu"]), incremental_violation)
                )
                if incremental_violation > 1e-9:
                    item["storageWorsenedVoltageHours"] += 1
        return summary

    def _summarize_line_loading_trace(self, rows: Iterable[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        summary: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            line = str(row.get("line") or row.get("name") or "").strip()
            if not line:
                continue
            current_a = self._optional_number(row, "current_a")
            loading_pct = self._optional_number(row, "loading_pct")
            normamps = self._optional_number(row, "normamps")
            terminal1_power_kw = self._optional_number(row, "terminal1_power_kw")
            if loading_pct is None and current_a is not None and normamps and normamps > 0:
                loading_pct = current_a / normamps * 100.0
            if current_a is None and loading_pct is None:
                continue
            item = summary.setdefault(
                self._metric_key(line),
                {
                    "line": line,
                    "currentA": None,
                    "loadingPct": None,
                    "normamps": normamps,
                    "emergamps": self._optional_number(row, "emergamps"),
                    "terminal1PowerKw": None,
                    "flowDirection": "forward",
                    "hourCount": 0,
                    "overloadHours": 0,
                },
            )
            item["hourCount"] += 1
            if current_a is not None:
                item["currentA"] = current_a if item["currentA"] is None else max(float(item["currentA"]), current_a)
            if loading_pct is not None:
                item["loadingPct"] = loading_pct if item["loadingPct"] is None else max(float(item["loadingPct"]), loading_pct)
                if loading_pct > 100.0:
                    item["overloadHours"] += 1
            if item.get("normamps") is None and normamps is not None:
                item["normamps"] = normamps
            emergamps = self._optional_number(row, "emergamps")
            if item.get("emergamps") is None and emergamps is not None:
                item["emergamps"] = emergamps
            if terminal1_power_kw is not None and (
                item.get("terminal1PowerKw") is None
                or abs(terminal1_power_kw) > abs(float(item.get("terminal1PowerKw") or 0.0))
            ):
                item["terminal1PowerKw"] = terminal1_power_kw
                item["flowDirection"] = "reverse" if terminal1_power_kw < -1e-9 else "forward"
        return summary

    @staticmethod
    def _trace_file_signature(path: Path) -> Dict[str, Any]:
        if not path.exists() or not path.is_file():
            return {"exists": False, "size": 0, "mtime_ns": None}
        stat = path.stat()
        return {
            "exists": True,
            "size": int(stat.st_size),
            "mtime_ns": int(stat.st_mtime_ns),
        }

    def _bus_voltage_base_fingerprint(self, bus_voltage_base_kv: Dict[str, float]) -> List[List[Any]]:
        fingerprint: List[List[Any]] = []
        for bus, value in sorted((bus_voltage_base_kv or {}).items()):
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(number) and number > 0:
                fingerprint.append([self._metric_key(bus), round(number, 9)])
        return fingerprint

    def _network_topology_trace_cache_metadata(
        self,
        case_dir: Path,
        bus_voltage_base_kv: Dict[str, float],
    ) -> Dict[str, Any]:
        return {
            "version": self.NETWORK_TOPOLOGY_SUMMARY_CACHE_VERSION,
            "bus_voltage_trace": self._trace_file_signature(case_dir / "best_bus_voltage_trace.csv"),
            "line_loading_trace": self._trace_file_signature(case_dir / "best_line_loading_trace.csv"),
            "bus_voltage_base_kv": self._bus_voltage_base_fingerprint(bus_voltage_base_kv),
        }

    @staticmethod
    def _network_topology_cache_diagnostics_metadata(metadata: Dict[str, Any]) -> Dict[str, Any]:
        base_fingerprint = metadata.get("bus_voltage_base_kv") if isinstance(metadata.get("bus_voltage_base_kv"), list) else []
        return {
            "version": metadata.get("version"),
            "bus_voltage_trace": metadata.get("bus_voltage_trace"),
            "line_loading_trace": metadata.get("line_loading_trace"),
            "bus_voltage_base_kv_count": len(base_fingerprint),
        }

    def _read_network_topology_trace_summary_cache(
        self,
        case_dir: Path,
        expected_metadata: Dict[str, Any],
    ) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]] | None:
        cache_path = case_dir / self.NETWORK_TOPOLOGY_SUMMARY_CACHE_FILE
        cache = self._safe_load_json(cache_path, default=None)
        if not isinstance(cache, dict) or cache.get("metadata") != expected_metadata:
            return None
        bus_voltage_summary = cache.get("bus_voltage_summary")
        line_loading_summary = cache.get("line_loading_summary")
        if not isinstance(bus_voltage_summary, dict) or not isinstance(line_loading_summary, dict):
            return None
        return bus_voltage_summary, line_loading_summary

    def _write_network_topology_trace_summary_cache(
        self,
        case_dir: Path,
        metadata: Dict[str, Any],
        bus_voltage_summary: Dict[str, Dict[str, Any]],
        line_loading_summary: Dict[str, Dict[str, Any]],
    ) -> None:
        cache_path = case_dir / self.NETWORK_TOPOLOGY_SUMMARY_CACHE_FILE
        payload = {
            "metadata": metadata,
            "bus_voltage_summary": bus_voltage_summary,
            "line_loading_summary": line_loading_summary,
            "created_at": self._now(),
        }
        try:
            cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            return

    def _load_or_build_network_topology_trace_summaries(
        self,
        case_dir: Path,
        bus_voltage_base_kv: Dict[str, float],
    ) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
        metadata = self._network_topology_trace_cache_metadata(case_dir, bus_voltage_base_kv)
        cache_path = case_dir / self.NETWORK_TOPOLOGY_SUMMARY_CACHE_FILE
        cache_exists = cache_path.exists()
        source_trace_missing = not metadata["bus_voltage_trace"]["exists"] or not metadata["line_loading_trace"]["exists"]
        cached = self._read_network_topology_trace_summary_cache(case_dir, metadata)
        if cached is not None:
            self._last_network_topology_cache_diagnostics = {
                "status": "hit",
                "reason": "metadata_match",
                "cache_file": str(cache_path),
                "metadata": self._network_topology_cache_diagnostics_metadata(metadata),
            }
            return cached

        bus_voltage_summary = self._summarize_bus_voltage_trace(
            self._iter_csv_dicts(case_dir / "best_bus_voltage_trace.csv"),
            bus_voltage_base_kv,
        )
        line_loading_summary = self._summarize_line_loading_trace(
            self._iter_csv_dicts(case_dir / "best_line_loading_trace.csv")
        )
        self._write_network_topology_trace_summary_cache(
            case_dir,
            metadata,
            bus_voltage_summary,
            line_loading_summary,
        )
        self._last_network_topology_cache_diagnostics = {
            "status": "missing" if source_trace_missing else "rebuilt",
            "reason": "source_trace_missing" if source_trace_missing else ("invalid_cache" if cache_exists else "missing_cache"),
            "cache_file": str(cache_path),
            "metadata": self._network_topology_cache_diagnostics_metadata(metadata),
            "bus_voltage_summary_count": len(bus_voltage_summary),
            "line_loading_summary_count": len(line_loading_summary),
        }
        return bus_voltage_summary, line_loading_summary

    def _match_line_loading_metric(self, row: Dict[str, Any], summary: Dict[str, Dict[str, Any]]) -> Dict[str, Any] | None:
        if not summary:
            return None
        candidates: set[str] = set()
        for value in [
            row.get("id"),
            row.get("name"),
            row.get("line"),
            row.get("line_id"),
            row.get("dss_line_name"),
        ]:
            key = self._metric_key(value)
            if not key or key == "unnamed":
                continue
            candidates.add(key)
            if key.startswith("line_"):
                candidates.add(key[5:])
            else:
                candidates.add(f"line_{key}")
        for key in candidates:
            if key in summary:
                return summary[key]
        for metric_key, metric in summary.items():
            for key in candidates:
                if metric_key.endswith(f"_{key}") or key.endswith(f"_{metric_key}"):
                    return metric
        return None

    def _build_network_topology_view(
        self,
        project_id: str,
        latest_task: Dict[str, Any],
        selected_case: str,
        case_dir: Path,
        hourly_rows: List[Dict[str, Any]],
        annual_summary: Dict[str, Any],
        best_result_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        project = self._load_project_data(project_id)
        network = project.get("network") if isinstance(project.get("network"), dict) else {}
        nodes = network.get("nodes") if isinstance(network.get("nodes"), list) else []
        edges = network.get("edges") if isinstance(network.get("edges"), list) else []
        if not nodes:
            return {
                "nodes": [],
                "edges": [],
                "summary": [],
                "warnings": ["当前项目未保存可视化拓扑，无法展示配电网结构。"],
            }

        node_map = {str(node.get("id")): node for node in nodes if node.get("id") is not None}
        bus_to_node = {self._node_bus_name(node): str(node.get("id")) for node in nodes if node.get("id") is not None}
        selected_node_id = self._match_selected_storage_node(nodes, selected_case, best_result_summary)
        target_node = node_map.get(selected_node_id or "")
        bus_voltage_base_kv = {
            self._metric_key(self._node_bus_name(node)): self._node_voltage_kv(node)
            for node in nodes
            if node.get("id") is not None and self._node_voltage_kv(node) > 0
        }
        bus_voltage_summary, line_loading_summary = self._load_or_build_network_topology_trace_summaries(
            case_dir,
            bus_voltage_base_kv,
        )
        has_bus_trace = bool(bus_voltage_summary)
        has_line_trace = bool(line_loading_summary)
        data_quality = "opendss" if has_bus_trace and has_line_trace else "mixed" if has_bus_trace or has_line_trace else "estimated"

        peak_before = max([self._number(row, "actual_net_load_kw") for row in hourly_rows] or [0.0])
        peak_after = max([self._number(row, "grid_exchange_kw") for row in hourly_rows] or [0.0])
        peak_reduction = max(0.0, peak_before - peak_after)
        storage_power = self._number(best_result_summary, "rated_power_kw")
        storage_energy = self._number(best_result_summary, "rated_energy_kwh")
        transformer_limit = self._node_transformer_limit_kw(target_node) if target_node else None
        target_margin_before = None if transformer_limit is None else transformer_limit - peak_before
        remaining_supply = None if transformer_limit is None else transformer_limit - peak_after
        target_margin_delta = None if target_margin_before is None or remaining_supply is None else remaining_supply - target_margin_before

        line_rows = self._merge_topology_edges_with_line_summary(project_id, latest_task, edges, bus_to_node)
        downstream_load = self._estimate_downstream_load_kw(nodes, line_rows)
        estimated_voltage_pu = self._estimate_node_voltage_profile_pu(nodes, line_rows, downstream_load)

        node_payload = []
        for idx, node in enumerate(nodes):
            params = node.get("params") if isinstance(node.get("params"), dict) else {}
            position = node.get("position") if isinstance(node.get("position"), dict) else {}
            node_id = str(node.get("id"))
            is_target = node_id == selected_node_id
            bus_name = self._node_bus_name(node)
            voltage_metric = bus_voltage_summary.get(self._metric_key(bus_name), {})
            baseline_voltage_min = voltage_metric.get("baselineVoltagePuMin")
            baseline_voltage_max = voltage_metric.get("baselineVoltagePuMax")
            voltage_min = voltage_metric.get("voltagePuMin")
            voltage_max = voltage_metric.get("voltagePuMax")
            voltage_source = "opendss" if voltage_metric else "estimated"
            design_load_kw = self._node_design_load_kw(node)
            node_limit_kw = self._node_transformer_limit_kw(node)
            if node_limit_kw is not None:
                margin_before_kw = (target_margin_before if is_target else node_limit_kw - design_load_kw)
                margin_after_kw = (remaining_supply if is_target else margin_before_kw)
                margin_delta_kw = None if margin_before_kw is None or margin_after_kw is None else margin_after_kw - margin_before_kw
            else:
                margin_before_kw = None
                margin_after_kw = None
                margin_delta_kw = None
            if voltage_min is None and voltage_max is None:
                estimated_voltage = estimated_voltage_pu.get(node_id)
                if estimated_voltage is not None:
                    voltage_min = max(0.0, float(estimated_voltage) - 0.005)
                    voltage_max = min(1.2, float(estimated_voltage) + 0.005)
                else:
                    voltage_source = "none"
            voltage_mid = None
            if voltage_min is not None and voltage_max is not None:
                voltage_mid = (float(voltage_min) + float(voltage_max)) / 2.0
            node_payload.append(
                {
                    "id": node_id,
                    "name": str(node.get("name") or node_id),
                    "type": str(node.get("type") or ""),
                    "bus": bus_name,
                    "x": self._number(position, "x", 120.0 + idx * 80.0),
                    "y": self._number(position, "y", 120.0 + idx * 45.0),
                    "voltageLevelKv": self._node_voltage_kv(node),
                    "voltagePu": voltage_mid,
                    "baselineVoltagePuMin": baseline_voltage_min,
                    "baselineVoltagePuMax": baseline_voltage_max,
                    "voltagePuMin": voltage_min,
                    "voltagePuMax": voltage_max,
                    "storageVoltageViolationIncrementPu": voltage_metric.get("storageVoltageViolationIncrementPu"),
                    "voltageDataQuality": voltage_source,
                    "baselineUndervoltageHours": voltage_metric.get("baselineUndervoltageHours"),
                    "baselineOvervoltageHours": voltage_metric.get("baselineOvervoltageHours"),
                    "storageWorsenedVoltageHours": voltage_metric.get("storageWorsenedVoltageHours"),
                    "undervoltageHours": voltage_metric.get("undervoltageHours"),
                    "overvoltageHours": voltage_metric.get("overvoltageHours"),
                    "designLoadKw": design_load_kw,
                    "capacityLimitKw": node_limit_kw,
                    "capacityMarginBeforeKw": margin_before_kw,
                    "capacityMarginAfterKw": margin_after_kw,
                    "capacityMarginDeltaKw": margin_delta_kw,
                    "storageTarget": is_target,
                    "storagePowerKw": storage_power if is_target else None,
                    "storageEnergyKwh": storage_energy if is_target else None,
                    "remainingSupplyKw": remaining_supply if is_target else None,
                    "optimizeStorage": self._bool_value(params.get("optimize_storage")) is True,
                }
            )

        edge_payload = []
        for row in line_rows:
            from_id = str(row.get("from_node_id") or "")
            to_id = str(row.get("to_node_id") or "")
            if from_id not in node_map or to_id not in node_map:
                continue
            is_transformer_link = self._is_transformer_link(node_map.get(from_id), node_map.get(to_id))
            child_id = str(row.get("child_node_id") or to_id)
            load_kw = downstream_load.get(str(row.get("id") or ""), 0.0)
            phases = int(self._number(row, "phases", 3))
            normamps = self._number(row, "normamps")
            to_voltage = self._node_voltage_kv(node_map.get(child_id) or node_map[to_id])
            estimated_current = self._estimate_line_current_a(load_kw, to_voltage, phases)
            estimated_load_rate = (estimated_current / normamps * 100.0) if normamps > 0 else None
            line_metric = self._match_line_loading_metric(row, line_loading_summary)
            current_a = None if is_transformer_link else (line_metric.get("currentA") if line_metric else estimated_current)
            load_rate = None if is_transformer_link else (line_metric.get("loadingPct") if line_metric else estimated_load_rate)
            flow_direction = str(line_metric.get("flowDirection") if line_metric else "forward")
            edge_payload.append(
                {
                    "id": str(row.get("id") or ""),
                    "name": str(row.get("name") or row.get("id") or ""),
                    "from_node_id": from_id,
                    "to_node_id": to_id,
                    "fromBus": str(row.get("from_bus") or ""),
                    "toBus": str(row.get("to_bus") or ""),
                    "linecode": str(row.get("linecode") or ""),
                    "lengthKm": self._number(row, "length_km"),
                    "normamps": normamps,
                    "emergamps": self._number(row, "emergamps", normamps * 1.25),
                    "currentA": current_a,
                    "estimatedCurrentA": None if is_transformer_link else estimated_current,
                    "estimatedLoadRatePct": None if is_transformer_link else estimated_load_rate,
                    "loadRatePct": load_rate,
                    "loadingPct": load_rate,
                    "terminal1PowerKw": None if is_transformer_link or not line_metric else line_metric.get("terminal1PowerKw"),
                    "flowDirection": "reverse" if flow_direction == "reverse" else "forward",
                    "dataQuality": "transformer_link" if is_transformer_link else ("opendss" if line_metric else "estimated"),
                    "downstreamLoadKw": None if is_transformer_link else load_kw,
                    "isTransformerLink": is_transformer_link,
                    "enabled": bool(row.get("enabled", True)),
                    "normallyOpen": bool(row.get("normally_open", False)),
                    "autoServiceLine": bool(row.get("auto_service_line")),
                    "serviceSecondaryKv": self._number(row, "service_secondary_kv"),
                    "serviceTransformerKva": self._number(row, "service_transformer_kva"),
                    "serviceResourceKva": self._number(row, "service_resource_kva"),
                    "serviceTransformerCurrentA": self._number(row, "service_transformer_current_a"),
                    "serviceResourceCurrentA": self._number(row, "service_resource_current_a"),
                    "serviceEquivalentMode": str(row.get("service_equivalent_mode") or ""),
                    "serviceCableName": str(row.get("service_cable_name") or ""),
                    "serviceCableParallel": self._number(row, "service_cable_parallel"),
                    "serviceEquivalentR1OhmPerKm": self._number(row, "service_equivalent_r1_ohm_per_km"),
                    "serviceEquivalentX1OhmPerKm": self._number(row, "service_equivalent_x1_ohm_per_km"),
                    "lineVoltageKv": self._number(row, "line_voltage_kv"),
                    "downstreamTransformerKva": self._number(row, "downstream_transformer_kva"),
                    "downstreamLoadKva": self._number(row, "downstream_load_kva"),
                    "downstreamApparentKva": self._number(row, "downstream_apparent_kva"),
                    "estimatedRequiredCurrentA": self._number(row, "estimated_required_current_a"),
                    "recommendedCurrentA": self._number(row, "recommended_current_a"),
                    "recommendedLinecode": str(row.get("recommended_linecode") or ""),
                    "capacityCheckStatus": str(row.get("capacity_check_status") or ""),
                    "capacityCheckMessage": str(row.get("capacity_check_message") or ""),
                }
            )

        max_load_rate = max([self._number(row, "loadRatePct") for row in edge_payload if row.get("loadRatePct") is not None] or [0.0])
        overloaded_lines = sum(1 for row in edge_payload if self._number(row, "loadRatePct") > 100.0)
        storage_worsened_voltage_nodes = sum(
            1
            for row in node_payload
            if self._number(row, "storageVoltageViolationIncrementPu") > 1e-9
        )
        baseline_voltage_values_min = [
            float(row["baselineVoltagePuMin"])
            for row in node_payload
            if row.get("baselineVoltagePuMin") is not None
        ]
        baseline_voltage_values_max = [
            float(row["baselineVoltagePuMax"])
            for row in node_payload
            if row.get("baselineVoltagePuMax") is not None
        ]
        voltage_values_min = [float(row["voltagePuMin"]) for row in node_payload if row.get("voltagePuMin") is not None]
        voltage_values_max = [float(row["voltagePuMax"]) for row in node_payload if row.get("voltagePuMax") is not None]
        total_design_load = sum(self._node_design_load_kw(node) for node in nodes)
        voltage_source_label = "OpenDSS真实" if has_bus_trace else "拓扑估算"
        summary = [
            {"name": "优化前峰值负荷", "value": peak_before, "unit": "kW"},
            {"name": "优化后峰值并网功率", "value": peak_after, "unit": "kW"},
            {"name": "储能释放供电裕度", "value": peak_reduction, "unit": "kW"},
            {"name": "推荐储能功率", "value": storage_power, "unit": "kW"},
            {"name": "推荐储能容量", "value": storage_energy, "unit": "kWh"},
            {"name": "估算变压器供电上限", "value": transformer_limit, "unit": "kW"},
            {"name": "储前供电裕度", "value": target_margin_before, "unit": "kW"},
            {"name": "储后供电裕度", "value": remaining_supply, "unit": "kW"},
            {"name": "供电裕度变化", "value": target_margin_delta, "unit": "kW"},
            {"name": "拓扑静态负荷合计", "value": total_design_load, "unit": "kW"},
            {"name": "储前最低电压标幺值", "value": min(baseline_voltage_values_min) if baseline_voltage_values_min else None, "unit": "pu"},
            {"name": "储前最高电压标幺值", "value": max(baseline_voltage_values_max) if baseline_voltage_values_max else None, "unit": "pu"},
            {"name": "储后最低电压标幺值", "value": min(voltage_values_min) if voltage_values_min else None, "unit": "pu"},
            {"name": "储后最高电压标幺值", "value": max(voltage_values_max) if voltage_values_max else None, "unit": "pu"},
            {"name": "储能新增电压越限节点", "value": storage_worsened_voltage_nodes, "unit": "个"},
            {"name": "电压数据来源", "value": voltage_source_label, "unit": ""},
            {"name": "最大真实线路负载率" if has_line_trace else "最大估算线路负载率", "value": max_load_rate, "unit": "%"},
            {"name": "真实过载线路数" if has_line_trace else "估算过载线路数", "value": overloaded_lines, "unit": "条"},
            {"name": "变压器越限小时", "value": self._number(annual_summary, "transformer_violation_hours"), "unit": "h"},
            {"name": "最大变压器缺口", "value": self._number(annual_summary, "max_transformer_slack_kw"), "unit": "kW"},
        ]
        if data_quality == "opendss":
            warnings = ["已读取 OpenDSS 逐时电压/线路电流导出，拓扑颜色使用真实潮流结果。"]
        elif data_quality == "mixed":
            warnings = ["只读取到部分 OpenDSS 潮流导出，缺失项仍按可视化拓扑和线路容量估算。"]
        else:
            warnings = [
                "节点电压和线路电流当前按可视化拓扑、静态负荷和线路容量估算；启用 OpenDSS oracle 并重新求解后，页面会切换为真实潮流热力图。",
            ]
        if selected_node_id is None:
            warnings.append("未能把当前结果目录精确匹配到拓扑中的某个负荷节点，储能配置以全局推荐方案展示。")

        return {
            "nodes": node_payload,
            "edges": edge_payload,
            "summary": summary,
            "warnings": warnings,
            "selectedNodeId": selected_node_id,
            "dataQuality": data_quality,
        }

    def _load_project_data(self, project_id: str) -> Dict[str, Any]:
        if self.project_service is not None and hasattr(self.project_service, "load_project"):
            project = self.project_service.load_project(project_id)
            if isinstance(project, dict):
                return project
            if hasattr(project, "model_dump"):
                return project.model_dump(mode="json")
            if hasattr(project, "dict"):
                return project.dict()
        project_file = self._project_dir(project_id) / "project.json"
        if project_file.exists():
            data = self._safe_load_json(project_file, default={})
            return data if isinstance(data, dict) else {}
        return {}

    def _merge_topology_edges_with_line_summary(
        self,
        project_id: str,
        latest_task: Dict[str, Any],
        topology_edges: List[Dict[str, Any]],
        bus_to_node: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        line_summary = self._line_summary_from_manifest(project_id, latest_task)
        if not line_summary:
            line_summary = self._line_summary_from_dss_files(project_id, latest_task)
        by_id = {str(row.get("id") or ""): row for row in line_summary if row.get("id")}
        merged: List[Dict[str, Any]] = []
        seen: set[str] = set()

        for edge in topology_edges:
            edge_id = self._safe_name(str(edge.get("id") or "line"))
            params = edge.get("params") if isinstance(edge.get("params"), dict) else {}
            row = dict(by_id.get(edge_id) or by_id.get(str(edge.get("id") or "")) or {})
            defaults = {
                "id": edge_id,
                "name": str(edge.get("name") or edge_id),
                "from_node_id": str(edge.get("from_node_id") or ""),
                "to_node_id": str(edge.get("to_node_id") or ""),
                "linecode": str(params.get("linecode") or params.get("line_code") or row.get("linecode") or "LC_MAIN"),
                "length_km": self._number(params, "length_km", self._number(row, "length_km", 1.0)),
                "r_ohm_per_km": self._number(params, "r_ohm_per_km", self._number(row, "r_ohm_per_km", 0.0)),
                "x_ohm_per_km": self._number(params, "x_ohm_per_km", self._number(row, "x_ohm_per_km", 0.0)),
                "normamps": self._number(params, "rated_current_a", self._number(row, "normamps", 0.0)),
                "emergamps": self._number(params, "emerg_current_a", self._number(row, "emergamps", 0.0)),
                "phases": int(self._number(params, "phases", self._number(row, "phases", 3))),
                "enabled": self._bool_value(params.get("enabled")) is not False,
                "normally_open": self._bool_value(params.get("normally_open")) is True,
            }
            defaults.update(row)
            defaults["from_node_id"] = str(defaults.get("from_node_id") or edge.get("from_node_id") or "")
            defaults["to_node_id"] = str(defaults.get("to_node_id") or edge.get("to_node_id") or "")
            merged.append(defaults)
            seen.add(str(defaults.get("id") or edge_id))

        for row in line_summary:
            line_id = str(row.get("id") or "")
            if not line_id or line_id in seen:
                continue
            fallback = dict(row)
            fallback["from_node_id"] = str(row.get("from_node_id") or bus_to_node.get(str(row.get("from_bus") or row.get("bus1") or ""), ""))
            fallback["to_node_id"] = str(row.get("to_node_id") or bus_to_node.get(str(row.get("to_bus") or row.get("bus2") or ""), ""))
            merged.append(fallback)

        return merged

    def _estimate_downstream_load_kw(self, nodes: List[Dict[str, Any]], edges: List[Dict[str, Any]]) -> Dict[str, float]:
        node_ids = {str(node.get("id")) for node in nodes if node.get("id") is not None}
        roots = {
            str(node.get("id"))
            for node in nodes
            if str(node.get("type")).strip().lower() in {"grid", "source", "transformer"} and node.get("id") is not None
        }
        adjacency: Dict[str, List[tuple[str, str]]] = {node_id: [] for node_id in node_ids}
        active_edges = []
        for edge in edges:
            if self._bool_value(edge.get("enabled")) is False or self._bool_value(edge.get("normally_open")) is True:
                continue
            from_id = str(edge.get("from_node_id") or "")
            to_id = str(edge.get("to_node_id") or "")
            if from_id not in node_ids or to_id not in node_ids:
                continue
            edge_id = str(edge.get("id") or "")
            active_edges.append(edge)
            adjacency[from_id].append((to_id, edge_id))
            adjacency[to_id].append((from_id, edge_id))

        parent: Dict[str, str | None] = {root: None for root in roots if root in node_ids}
        parent_edge: Dict[str, str] = {}
        queue = list(parent.keys()) or list(node_ids)[:1]
        for root in queue:
            parent.setdefault(root, None)
        while queue:
            current = queue.pop(0)
            for neighbor, edge_id in adjacency.get(current, []):
                if neighbor in parent:
                    continue
                parent[neighbor] = current
                parent_edge[neighbor] = edge_id
                queue.append(neighbor)

        children: Dict[str, List[str]] = {node_id: [] for node_id in node_ids}
        for node_id, parent_id in parent.items():
            if parent_id is not None:
                children.setdefault(parent_id, []).append(node_id)

        load_by_node = {str(node.get("id")): self._node_design_load_kw(node) for node in nodes if node.get("id") is not None}
        subtree_cache: Dict[str, float] = {}

        def subtree_load(node_id: str) -> float:
            if node_id in subtree_cache:
                return subtree_cache[node_id]
            total = load_by_node.get(node_id, 0.0)
            for child in children.get(node_id, []):
                total += subtree_load(child)
            subtree_cache[node_id] = total
            return total

        result: Dict[str, float] = {}
        for edge in active_edges:
            edge_id = str(edge.get("id") or "")
            from_id = str(edge.get("from_node_id") or "")
            to_id = str(edge.get("to_node_id") or "")
            if parent.get(to_id) == from_id:
                child = to_id
            elif parent.get(from_id) == to_id:
                child = from_id
            else:
                child = to_id
            edge["child_node_id"] = child
            result[edge_id] = subtree_load(child)
        return result

    def _estimate_node_voltage_profile_pu(
        self,
        nodes: List[Dict[str, Any]],
        edges: List[Dict[str, Any]],
        downstream_load: Dict[str, float],
    ) -> Dict[str, float]:
        node_ids = {str(node.get("id")) for node in nodes if node.get("id") is not None}
        if not node_ids:
            return {}

        roots = [
            str(node.get("id"))
            for node in nodes
            if str(node.get("type")).strip().lower() in {"grid", "source", "transformer"} and node.get("id") is not None
        ]
        if not roots:
            roots = [next(iter(node_ids))]

        adjacency: Dict[str, List[tuple[str, Dict[str, Any]]]] = {node_id: [] for node_id in node_ids}
        for edge in edges:
            if self._bool_value(edge.get("enabled")) is False or self._bool_value(edge.get("normally_open")) is True:
                continue
            from_id = str(edge.get("from_node_id") or "")
            to_id = str(edge.get("to_node_id") or "")
            if from_id not in node_ids or to_id not in node_ids:
                continue
            adjacency[from_id].append((to_id, edge))
            adjacency[to_id].append((from_id, edge))

        node_map = {str(node.get("id")): node for node in nodes if node.get("id") is not None}
        voltage: Dict[str, float] = {root: 1.0 for root in roots if root in node_ids}
        queue = [root for root in roots if root in node_ids]

        while queue:
            current = queue.pop(0)
            parent_voltage = voltage.get(current, 1.0)
            for neighbor, edge in adjacency.get(current, []):
                if neighbor in voltage:
                    continue
                edge_id = str(edge.get("id") or "")
                child_id = str(edge.get("child_node_id") or neighbor)
                load_kw = downstream_load.get(edge_id, 0.0)
                child_node = node_map.get(child_id) or node_map.get(neighbor) or {}
                voltage_kv = self._node_voltage_kv(child_node)
                phases = int(self._number(edge, "phases", 3))
                drop = self._estimate_voltage_drop_pu(edge, load_kw, voltage_kv, phases)
                voltage[neighbor] = max(0.88, min(1.08, parent_voltage - drop))
                queue.append(neighbor)

        for node_id in node_ids:
            voltage.setdefault(node_id, 1.0)
        return voltage

    @staticmethod
    def _is_transformer_link(node_a: Dict[str, Any] | None, node_b: Dict[str, Any] | None) -> bool:
        if not node_a or not node_b:
            return False
        types = {str(node_a.get("type") or "").strip().lower(), str(node_b.get("type") or "").strip().lower()}
        return bool(types & {"grid", "source"}) and "transformer" in types

    def _estimate_voltage_drop_pu(self, edge: Dict[str, Any], load_kw: float, voltage_kv: float, phases: int) -> float:
        if load_kw <= 0 or voltage_kv <= 0:
            return 0.0
        length_km = max(0.0, self._number(edge, "length_km", 0.0))
        r_ohm = max(0.0, self._number(edge, "r_ohm_per_km", 0.0)) * length_km
        x_ohm = max(0.0, self._number(edge, "x_ohm_per_km", 0.0)) * length_km
        if r_ohm <= 0 and x_ohm <= 0:
            return 0.0
        pf = 0.95
        q_kvar = load_kw * math.tan(math.acos(pf))
        voltage_ll_kv = self._estimate_line_line_voltage_kv(voltage_kv)
        denominator = max(voltage_ll_kv * voltage_ll_kv * 1000.0, 1.0)
        drop = (r_ohm * load_kw + x_ohm * q_kvar) / denominator
        return max(0.0, min(0.08, drop))

    def _match_selected_storage_node(
        self,
        nodes: List[Dict[str, Any]],
        selected_case: str,
        best_result_summary: Dict[str, Any],
    ) -> str | None:
        candidates = [
            selected_case,
            str(best_result_summary.get("internal_model_id") or ""),
            str(best_result_summary.get("model_id") or ""),
            str(best_result_summary.get("load_id") or ""),
        ]
        normalized = [self._safe_name(item).lower() for item in candidates if item]
        for node in nodes:
            node_id = str(node.get("id") or "")
            params = node.get("params") if isinstance(node.get("params"), dict) else {}
            values = [
                node_id,
                str(node.get("name") or ""),
                self._safe_name(node_id),
                str(params.get("dss_load_name") or ""),
                str(params.get("node_id") or ""),
            ]
            node_id_num = str(params.get("node_id") or "").strip()
            if node_id_num:
                values.extend([f"node{int(float(node_id_num)):02d}", f"load_{int(float(node_id_num)):02d}"])
            haystack = " ".join(self._safe_name(value).lower() for value in values if value)
            if any(item and item in haystack for item in normalized):
                return node_id
        for node in nodes:
            params = node.get("params") if isinstance(node.get("params"), dict) else {}
            if self._bool_value(params.get("optimize_storage")) is True:
                return str(node.get("id") or "")
        return None

    def _node_bus_name(self, node: Dict[str, Any]) -> str:
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        explicit = str(params.get("dss_bus_name") or params.get("bus_name") or "").strip()
        if explicit:
            return self._safe_name(explicit)
        node_type = str(node.get("type") or "").strip().lower()
        if node_type in {"grid", "source"}:
            return self._safe_name(str(params.get("source_bus") or "sourcebus"))
        if node_type == "transformer":
            return self._safe_name(str(params.get("secondary_bus_name") or "n0"))
        node_id = params.get("node_id")
        if node_type == "load" and node_id not in (None, ""):
            return self._safe_name(f"n{int(float(node_id))}")
        return self._safe_name(str(node.get("id") or "node"))

    def _node_voltage_kv(self, node: Dict[str, Any]) -> float:
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        node_type = str(node.get("type") or "").strip().lower()
        if node_type in {"grid", "source"}:
            return self._number(params, "base_kv", 110.0)
        if node_type == "load":
            return self._number(params, "target_kv_ln", 10.0)
        return self._number(params, "voltage_level_kv", 10.0)

    def _node_design_load_kw(self, node: Dict[str, Any]) -> float:
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        if str(node.get("type") or "").strip().lower() != "load":
            return 0.0
        return self._number(params, "design_kw")

    def _node_transformer_limit_kw(self, node: Dict[str, Any] | None) -> float | None:
        if not node:
            return None
        params = node.get("params") if isinstance(node.get("params"), dict) else {}
        kva = self._number(params, "transformer_capacity_kva")
        if kva <= 0:
            return None
        pf = self._number(params, "transformer_pf_limit", self._number(params, "pf", 0.95))
        reserve = self._number(params, "transformer_reserve_ratio", 0.15)
        return max(0.0, kva * pf * (1.0 - reserve))

    def _estimate_line_current_a(self, load_kw: float, voltage_kv: float, phases: int) -> float:
        if load_kw <= 0 or voltage_kv <= 0:
            return 0.0
        voltage_ll_kv = self._estimate_line_line_voltage_kv(voltage_kv)
        pf = 0.95
        return load_kw / max(math.sqrt(3.0) * voltage_ll_kv * pf, 1e-9)

    @staticmethod
    def _estimate_line_line_voltage_kv(voltage_kv: float) -> float:
        if voltage_kv <= 0:
            return 0.0
        # 5.7735 kV 等字段是 10 kV 系统的相电压；线路载流量按线电压口径估算。
        if voltage_kv <= 7.0:
            return voltage_kv * math.sqrt(3.0)
        return voltage_kv

    def _safe_name(self, value: str) -> str:
        chars = []
        for ch in str(value).strip():
            if ch.isalnum() or ch in {"_", "-"}:
                chars.append(ch)
            else:
                chars.append("_")
        return "".join(chars).strip("_") or "unnamed"

    def _add_moving_average(self, rows: List[Dict[str, Any]], source_key: str, target_key: str, window: int = 7) -> None:
        if not rows:
            return
        half_window = window // 2
        values = [self._number(row, source_key) for row in rows]
        for idx, row in enumerate(rows):
            start = max(0, idx - half_window)
            end = min(len(values), idx + half_window + 1)
            subset = values[start:end]
            row[target_key] = sum(subset) / max(len(subset), 1)

    def _build_cashflow_chart(self, rows: List[Dict[str, Any]], financial_summary: Dict[str, Any]) -> List[Dict[str, Any]]:
        cumulative = -self._number(financial_summary, "initial_investment_yuan") / 10000.0
        chart: List[Dict[str, Any]] = []
        for row in rows:
            discounted = self._number(row, "discounted_net_cashflow_yuan") / 10000.0
            cumulative += discounted
            chart.append(
                {
                    "year": int(self._number(row, "year")),
                    "revenueWan": self._number(row, "revenue_yuan") / 10000.0,
                    "operatingRevenueWan": self._number(row, "operating_revenue_yuan") / 10000.0,
                    "arbitrageRevenueWan": self._number(row, "arbitrage_revenue_yuan") / 10000.0,
                    "demandSavingWan": self._number(row, "demand_saving_yuan") / 10000.0,
                    "auxiliaryServiceRevenueWan": self._number(row, "auxiliary_service_revenue_yuan") / 10000.0,
                    "capacityRevenueWan": self._number(row, "capacity_revenue_yuan") / 10000.0,
                    "lossReductionRevenueWan": self._number(row, "loss_reduction_revenue_yuan") / 10000.0,
                    "operatingCostWan": -self._number(row, "operating_cost_yuan") / 10000.0,
                    "degradationCostWan": -self._number(row, "degradation_cost_yuan") / 10000.0,
                    "omCostWan": -self._number(row, "om_cost_yuan") / 10000.0,
                    "replacementCostWan": -self._number(row, "replacement_cost_yuan") / 10000.0,
                    "salvageValueWan": self._number(row, "salvage_value_yuan") / 10000.0,
                    "netCashflowWan": self._number(row, "net_cashflow_yuan") / 10000.0,
                    "discountedNetCashflowWan": discounted,
                    "cumulativeDiscountedWan": cumulative,
                }
            )
        return chart

    def _build_capital_breakdown(self, row: Dict[str, Any]) -> List[Dict[str, Any]]:
        items = [
            ("容量侧投资", "energy_capex_yuan"),
            ("功率侧投资", "power_capex_yuan"),
            ("安全附加", "safety_markup_yuan"),
            ("集成附加", "integration_markup_yuan"),
            ("其他投资", "other_capex_yuan"),
        ]
        return [{"name": name, "valueWan": self._number(row, key) / 10000.0} for name, key in items]

    def _build_annual_value_breakdown(self, row: Dict[str, Any]) -> List[Dict[str, Any]]:
        service_net = (
            self._number(row, "annual_auxiliary_service_revenue_yuan")
            or (
                self._number(row, "annual_service_capacity_revenue_yuan")
                + self._number(row, "annual_service_delivery_revenue_yuan")
                - self._number(row, "annual_service_penalty_yuan")
            )
        )
        grid_penalty = self._number(row, "annual_transformer_penalty_yuan") + self._number(row, "annual_voltage_penalty_yuan")
        items = [
            ("套利收益", self._number(row, "annual_arbitrage_revenue_yuan")),
            ("需量收益", self._number(row, "annual_demand_saving_yuan")),
            ("辅助服务收益", service_net),
            ("容量收益", self._number(row, "annual_capacity_revenue_yuan")),
            ("降损收益", self._number(row, "annual_loss_reduction_revenue_yuan")),
            ("政府补贴", self._number(row, "government_subsidy_yuan")),
            ("退化成本", -self._number(row, "annual_degradation_cost_yuan")),
            ("运维成本", -self._number(row, "annual_om_cost_yuan")),
            ("更换成本年化", -self._number(row, "annual_replacement_equivalent_cost_yuan")),
            ("网侧罚金", -grid_penalty),
        ]
        return [{"name": name, "valueWan": value / 10000.0} for name, value in items]

    def _build_financial_metrics(self, financial_row: Dict[str, Any], annual_row: Dict[str, Any]) -> List[Dict[str, Any]]:
        return [
            {"name": "NPV", "value": self._number(financial_row, "npv_yuan") / 10000.0, "unit": "万元"},
            {"name": "IRR", "value": self._number(financial_row, "irr") * 100.0, "unit": "%"},
            {"name": "静态回收期", "value": self._optional_number(financial_row, "simple_payback_years"), "unit": "年"},
            {"name": "折现回收期", "value": self._optional_number(financial_row, "discounted_payback_years"), "unit": "年"},
            {"name": "初始投资", "value": self._number(financial_row, "initial_investment_yuan") / 10000.0, "unit": "万元"},
            {"name": "补贴后投资", "value": self._number(financial_row, "initial_net_investment_yuan") / 10000.0, "unit": "万元"},
            {"name": "政府补贴", "value": self._number(financial_row, "government_subsidy_yuan") / 10000.0, "unit": "万元"},
            {"name": "年降损收益", "value": self._number(financial_row, "annual_loss_reduction_revenue_yuan") / 10000.0, "unit": "万元"},
            {"name": "年容量收益", "value": self._number(financial_row, "annual_capacity_revenue_yuan") / 10000.0, "unit": "万元"},
            {"name": "年更换折算", "value": self._number(financial_row, "annual_replacement_equivalent_cost_yuan") / 10000.0, "unit": "万元"},
            {"name": "额定功率", "value": self._number(financial_row, "rated_power_kw"), "unit": "kW"},
            {"name": "额定容量", "value": self._number(financial_row, "rated_energy_kwh"), "unit": "kWh"},
            {"name": "年等效循环", "value": self._number(annual_row or financial_row, "annual_equivalent_full_cycles"), "unit": "次"},
        ]

    def _build_pareto_chart(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chart: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows):
            chart.append(
                {
                    "index": idx + 1,
                    "strategyId": row.get("strategy_id"),
                    "ratedPowerKw": self._number(row, "rated_power_kw"),
                    "ratedEnergyKwh": self._number(row, "rated_energy_kwh"),
                    "durationH": self._number(row, "duration_h"),
                    "npvWan": self._number(row, "npv_yuan") / 10000.0,
                    "initialInvestmentWan": self._number(row, "initial_investment_yuan") / 10000.0,
                    "paybackYears": self._optional_number(row, "simple_payback_years"),
                    "annualCycles": self._number(row, "annual_equivalent_full_cycles"),
                    "feasible": bool(row.get("feasible")),
                    "totalViolation": self._number(row, "total_violation"),
                }
            )
        return chart

    def _build_history_chart(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chart: List[Dict[str, Any]] = []
        for row in rows:
            chart.append(
                {
                    "generation": int(self._number(row, "generation")),
                    "populationSize": int(self._number(row, "population_size")),
                    "feasibleCount": int(self._number(row, "feasible_count")),
                    "archiveSize": int(self._number(row, "archive_size")),
                    "bestNpvWan": self._optional_number(row, "best_npv_yuan") / 10000.0
                    if self._optional_number(row, "best_npv_yuan") is not None
                    else None,
                    "avgNpvWan": self._optional_number(row, "avg_npv_yuan") / 10000.0
                    if self._optional_number(row, "avg_npv_yuan") is not None
                    else None,
                    "bestPaybackYears": self._optional_number(row, "best_payback_years"),
                }
            )
        return chart
