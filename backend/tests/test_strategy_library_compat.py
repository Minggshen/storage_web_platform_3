from __future__ import annotations

import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

_BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
if _BACKEND_DIR not in sys.path:
    sys.path.insert(0, _BACKEND_DIR)

from services.asset_binding_service import AssetBindingService  # noqa: E402
from services.build_export_service import BuildExportService  # noqa: E402
from services.strategy_service import validate_strategy_library  # noqa: E402


def _v2_device_row() -> dict[str, object]:
    return {
        "enabled": 1,
        "manufacturer": "测试厂商",
        "device_model": "SafeBox-100",
        "rated_power_kw": 100,
        "rated_energy_kwh": 215,
        "duration_hour": 2,
        "battery_chemistry": "LFP_314Ah",
        "cooling_class": "pack_liquid_power_air",
        "cooling_note": "PACK液冷，PCS风冷",
        "ip_system": "IP55",
        "ip_pack": "IP67",
        "ip_pcs": "IP66",
        "corrosion_grade": "C3",
        "corrosion_optional_grade": "C5",
        "manual_safety_grade": "A",
        "round_trip_efficiency": 0.88,
        "c_rate_charge_max": 0.5,
        "c_rate_discharge_max": 0.5,
        "energy_unit_price_yuan_per_kwh": 760,
        "power_related_capex_yuan_per_kw": 260,
        "annual_om_ratio": 0.02,
        "soc_min": 0.1,
        "soc_max": 0.9,
        "operating_temp_min_c": -20,
        "operating_temp_max_c": 55,
        "cycle_life": 8000,
        "fire_detection_class": "composite_detection",
        "fire_suppression_class": "aerosol",
        "explosion_protection_class": "undisclosed",
        "propagation_protection_class": "pack_level",
        "ems_model": "EMS300CP",
        "certification_tokens": "GB_T_36276;IEC_62619",
        "weight_kg": 2600,
        "dimensions_mm": "1500x1300x2300",
        "is_default_candidate": 1,
        "ems_package_name": "EMS-A",
    }


def _write_v2_strategy_library(path: Path) -> None:
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame(
            [
                {"key": "schema_version", "value": "device_library_v2", "说明": "schema"},
                {"key": "template_version", "value": "2026.06", "说明": "template"},
            ]
        ).to_excel(writer, sheet_name="元数据", index=False)
        pd.DataFrame([_v2_device_row()]).to_excel(writer, sheet_name="设备库", index=False)
        pd.DataFrame(
            [
                {"dimension": "thermal", "label_zh": "热管理", "default_weight": 0.5},
                {"dimension": "certification", "label_zh": "认证", "default_weight": 0.5},
            ]
        ).to_excel(writer, sheet_name="安全权重", index=False)
        pd.DataFrame(
            [
                {
                    "dimension": "thermal",
                    "source_column": "cooling_class",
                    "pattern": "^pack_liquid_power_air$",
                    "score": 85,
                    "priority": 7,
                    "note": "PACK液冷",
                },
                {
                    "dimension": "thermal",
                    "source_column": "cooling_class",
                    "pattern": "",
                    "score": 60,
                    "priority": 0,
                    "note": "冷却默认",
                },
                {
                    "dimension": "certification",
                    "source_column": "certification_tokens",
                    "pattern": "GB_T_36276",
                    "score": 20,
                    "priority": 10,
                    "note": "GB",
                },
                {
                    "dimension": "certification",
                    "source_column": "certification_tokens",
                    "pattern": "IEC_62619",
                    "score": 25,
                    "priority": 10,
                    "note": "IEC",
                },
                {
                    "dimension": "certification",
                    "source_column": "certification_tokens",
                    "pattern": "",
                    "score": 0,
                    "priority": 0,
                    "note": "未披露",
                },
            ]
        ).to_excel(writer, sheet_name="安全评分规则", index=False)

    wb = load_workbook(path)
    wb.active = wb.sheetnames.index("安全评分规则")
    wb.save(path)


def test_device_library_validation_reads_v2_device_sheet_when_active_sheet_differs(tmp_path) -> None:
    path = tmp_path / "strategy.xlsx"
    _write_v2_strategy_library(path)

    report, records = AssetBindingService(project_service=object()).validate_device_library_file(path)

    assert report.ok is True
    assert len(records) == 1
    assert records[0].vendor == "测试厂商"
    assert records[0].model == "SafeBox-100"
    assert records[0].cooling_type == "pack_liquid_power_air"
    assert records[0].ip_system == "IP55"
    assert records[0].corrosion_grade == "C3"
    assert records[0].efficiency_pct == 88.0


def test_strategy_library_validation_accepts_v2_device_sheet(tmp_path) -> None:
    path = tmp_path / "strategy.xlsx"
    _write_v2_strategy_library(path)

    checks, summary = validate_strategy_library(path)

    assert all(check["ok"] for check in checks)
    assert summary["device_rows"] == 1
    assert summary["enabled_device_rows"] == 1
    assert summary["ems_rows"] == 0
    assert summary["default_ems_package"] == "EMS-A"
    assert summary["safety_sheets_present"] is True


def test_strategy_library_validation_rejects_non_v2_workbook(tmp_path) -> None:
    path = tmp_path / "legacy.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"enabled": 1, "manufacturer": "旧厂商", "device_model": "Old"}]).to_excel(
            writer,
            sheet_name="设备库",
            index=False,
        )

    checks, summary = validate_strategy_library(path)

    assert any(not check["ok"] and check["name"] == "策略库格式" for check in checks)
    assert summary["device_rows"] == 1


def test_solver_workspace_prefers_original_v2_strategy_library(tmp_path) -> None:
    source = tmp_path / "uploaded_strategy.xlsx"
    storage_dir = tmp_path / "solver_workspace" / "inputs" / "storage"
    storage_dir.mkdir(parents=True)
    _write_v2_strategy_library(source)

    project = {
        "device_library": {
            "asset": {
                "metadata": {
                    "stored_path": str(source),
                },
            },
            "records": [
                {
                    "enabled": True,
                    "vendor": "将被忽略",
                    "model": "OldRecord",
                    "rated_power_kw": 1,
                    "rated_energy_kwh": 2,
                }
            ],
        }
    }
    warnings: list[str] = []
    errors: list[str] = []

    rel_path, target = BuildExportService(data_root=tmp_path)._prepare_strategy_library(
        project,
        storage_dir,
        errors,
        warnings,
    )

    assert rel_path == "inputs/storage/工商业储能设备策略库.xlsx"
    assert target is not None
    assert not errors
    assert "设备库" in pd.ExcelFile(target).sheet_names
    assert "安全权重" in pd.ExcelFile(target).sheet_names
    assert "安全评分规则" in pd.ExcelFile(target).sheet_names
    copied = pd.read_excel(target, sheet_name="设备库")
    assert copied.loc[0, "manufacturer"] == "测试厂商"


def test_solver_workspace_rejects_cached_records_without_v2_source_file(tmp_path) -> None:
    storage_dir = tmp_path / "solver_workspace" / "inputs" / "storage"
    storage_dir.mkdir(parents=True)
    project = {
        "device_library": {
            "records": [
                {
                    "enabled": True,
                    "vendor": "缓存厂商",
                    "model": "CachedOnly",
                    "rated_power_kw": 1,
                    "rated_energy_kwh": 2,
                }
            ],
        }
    }
    warnings: list[str] = []
    errors: list[str] = []

    rel_path, target = BuildExportService(data_root=tmp_path)._prepare_strategy_library(
        project,
        storage_dir,
        errors,
        warnings,
    )

    assert rel_path is None
    assert target is None
    assert errors == ["未绑定 device_library_v2 设备策略库，无法生成求解器 strategy-library 输入。"]
